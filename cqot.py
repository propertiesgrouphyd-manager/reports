# ==============================
# ULTRA FAST ASYNC MULTI PROPERTY AUTOMATION
# HOURLY CASH + QR + ONLINE + DISCOUNT + TOTAL REPORT
# ==============================

import os
import json
import asyncio
import aiohttp
from datetime import datetime, timedelta
import traceback
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from io import BytesIO
import pytz


IST = pytz.timezone("Asia/Kolkata")
now = datetime.now(IST)

MAX_FULL_RUN_RETRIES = 5
FULL_RUN_RETRY_DELAY = 10

PROP_PARALLEL_LIMIT = 3
DETAIL_PARALLEL_LIMIT = 10

prop_semaphore = asyncio.Semaphore(PROP_PARALLEL_LIMIT)

DETAIL_TIMEOUT = 25
BATCH_TIMEOUT = 35


# ================= TELEGRAM =================

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
if not TELEGRAM_BOT_TOKEN:
    raise RuntimeError("❌ TELEGRAM_BOT_TOKEN missing")

CHAT_MAP = json.loads(os.getenv("TELEGRAM_CHAT_MAP", "{}"))

def get_chat_id(name: str):
    if name not in CHAT_MAP:
        raise RuntimeError(f"❌ Chat ID not configured: {name}")
    return int(CHAT_MAP[name])

TELEGRAM_CHAT_ID = get_chat_id("6am")


async def send_telegram_excel_buffer(buffer, filename, caption=None):

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"

    data = aiohttp.FormData()
    data.add_field("chat_id", str(TELEGRAM_CHAT_ID))

    if caption:
        data.add_field("caption", caption)

    data.add_field(
        "document",
        buffer,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    async with aiohttp.ClientSession() as session:
        async with session.post(url, data=data, timeout=120) as resp:
            if resp.status != 200:
                text = await resp.text()
                raise RuntimeError(f"Telegram send failed: {text}")


# ================= PROPERTIES =================

PROPERTIES_RAW = json.loads(os.getenv("OYO_PROPERTIES", "{}"))
PROPERTIES = {int(k): v for k, v in PROPERTIES_RAW.items()}

if not PROPERTIES:
    raise RuntimeError("❌ OYO_PROPERTIES secret missing or empty")


# ================= COLOR =================

def get_hour_color(hour):

    palette = [
        "EEF3FB","E8F0FA","E3EDFA","DEEAFA",
        "D9F2FF","DFF7FF","E6FBFF","FFF9DB",
        "FFF4CC","FFEFB3","FFE699","FFDD80",
        "FFE0CC","FFD6B3","FFCC99","FFC280",
        "FFD9D9","FFD1D1","FFC9C9","FFC1C1",
        "F3E5F5","EDE7F6","E8EAF6","E3F2FD"
    ]

    return palette[hour % 24]


# ================= FETCH DETAILS =================

async def fetch_booking_details(session, P, booking_no):

    url = "https://www.oyoos.com/hms_ms/api/v1/visibility/booking_details_with_entities"

    params = {
        "qid": P["QID"],
        "booking_id": booking_no,
        "role": 0,
        "platform": "OYOOS",
        "country_code": 1
    }

    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}

    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    for attempt in range(1, 4):

        try:
            async with session.get(
                url,
                params=params,
                headers=headers,
                cookies=cookies,
                timeout=DETAIL_TIMEOUT
            ) as r:

                if r.status != 200:
                    raise RuntimeError("DETAIL API FAILED")

                data = await r.json()

                booking = next(
                    iter(data.get("entities", {}).get("bookings", {}).values()),
                    {}
                )

                payments = booking.get("payments", [])

                events = []

                for p in payments:

                    amt = float(p.get("amount", 0) or 0)
                    if amt <= 0:
                        continue

                    created_at = str(p.get("created_at") or "").strip()
                    if not created_at:
                        continue

                    try:
                        dt = datetime.fromisoformat(created_at.replace("Z", ""))
                        dt = dt.astimezone(IST)
                    except Exception:
                        continue

                    mode_raw = p.get("mode", "")

                    if mode_raw == "Cash at Hotel":
                        bucket = "cash"
                    elif mode_raw == "UPI QR":
                        bucket = "qr"
                    elif mode_raw == "oyo_wizard_discount":
                        bucket = "discount"
                    else:
                        bucket = "online"

                    events.append({
                        "date": dt.strftime("%Y-%m-%d"),
                        "hour": dt.hour,
                        "mode": bucket,
                        "amt": amt
                    })

                return events

        except Exception:
            await asyncio.sleep(2 + attempt)

    raise RuntimeError("DETAIL FETCH FAILED")


# ================= FETCH BOOKINGS =================

async def fetch_bookings_batch(session, offset, f, t, P):

    url = "https://www.oyoos.com/hms_ms/api/v1/get_booking_with_ids"

    params = {
        "qid": P["QID"],
        "checkin_from": f,
        "checkin_till": t,
        "batch_count": 100,
        "batch_offset": offset,
        "visibility_required": "true",
        "additionalParams": "payment_hold_transaction,guest,stay_details",
        "decimal_price": "true",
        "ascending": "true",
        "sort_on": "checkin_date"
    }

    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}

    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    async with session.get(
        url,
        params=params,
        cookies=cookies,
        headers=headers,
        timeout=BATCH_TIMEOUT
    ) as r:

        if r.status != 200:
            raise RuntimeError("BATCH API FAILED")

        return await r.json()


# ================= PROCESS PROPERTY =================

async def process_property(P, TF, TT, HF, HT):

    print(f"PROCESSING → {P['name']}")

    tf_dt = datetime.strptime(TF, "%Y-%m-%d").date()
    tt_dt = datetime.strptime(TT, "%Y-%m-%d").date()

    async with aiohttp.ClientSession() as session:

        detail_semaphore = asyncio.Semaphore(DETAIL_PARALLEL_LIMIT)
        detail_cache = {}

        async def limited_detail_call(booking_no):

            async with detail_semaphore:

                if booking_no in detail_cache:
                    return detail_cache[booking_no]

                res = await fetch_booking_details(session, P, booking_no)
                detail_cache[booking_no] = res

                return res


        hourly_cash = {
            h: {"cash":0.0,"qr":0.0,"online":0.0,"discount":0.0,"total":0.0}
            for h in range(24)
        }

        offset = 0

        while True:

            data = await fetch_bookings_batch(session, offset, HF, HT, P)

            if not data or not data.get("bookingIds"):
                break

            bookings = data.get("entities", {}).get("bookings", {})
            if not bookings:
                break

            tasks = []

            for b in bookings.values():

                status = (b.get("status") or "").strip()

                if status not in ["Checked In", "Checked Out"]:
                    continue

                booking_no = b.get("booking_no")

                if not booking_no:
                    continue

                tasks.append(limited_detail_call(booking_no))

            results = await asyncio.gather(*tasks, return_exceptions=True)

            for res in results:

                if isinstance(res, Exception):
                    continue

                for ev in res or []:

                    try:
                        d_dt = datetime.strptime(ev["date"], "%Y-%m-%d").date()
                    except Exception:
                        continue

                    if not (tf_dt <= d_dt <= tt_dt):
                        continue

                    h = ev["hour"]
                    amt = float(ev["amt"])
                    mode = ev["mode"]

                    hourly_cash[h][mode] += amt
                    hourly_cash[h]["total"] += amt

            if len(data.get("bookingIds", [])) < 100:
                break

            offset += 100

        return (P["name"], hourly_cash)


# ================= RETRY =================

async def run_property_with_retry(P, TF, TT, HF, HT, retries=3):

    last_error = None

    for attempt in range(1, retries + 1):

        try:
            return await process_property(P, TF, TT, HF, HT)

        except Exception as e:

            last_error = e

            print(f"RETRY {attempt}/{retries} → {P['name']} :: {e}")

            await asyncio.sleep(2 + attempt * 2)

    raise RuntimeError(f"PROPERTY FAILED → {P['name']}") from last_error


async def run_property_limited(P, TF, TT, HF, HT):

    async with prop_semaphore:

        return await run_property_with_retry(P, TF, TT, HF, HT)



# ================= AUTOFIT =================

def autofit_columns(ws):

    for column_cells in ws.columns:

        max_length = 0
        col_letter = column_cells[0].column_letter

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted = (max_length * 1.4) + 4

        # Minimum sizes for important columns
        if col_letter == "A":
            adjusted = max(adjusted, 16)

        if col_letter == "B":
            adjusted = max(adjusted, 22)

        ws.column_dimensions[col_letter].width = min(adjusted, 45)

# ================= MAIN =================

async def main():

    print("========================================")
    print(" HOURLY COLLECTION REPORT")
    print("========================================")

    global now
    now = datetime.now(IST)

    target_date = (now - timedelta(days=1)).date()

    TF = target_date.strftime("%Y-%m-%d")
    TT = TF

    HF = (target_date - timedelta(days=30)).strftime("%Y-%m-%d")
    HT = TF

    display_date = datetime.strptime(TF, "%Y-%m-%d").strftime("%d-%m-%Y")

    pending = {k: v for k, v in PROPERTIES.items()}
    success_results = {}

    for run_attempt in range(1, MAX_FULL_RUN_RETRIES + 1):

        if not pending:
            break

        tasks = [run_property_limited(P, TF, TT, HF, HT) for P in pending.values()]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        new_pending = {}

        for key, (P, result) in zip(list(pending.keys()), zip(pending.values(), results)):

            if isinstance(result, Exception):
                new_pending[key] = P
                continue

            success_results[key] = result

        pending = new_pending

        if pending:
            await asyncio.sleep(FULL_RUN_RETRY_DELAY)

    valid_results = [success_results[k] for k in PROPERTIES.keys() if k in success_results]

    wb = Workbook()
    wb.remove(wb.active)

    consolidated = {
        h: {"cash":0.0,"qr":0.0,"online":0.0,"discount":0.0,"total":0.0}
        for h in range(24)
    }

    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    def hour_label(h):
        start = datetime(2000, 1, 1, h, 0)
        end = start + timedelta(hours=1)
        return f"{start.strftime('%I%p').lstrip('0')} - {end.strftime('%I%p').lstrip('0')}"

    def create_sheet(ws, hourly_cash):

        ws.append(["Date", "Time (Hourly)", "Cash", "QR", "Online", "Discount", "Total"])

        header_fill = PatternFill("solid", fgColor="1F4E78")

        for col in range(1, 8):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")

        sum_cash = sum_qr = sum_online = sum_discount = sum_total = 0

        for h in range(24):

            row = hourly_cash.get(h, {"cash":0,"qr":0,"online":0,"discount":0,"total":0})

            cash = round(row["cash"], 2)
            qr = round(row["qr"], 2)
            online = round(row["online"], 2)
            discount = round(row["discount"], 2)
            total = round(row["total"], 2)

            sum_cash += cash
            sum_qr += qr
            sum_online += online
            sum_discount += discount
            sum_total += total

            ws.append([
                display_date,
                hour_label(h),
                cash,
                qr,
                online,
                discount,
                total
            ])

            r = ws.max_row
            fill = PatternFill("solid", fgColor=get_hour_color(h))

            for c in range(1, 8):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill
                cell.border = thin
                cell.alignment = Alignment(horizontal="center")

        ws.append([
            "",
            "TOTAL",
            round(sum_cash,2),
            round(sum_qr,2),
            round(sum_online,2),
            round(sum_discount,2),
            round(sum_total,2)
        ])

        total_row = ws.max_row

        for c in range(1, 8):
            cell = ws.cell(row=total_row, column=c)
            cell.fill = PatternFill("solid", fgColor="000000")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        

        chart_titles = ["Cash", "QR", "Online", "Discount", "Total"]

        base_chart_row = ws.max_row + 3
        chart_gap = 22

        for i, col in enumerate(range(3, 8)):

            chart = BarChart()
            chart.title = f"Hourly {chart_titles[i]}"
            chart.height = 12
            chart.width = 26
            chart.legend = None

            data = Reference(ws, min_col=col, min_row=1, max_row=25)
            cats = Reference(ws, min_col=2, min_row=2, max_row=25)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            series = chart.series[0]
            points = []

            for h in range(24):
                dp = DataPoint(idx=h)
                dp.graphicalProperties.solidFill = get_hour_color(h)
                points.append(dp)

            series.dPt = points

            chart_row = base_chart_row + (i * chart_gap)
            ws.add_chart(chart, f"A{chart_row}")
        autofit_columns(ws)

    for name, hourly_cash in valid_results:

        ws = wb.create_sheet(name[:31])
        create_sheet(ws, hourly_cash)

        for h in range(24):
            for k in consolidated[h]:
                consolidated[h][k] += hourly_cash[h][k]


    ws = wb.create_sheet("CONSOLIDATED")
    create_sheet(ws, consolidated)


    # ================= PROPERTY RANKING =================

    ranking_data = []

    for name, hourly_cash in valid_results:

        total_cash = sum(v["cash"] for v in hourly_cash.values())
        total_qr = sum(v["qr"] for v in hourly_cash.values())
        total_online = sum(v["online"] for v in hourly_cash.values())
        total_discount = sum(v["discount"] for v in hourly_cash.values())
        total_total = sum(v["total"] for v in hourly_cash.values())

        ranking_data.append({
            "name": name,
            "cash": total_cash,
            "qr": total_qr,
            "online": total_online,
            "discount": total_discount,
            "total": total_total
        })

    ranking_data.sort(key=lambda x: x["total"], reverse=True)

    ws = wb.create_sheet("PROPERTY RANKING")

    headers = ["Rank", "Property", "Cash", "QR", "Online", "Discount", "Total", "Badge"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")

    for col in range(1, 9):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")

    widths = [10, 28, 14, 14, 14, 14, 16, 18]

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w


    def get_medal(rank):
        if rank == 1:
            return "🥇 Gold"
        if rank == 2:
            return "🥈 Silver"
        if rank == 3:
            return "🥉 Bronze"
        return ""

    rank = 1

    for idx, p in enumerate(ranking_data):

        medal = get_medal(rank)

        ws.append([
            rank,
            p["name"],
            round(p["cash"], 2),
            round(p["qr"], 2),
            round(p["online"], 2),
            round(p["discount"], 2),
            round(p["total"], 2),
            medal
        ])

        r = ws.max_row
        fill = PatternFill("solid", fgColor=get_hour_color(idx))

        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")

        rank += 1
    autofit_columns(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    await send_telegram_excel_buffer(
        buffer,
        filename=f"Collection_Daily.xlsx",
        caption="📊 Hourly Collection Report"
    )

    print("✅ EXCEL SENT SUCCESSFULLY")


# ================= RUN =================

if __name__ == "__main__":

    try:
        asyncio.run(main())

    except Exception:

        print("SCRIPT CRASHED")
        traceback.print_exc()
