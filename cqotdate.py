# ==============================
# ULTRA FAST ASYNC MULTI PROPERTY AUTOMATION
# HOURLY CASH + QR + ONLINE + TOTAL REPORT
# ==============================

import os
import json
import asyncio
import aiohttp
from datetime import datetime, timedelta
import traceback
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from io import BytesIO
import pytz


IST = pytz.timezone("Asia/Kolkata")
now = datetime.now(IST)

MAX_FULL_RUN_RETRIES = 5
FULL_RUN_RETRY_DELAY = 10

PROP_PARALLEL_LIMIT = 3
DETAIL_PARALLEL_LIMIT = 10

prop_semaphore = asyncio.Semaphore(PROP_PARALLEL_LIMIT)

DETAIL_TIMEOUT = 25
BATCH_TIMEOUT = 35


# ================= TELEGRAM =================

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
if not TELEGRAM_BOT_TOKEN:
    raise RuntimeError("❌ TELEGRAM_BOT_TOKEN missing")

CHAT_MAP = json.loads(os.getenv("TELEGRAM_CHAT_MAP", "{}"))

def get_chat_id(name: str):
    if name not in CHAT_MAP:
        raise RuntimeError(f"❌ Chat ID not configured: {name}")
    return int(CHAT_MAP[name])

TELEGRAM_CHAT_ID = get_chat_id("6am")


async def send_telegram_excel_buffer(buffer, filename, caption=None):

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"

    data = aiohttp.FormData()
    data.add_field("chat_id", str(TELEGRAM_CHAT_ID))

    if caption:
        data.add_field("caption", caption)

    data.add_field(
        "document",
        buffer,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    async with aiohttp.ClientSession() as session:
        async with session.post(url, data=data, timeout=120) as resp:
            if resp.status != 200:
                text = await resp.text()
                raise RuntimeError(f"Telegram send failed: {text}")


# ================= PROPERTIES =================

PROPERTIES_RAW = json.loads(os.getenv("OYO_PROPERTIES", "{}"))
PROPERTIES = {int(k): v for k, v in PROPERTIES_RAW.items()}

if not PROPERTIES:
    raise RuntimeError("❌ OYO_PROPERTIES secret missing or empty")


# ================= COLOR =================

# ================= MONTH DAY COLOR =================

def get_hour_color(day_index: int, total_days: int = 31):
    """
    Unique color for each day in month.
    Natural sun → sunset → night progression.
    Works for 28–31 days.
    """

    if total_days <= 1:
        total_days = 31

    # Normalize position 0 → 1 across month
    t = day_index / (total_days - 1)

    # Solar gradient anchors
    colors = [
        (238, 243, 251),  # dawn blue
        (217, 242, 255),  # morning sky
        (255, 244, 204),  # sunlight
        (255, 221, 128),  # noon warm
        (255, 204, 153),  # evening
        (255, 193, 193),  # sunset
        (232, 234, 246),  # twilight
        (227, 242, 253)   # night blue
    ]

    # Map t across segments
    seg = t * (len(colors) - 1)
    i = int(seg)
    frac = seg - i

    if i >= len(colors) - 1:
        r, g, b = colors[-1]
    else:
        r1, g1, b1 = colors[i]
        r2, g2, b2 = colors[i + 1]

        r = int(r1 + (r2 - r1) * frac)
        g = int(g1 + (g2 - g1) * frac)
        b = int(b1 + (b2 - b1) * frac)

    return f"{r:02X}{g:02X}{b:02X}"


# ================= FETCH DETAILS =================

async def fetch_booking_details(session, P, booking_no):

    url = "https://www.oyoos.com/hms_ms/api/v1/visibility/booking_details_with_entities"

    params = {
        "qid": P["QID"],
        "booking_id": booking_no,
        "role": 0,
        "platform": "OYOOS",
        "country_code": 1
    }

    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}

    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    for attempt in range(1, 4):

        try:
            async with session.get(
                url,
                params=params,
                headers=headers,
                cookies=cookies,
                timeout=DETAIL_TIMEOUT
            ) as r:

                if r.status != 200:
                    raise RuntimeError("DETAIL API FAILED")

                data = await r.json()

                booking = next(
                    iter(data.get("entities", {}).get("bookings", {}).values()),
                    {}
                )

                payments = booking.get("payments", [])

                events = []

                for p in payments:

                    amt = float(p.get("amount", 0) or 0)
                    if amt <= 0:
                        continue

                    created_at = str(p.get("created_at") or "").strip()
                    if not created_at:
                        continue

                    try:
                        dt = datetime.fromisoformat(created_at.replace("Z", ""))
                        dt = dt.astimezone(IST)
                    except Exception:
                        continue

                    mode_raw = p.get("mode", "")

                    if mode_raw == "Cash at Hotel":
                        bucket = "cash"
                    elif mode_raw == "UPI QR":
                        bucket = "qr"
                    elif mode_raw == "oyo_wizard_discount":
                        continue
                    else:
                        bucket = "online"

                    events.append({
                        "date": dt.strftime("%Y-%m-%d"),
                        "hour": dt.hour,
                        "mode": bucket,
                        "amt": amt
                    })

                return events

        except Exception:
            await asyncio.sleep(2 + attempt)

    raise RuntimeError("DETAIL FETCH FAILED")


# ================= FETCH BOOKINGS =================

async def fetch_bookings_batch(session, offset, f, t, P):

    url = "https://www.oyoos.com/hms_ms/api/v1/get_booking_with_ids"

    params = {
        "qid": P["QID"],
        "checkin_from": f,
        "checkin_till": t,
        "batch_count": 100,
        "batch_offset": offset,
        "visibility_required": "true",
        "additionalParams": "payment_hold_transaction,guest,stay_details",
        "decimal_price": "true",
        "ascending": "true",
        "sort_on": "checkin_date"
    }

    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}

    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    async with session.get(
        url,
        params=params,
        cookies=cookies,
        headers=headers,
        timeout=BATCH_TIMEOUT
    ) as r:

        if r.status != 200:
            raise RuntimeError("BATCH API FAILED")

        return await r.json()


# ================= PROCESS PROPERTY =================

async def process_property(P, TF, TT, HF, HT):

    print(f"PROCESSING → {P['name']}")

    tf_dt = datetime.strptime(TF, "%Y-%m-%d").date()
    tt_dt = datetime.strptime(TT, "%Y-%m-%d").date()

    async with aiohttp.ClientSession() as session:

        detail_semaphore = asyncio.Semaphore(DETAIL_PARALLEL_LIMIT)
        detail_cache = {}

        async def limited_detail_call(booking_no):
            async with detail_semaphore:
                if booking_no in detail_cache:
                    return detail_cache[booking_no]
                res = await fetch_booking_details(session, P, booking_no)
                detail_cache[booking_no] = res
                return res

        # ===== DATE MAP =====
        date_map = {}

        d = tf_dt
        while d <= tt_dt:
            date_map[d] = {
                "cash": 0.0,
                "qr": 0.0,
                "online": 0.0,
                "total": 0.0
            }
            d += timedelta(days=1)

        offset = 0

        while True:

            data = await fetch_bookings_batch(session, offset, HF, HT, P)

            if not data or not data.get("bookingIds"):
                break

            bookings = data.get("entities", {}).get("bookings", {})
            if not bookings:
                raise RuntimeError("BOOKINGS EMPTY")

            tasks = []
            mapping = []

            for b in bookings.values():

                status = (b.get("status") or "").strip()
                if status not in ["Checked In", "Checked Out"]:
                    continue

                booking_no = b.get("booking_no")
                if not booking_no:
                    continue

                tasks.append(limited_detail_call(booking_no))
                mapping.append(b)

            results = await asyncio.gather(*tasks, return_exceptions=True)

            for res in results:

                if isinstance(res, Exception):
                    continue

                for ev in res or []:

                    try:
                        d_dt = datetime.strptime(ev["date"], "%Y-%m-%d").date()
                    except:
                        continue

                    if not (tf_dt <= d_dt <= tt_dt):
                        continue

                    amt = float(ev["amt"])
                    mode = ev["mode"]

                    date_map[d_dt][mode] += amt
                    date_map[d_dt]["total"] += amt

            if len(data.get("bookingIds", [])) < 100:
                break

            offset += 100

        return (P["name"], date_map)

# ================= RETRY =================

async def run_property_with_retry(P, TF, TT, HF, HT, retries=5):

    last_error = None

    for attempt in range(1, retries + 1):

        try:
            return await process_property(P, TF, TT, HF, HT)

        except Exception as e:

            last_error = e

            print(f"RETRY {attempt}/{retries} → {P['name']} :: {e}")

            await asyncio.sleep(2 + attempt * 2)

    raise RuntimeError(f"PROPERTY FAILED → {P['name']}") from last_error


# ================= PARALLEL LIMITER =================

async def run_property_limited(P, TF, TT, HF, HT):

    async with prop_semaphore:
        return await run_property_with_retry(P, TF, TT, HF, HT)


# ================ MAIN =================
async def main():

    print("========================================")
    print(" DATE-WISE COLLECTION REPORT")
    print("========================================")

    global now
    now = datetime.now(IST)

    # ================= DATE RANGE =================
    target_date = (now - timedelta(days=1)).date()

    TF = target_date.replace(day=1).strftime("%Y-%m-%d")
    TT = target_date.strftime("%Y-%m-%d")

    HF = (target_date - timedelta(days=120)).strftime("%Y-%m-%d")
    HT = TT

    display_month = datetime.strptime(TT, "%Y-%m-%d").strftime("%B %Y")

    # ================= SMART RETRY (ONLY FAILED PROPERTIES) =================
    pending = {k: v for k, v in PROPERTIES.items()}
    success_results = {}

    for run_attempt in range(1, MAX_FULL_RUN_RETRIES + 1):

        if not pending:
            break

        print(f"\n🔁 PARTIAL RUN ATTEMPT {run_attempt}/{MAX_FULL_RUN_RETRIES}")
        print(f"⏳ Pending Properties: {len(pending)}")

        tasks = [run_property_limited(P, TF, TT, HF, HT) for P in pending.values()]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        new_pending = {}

        for key, (P, result) in zip(list(pending.keys()), zip(pending.values(), results)):

            if isinstance(result, Exception):
                print(f"❌ FAILED → {P['name']} :: {result}")
                new_pending[key] = P
                continue

            success_results[key] = result
            print(f"✅ OK → {P['name']}")

        pending = new_pending

        if pending:

            if run_attempt == MAX_FULL_RUN_RETRIES:

                failed_names = [p["name"] for p in pending.values()]

                raise RuntimeError(
                    f"FINAL FAILURE: Properties failed after retries: {failed_names}"
                )

            print(f"🔁 RETRYING ONLY FAILED PROPERTIES after {FULL_RUN_RETRY_DELAY}s...")
            await asyncio.sleep(FULL_RUN_RETRY_DELAY)

    # ================= FINAL VERIFICATION =================
    valid_results = [success_results[k] for k in PROPERTIES.keys() if k in success_results]

    if len(valid_results) != len(PROPERTIES):

        missing = [PROPERTIES[k]["name"] for k in PROPERTIES.keys() if k not in success_results]

        raise RuntimeError(f"DATA INCOMPLETE: Missing properties: {missing}")

    print("✅ DATA VERIFIED — ALL PROPERTIES PRESENT")

    # ================= EXCEL =================
    wb = Workbook()
    wb.remove(wb.active)

    # ===== DATE LIST =====
    start_dt = datetime.strptime(TF, "%Y-%m-%d").date()
    end_dt = datetime.strptime(TT, "%Y-%m-%d").date()

    date_list = []
    d = start_dt
    while d <= end_dt:
        date_list.append(d)
        d += timedelta(days=1)

    consolidated = {
        d: {"cash": 0.0, "qr": 0.0, "online": 0.0, "total": 0.0}
        for d in date_list
    }

    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # ================= SHEET BUILDER =================
    def create_sheet(ws, date_map):

        ws.append(["Date", "Cash", "QR", "Online", "Total"])

        header_fill = PatternFill("solid", fgColor="1F4E78")

        for col in range(1, 6):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")

        sum_cash = sum_qr = sum_online = sum_total = 0

        for idx, d in enumerate(date_list):

            row = date_map.get(d, {"cash":0,"qr":0,"online":0,"total":0})

            cash = round(row["cash"], 2)
            qr = round(row["qr"], 2)
            online = round(row["online"], 2)
            total = round(row["total"], 2)

            sum_cash += cash
            sum_qr += qr
            sum_online += online
            sum_total += total

            ws.append([
                d.strftime("%d-%m-%Y"),
                cash,
                qr,
                online,
                total
            ])

            r = ws.max_row
            fill = PatternFill("solid", fgColor=get_hour_color(idx, len(date_list)))

            for c in range(1, 6):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill
                cell.border = thin
                cell.alignment = Alignment(horizontal="center")

        ws.append([
            "TOTAL",
            round(sum_cash,2),
            round(sum_qr,2),
            round(sum_online,2),
            round(sum_total,2)
        ])

        total_row = ws.max_row

        for c in range(1, 6):
            cell = ws.cell(row=total_row, column=c)
            cell.fill = PatternFill("solid", fgColor="000000")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 16


    # ===== CHARTS =====
    chart_titles = ["Cash", "QR", "Online", "Total"]

    base_chart_row = ws.max_row + 3
    chart_gap = 22

    for i, col in enumerate(range(2, 6)):

        chart = BarChart()
        chart.title = f"{chart_titles[i]} Trend"
        chart.height = 12
        chart.width = 26
        chart.legend = None

        data = Reference(ws, min_col=col, min_row=1, max_row=len(date_list)+1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(date_list)+1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        series = chart.series[0]
        points = []

        for idx in range(len(date_list)):
            dp = DataPoint(idx=idx)
            dp.graphicalProperties.solidFill = get_hour_color(idx)
            points.append(dp)

        series.dPt = points

        chart_row = base_chart_row + (i * chart_gap)
        ws.add_chart(chart, f"A{chart_row}")
    

    # ================= PROPERTY SHEETS =================
    for name, date_map in valid_results:

        ws = wb.create_sheet(name[:31])
        create_sheet(ws, date_map)

        for d in date_list:
            for k in consolidated[d]:
                consolidated[d][k] += date_map[d][k]

    # ================= CONSOLIDATED =================
    ws = wb.create_sheet("CONSOLIDATED")
    create_sheet(ws, consolidated)

    # ================= PROPERTY RANKING =================
    ranking_data = []

    for name, date_map in valid_results:

        total_cash = sum(v["cash"] for v in date_map.values())
        total_qr = sum(v["qr"] for v in date_map.values())
        total_online = sum(v["online"] for v in date_map.values())
        total_total = sum(v["total"] for v in date_map.values())

        ranking_data.append({
            "name": name,
            "cash": total_cash,
            "qr": total_qr,
            "online": total_online,
            "total": total_total
        })

    ranking_data.sort(key=lambda x: x["total"], reverse=True)

    ws = wb.create_sheet("PROPERTY RANKING")

    headers = ["Rank", "Property", "Cash", "QR", "Online", "Total", "Badge"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")

    for col in range(1, 8):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")

    widths = [10, 28, 14, 14, 14, 16, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    def get_medal(rank):
        if rank == 1:
            return "🥇 Gold"
        if rank == 2:
            return "🥈 Silver"
        if rank == 3:
            return "🥉 Bronze"
        return ""

    rank = 1

    for idx, p in enumerate(ranking_data):

        medal = get_medal(rank)

        ws.append([
            rank,
            p["name"],
            round(p["cash"], 2),
            round(p["qr"], 2),
            round(p["online"], 2),
            round(p["total"], 2),
            medal
        ])

        r = ws.max_row
        fill = PatternFill("solid", fgColor=get_hour_color(idx, len(date_list)))
        
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")

        rank += 1

    # ================= SAVE + TELEGRAM =================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    await send_telegram_excel_buffer(
        buffer,
        filename=f"Collection_{display_month}.xlsx",
        caption="📊 Date-wise Collection Report"
    )

    print("✅ EXCEL SENT SUCCESSFULLY")
            
# ================= RUN =================

if __name__ == "__main__":

    try:
        asyncio.run(main())

    except Exception:

        print("SCRIPT CRASHED")
        traceback.print_exc()
