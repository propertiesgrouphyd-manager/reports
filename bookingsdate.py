# ==============================
# ULTRA FAST ASYNC MULTI PROPERTY AUTOMATION
# DATE-WISE BOOKING MODE REPORT (ULTRA FAST)
# WITH CHARTS + PROPERTY RANKING
# ==============================

import os
import json
import asyncio
import aiohttp
from datetime import datetime, timedelta
import traceback
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from io import BytesIO
import pytz


IST = pytz.timezone("Asia/Kolkata")

MAX_FULL_RUN_RETRIES = 5
FULL_RUN_RETRY_DELAY = 10

PROP_PARALLEL_LIMIT = 4
prop_semaphore = asyncio.Semaphore(PROP_PARALLEL_LIMIT)

BATCH_TIMEOUT = 40


# ================= TELEGRAM =================

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
CHAT_MAP = json.loads(os.getenv("TELEGRAM_CHAT_MAP", "{}"))

def get_chat_id(name: str):
    return int(CHAT_MAP[name])

TELEGRAM_CHAT_ID = get_chat_id("6am")


async def send_telegram_excel_buffer(buffer, filename, caption=None):

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"

    data = aiohttp.FormData()
    data.add_field("chat_id", str(TELEGRAM_CHAT_ID))

    if caption:
        data.add_field("caption", caption)

    data.add_field(
        "document",
        buffer,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    async with aiohttp.ClientSession() as session:
        async with session.post(url, data=data, timeout=120) as resp:
            if resp.status != 200:
                raise RuntimeError(await resp.text())


# ================= PROPERTIES =================

PROPERTIES_RAW = json.loads(os.getenv("OYO_PROPERTIES", "{}"))
PROPERTIES = {int(k): v for k, v in PROPERTIES_RAW.items()}


# ================= COLOR =================
# ================= MONTH DAY COLOR =================

def get_hour_color(day_index: int, total_days: int = 31):
    """
    Unique color for each day in month.
    Natural sun → sunset → night progression.
    Works for 28–31 days.
    """

    if total_days <= 1:
        total_days = 31

    # Normalize position 0 → 1 across month
    t = day_index / (total_days - 1)

    # Solar gradient anchors
    colors = [
        (238, 243, 251),  # dawn blue
        (217, 242, 255),  # morning sky
        (255, 244, 204),  # sunlight
        (255, 221, 128),  # noon warm
        (255, 204, 153),  # evening
        (255, 193, 193),  # sunset
        (232, 234, 246),  # twilight
        (227, 242, 253)   # night blue
    ]

    # Map t across segments
    seg = t * (len(colors) - 1)
    i = int(seg)
    frac = seg - i

    if i >= len(colors) - 1:
        r, g, b = colors[-1]
    else:
        r1, g1, b1 = colors[i]
        r2, g2, b2 = colors[i + 1]

        r = int(r1 + (r2 - r1) * frac)
        g = int(g1 + (g2 - g1) * frac)
        b = int(b1 + (b2 - b1) * frac)

    return f"{r:02X}{g:02X}{b:02X}"


# ================= BOOKING SOURCE =================

def get_booking_source(b):

    source = str(b.get("source", "") or "").strip()
    ota = str(b.get("ota_source", "") or "").strip()
    sub = str(b.get("sub_source", "") or "").strip()
    corp = bool(b.get("is_corporate", False))
    booking_identifier = str(b.get("booking_identifier", "") or "").strip()

    if booking_identifier == "TA":
        return "TA"

    if source == "Walk In":
        return "Walk-in"

    if corp or sub == "corporate":
        return "CB"

    if "Booking.com" in ota:
        return "BDC"

    if "GoMMT" in ota:
        return "MMT"

    if "Agoda" in ota:
        return "Agoda"

    if source == "Travel Agent" or sub == "TPO":
        return "TA"

    if source in [
        "Android App","IOS App","Web Booking",
        "Mobile Web Booking","Website Booking","Direct"
    ]:
        return "OYO"

    return "OBA"


# ================= FETCH BOOKINGS =================

async def fetch_bookings_batch(session, offset, f, t, P):

    url = "https://www.oyoos.com/hms_ms/api/v1/get_booking_with_ids"

    params = {
        "qid": P["QID"],
        "checkin_from": f,
        "checkin_till": t,
        "batch_count": 100,
        "batch_offset": offset,
        "visibility_required": "true",
        "additionalParams": "payment_hold_transaction,guest,stay_details",
        "decimal_price": "true",
        "ascending": "true",
        "sort_on": "checkin_date"
    }

    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}

    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    async with session.get(
        url, params=params, cookies=cookies,
        headers=headers, timeout=BATCH_TIMEOUT
    ) as r:

        if r.status != 200:
            raise RuntimeError("BATCH FAIL")

        return await r.json()




async def process_property(P, TF, TT, HF, HT):

    print(f"PROCESSING → {P['name']}")

    tf_dt = datetime.strptime(TF, "%Y-%m-%d").date()
    tt_dt = datetime.strptime(TT, "%Y-%m-%d").date()

    # ================= DATE MAP =================

    date_map = {}

    d = tf_dt
    while d <= tt_dt:
        date_map[d] = {
            "OYO":0,"Walk-in":0,"MMT":0,"BDC":0,
            "Agoda":0,"CB":0,"TA":0,"OBA":0
        }
        d += timedelta(days=1)

    # ================= SESSION =================

    async with aiohttp.ClientSession() as session:

        offset = 0
        seen_ids = set()          # ⭐ prevents pagination skip / duplicates
        empty_page_streak = 0     # ⭐ safe termination guard

        while True:

            data = await fetch_bookings_batch(session, offset, HF, HT, P)

            if not data:
                break

            bookings = data.get("entities", {}).get("bookings", {})

            if not bookings:
                empty_page_streak += 1
                if empty_page_streak >= 2:
                    break
                offset += 100
                continue

            empty_page_streak = 0

            # ================= LOOP BOOKINGS =================

            for bid, b in bookings.items():

                # ⭐ DEDUP FIX
                if bid in seen_ids:
                    continue

                seen_ids.add(bid)

                status = (b.get("status") or "").strip()

                if status not in ["Checked In", "Checked Out"]:
                    continue

                # ================= CHECKIN ONLY =================

                checkin_str = b.get("checkin")

                if not checkin_str:
                    continue

                try:
                    ci = datetime.strptime(checkin_str, "%Y-%m-%d").date()
                except Exception:
                    continue

                if not (tf_dt <= ci <= tt_dt):
                    continue

                # ================= SOURCE =================

                src = get_booking_source(b)

                if src not in date_map[ci]:
                    src = "OBA"

                rooms = int(
                    b.get("no_of_rooms")
                    or b.get("oyo_rooms")
                    or 1
                )

                date_map[ci][src] += rooms

            # ================= PAGINATION CONTROL =================

            if len(bookings) < 100:
                break

            offset += 100

    return (P["name"], date_map)


# ================= RETRY =================

async def run_property_with_retry(P, TF, TT, HF, HT, retries=3):

    for attempt in range(retries):

        try:
            return await process_property(P, TF, TT, HF, HT)

        except Exception as e:

            print(f"RETRY {attempt+1} → {P['name']} :: {e}")
            await asyncio.sleep(2)

    raise RuntimeError("PROPERTY FAILED")


async def run_property_limited(P, TF, TT, HF, HT):
    async with prop_semaphore:
        return await run_property_with_retry(P, TF, TT, HF, HT)



def autofit_columns(ws):

    for column_cells in ws.columns:

        max_length = 0
        col_letter = column_cells[0].column_letter

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # ===== PREMIUM SPACING =====
        adjusted_width = (max_length * 1.4) + 4   # extra breathing space

        # ===== MINIMUM WIDTH RULES =====
        if col_letter == "A":     # Date / Rank columns
            adjusted_width = max(adjusted_width, 16)

        if col_letter == "B":     # Property name usually
            adjusted_width = max(adjusted_width, 28)

        if col_letter in ["C","D","E","F","G","H","I","J"]:
            adjusted_width = max(adjusted_width, 14)

        ws.column_dimensions[col_letter].width = min(adjusted_width, 45)


# ================= MAIN =================

async def main():

    now = datetime.now(IST)
    target_date = (now - timedelta(days=1)).date()

    TF = target_date.replace(day=1).strftime("%Y-%m-%d")
    TT = target_date.strftime("%Y-%m-%d")

    HF = (target_date - timedelta(days=120)).strftime("%Y-%m-%d")
    HT = TT

    display_month = target_date.strftime("%B %Y")

    # ================= FETCH =================

    # ================= FETCH =================

    pending = {k: v for k, v in PROPERTIES.items()}
    success_results = {}

    for run_attempt in range(1, MAX_FULL_RUN_RETRIES + 1):

        if not pending:
            break

        print(f"\n🔁 PARTIAL RUN ATTEMPT {run_attempt}/{MAX_FULL_RUN_RETRIES}")
        print(f"⏳ Pending Properties: {len(pending)}")

        tasks = [run_property_limited(P, TF, TT, HF, HT) for P in pending.values()]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        new_pending = {}

        for key, (P, result) in zip(list(pending.keys()), zip(pending.values(), results)):

            if isinstance(result, Exception):
                print(f"❌ FAILED → {P['name']} :: {result}")
                new_pending[key] = P
                continue

            success_results[key] = result
            print(f"✅ OK → {P['name']}")

        pending = new_pending

        if pending:

            if run_attempt == MAX_FULL_RUN_RETRIES:
                failed_names = [p["name"] for p in pending.values()]
                raise RuntimeError(
                    f"FINAL FAILURE: Properties failed after retries: {failed_names}"
                )

            print(f"🔁 RETRYING FAILED PROPERTIES AFTER {FULL_RUN_RETRY_DELAY}s...")
            await asyncio.sleep(FULL_RUN_RETRY_DELAY)


    # ================= FINAL VERIFICATION =================

    if len(success_results) != len(PROPERTIES):
        missing = [
            PROPERTIES[k]["name"]
            for k in PROPERTIES.keys()
            if k not in success_results
        ]
        raise RuntimeError(f"DATA INCOMPLETE: Missing properties: {missing}")


    results = [
        success_results[k]
        for k in PROPERTIES.keys()
        if k in success_results
    ]


    # ================= DATE LIST =================

    start_dt = datetime.strptime(TF, "%Y-%m-%d").date()
    end_dt = datetime.strptime(TT, "%Y-%m-%d").date()

    date_list = []

    d = start_dt
    while d <= end_dt:
        date_list.append(d)
        d += timedelta(days=1)

    # ================= EXCEL =================

    wb = Workbook()
    wb.remove(wb.active)

    consolidated = {
        d: {"OYO":0,"Walk-in":0,"MMT":0,"BDC":0,
            "Agoda":0,"CB":0,"TA":0,"OBA":0}
        for d in date_list
    }

    property_totals = []

    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # ================= SHEET BUILDER =================

    def create_sheet(ws, date_map):

        headers = [
            "Date",
            "OYO","Walk-in","MMT","BDC",
            "Agoda","CB","TA","OBA","Total"
        ]

        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1F4E78")

        for col in range(1,len(headers)+1):

            c = ws.cell(row=1,column=col)
            c.fill = header_fill
            c.font = Font(bold=True,color="FFFFFF")
            c.alignment = Alignment(horizontal="center")

        totals = [0]*8

        for idx,d in enumerate(date_list):

            row = date_map[d]

            values = [
                row["OYO"],row["Walk-in"],row["MMT"],row["BDC"],
                row["Agoda"],row["CB"],row["TA"],row["OBA"]
            ]

            total = sum(values)

            for i,v in enumerate(values):
                totals[i]+=v

            ws.append([
                d.strftime("%d-%m-%Y"),
                *values,
                total
            ])

            r = ws.max_row
            fill = PatternFill("solid", fgColor=get_hour_color(idx, len(date_list)))

            for c in range(1,len(headers)+1):

                cell = ws.cell(row=r,column=c)
                cell.fill = fill
                cell.border = thin
                cell.alignment = Alignment(horizontal="center")

        ws.append(["TOTAL",*totals,sum(totals)])

        total_row = ws.max_row

        for col in range(1,len(headers)+1):

            cell = ws.cell(row=total_row,column=col)
            cell.fill = PatternFill("solid", fgColor="000000")
            cell.font = Font(bold=True,color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # ===== CHARTS =====

        start_chart_row = ws.max_row + 3
        chart_gap = 22

        for i,col in enumerate(range(2,11)):

            chart = BarChart()
            chart.height = 12
            chart.width = 26
            chart.title = headers[col-1]
            chart.legend = None

            data = Reference(ws, min_col=col, min_row=1,
                             max_row=len(date_list)+1)

            cats = Reference(ws, min_col=1, min_row=2,
                             max_row=len(date_list)+1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            series = chart.series[0]
            pts = []

            for idx in range(len(date_list)):

                dp = DataPoint(idx=idx)
                dp.graphicalProperties.solidFill = get_hour_color(idx)
                pts.append(dp)

            series.dPt = pts

            ws.add_chart(chart, f"A{start_chart_row + i*chart_gap}")

    # ================= PROPERTY SHEETS =================

    for name,date_map in results:

        ws = wb.create_sheet(name[:31])
        create_sheet(ws,date_map)
        autofit_columns(ws)

        total_prop = 0

        for d in date_list:

            for k in consolidated[d]:
                consolidated[d][k]+=date_map[d][k]

            total_prop += sum(date_map[d].values())

        property_totals.append({
            "name":name,
            "total":total_prop
        })

    # ================= CONSOLIDATED =================

    ws = wb.create_sheet("CONSOLIDATED")
    create_sheet(ws,consolidated)
    autofit_columns(ws)

    # ================= PROPERTY RANKING =================

    ws = wb.create_sheet("PROPERTY RANKING")

    headers = ["Rank","Property","Total Bookings","Badge"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")

    for col in range(1,5):

        c = ws.cell(row=1,column=col)
        c.fill = header_fill
        c.font = Font(bold=True,color="FFFFFF")
        c.alignment = Alignment(horizontal="center")

    def get_medal(rank):

        if rank == 1: return "🥇 Gold"
        if rank == 2: return "🥈 Silver"
        if rank == 3: return "🥉 Bronze"
        return ""

    property_totals.sort(key=lambda x:x["total"], reverse=True)

    rank = 1

    for idx,p in enumerate(property_totals):

        medal = get_medal(rank)

        ws.append([rank,p["name"],p["total"],medal])

        r = ws.max_row
        fill = PatternFill("solid", fgColor=get_hour_color(idx, len(date_list)))

        for c in range(1,5):

            cell = ws.cell(row=r,column=c)
            cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")

        rank+=1
    autofit_columns(ws)

    # ================= SAVE =================

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    await send_telegram_excel_buffer(
        buffer,
        filename=f"Bookings_Monthly.xlsx",
        caption="📊 Date-wise Booking Mode Report"
    )


# ================= RUN =================

if __name__ == "__main__":

    try:
        asyncio.run(main())

    except Exception:

        traceback.print_exc()
