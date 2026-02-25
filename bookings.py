# ==============================
# ULTRA FAST ASYNC MULTI PROPERTY AUTOMATION
# HOURLY BOOKING MODE REPORT
# WITH PROPERTY RANKING
# ==============================

import os
import json
import asyncio
import aiohttp
from datetime import datetime, timedelta
import traceback
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from io import BytesIO
import pytz

IST = pytz.timezone("Asia/Kolkata")

MAX_FULL_RUN_RETRIES = 5
FULL_RUN_RETRY_DELAY = 10

PROP_PARALLEL_LIMIT = 3
prop_semaphore = asyncio.Semaphore(PROP_PARALLEL_LIMIT)

BATCH_TIMEOUT = 35


# ================= TELEGRAM =================

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
CHAT_MAP = json.loads(os.getenv("TELEGRAM_CHAT_MAP", "{}"))

def get_chat_id(name: str):
    return int(CHAT_MAP[name])

TELEGRAM_CHAT_ID = get_chat_id("6am")


async def send_telegram_excel_buffer(buffer, filename, caption=None):

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"

    data = aiohttp.FormData()
    data.add_field("chat_id", str(TELEGRAM_CHAT_ID))

    if caption:
        data.add_field("caption", caption)

    data.add_field(
        "document",
        buffer,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    async with aiohttp.ClientSession() as session:
        async with session.post(url, data=data, timeout=120) as resp:
            if resp.status != 200:
                raise RuntimeError(await resp.text())


# ================= PROPERTIES =================

PROPERTIES_RAW = json.loads(os.getenv("OYO_PROPERTIES", "{}"))
PROPERTIES = {int(k): v for k, v in PROPERTIES_RAW.items()}


# ================= COLOR THEME =================

def get_hour_color(hour):

    palette = [
        "EEF3FB","E8F0FA","E3EDFA","DEEAFA",
        "D9F2FF","DFF7FF","E6FBFF","FFF9DB",
        "FFF4CC","FFEFB3","FFE699","FFDD80",
        "FFE0CC","FFD6B3","FFCC99","FFC280",
        "FFD9D9","FFD1D1","FFC9C9","FFC1C1",
        "F3E5F5","EDE7F6","E8EAF6","E3F2FD"
    ]

    return palette[hour % 24]


# ================= BOOKING SOURCE =================

def get_booking_source(b):

    source = str(b.get("source", "") or "")
    ota = str(b.get("ota_source", "") or "")
    sub = str(b.get("sub_source", "") or "")
    corp = bool(b.get("is_corporate", False))

    if source == "Walk In":
        return "Walk-in"

    if corp or sub == "corporate":
        return "CB"

    if "Booking.com" in ota:
        return "BDC"

    if "GoMMT" in ota:
        return "MMT"

    if "Agoda" in ota:
        return "Agoda"

    if source in ["Android App","IOS App","Web Booking","Mobile Web Booking","Website Booking","Direct"]:
        return "OYO"

    if source == "Travel Agent":
        return "TA"

    return "OBA"


# ================= FETCH =================

async def fetch_bookings_batch(session, offset, f, t, P):

    url = "https://www.oyoos.com/hms_ms/api/v1/get_booking_with_ids"

    params = {
        "qid": P["QID"],
        "checkin_from": f,
        "checkin_till": t,
        "batch_count": 100,
        "batch_offset": offset
    }

    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}

    headers = {
        "accept": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    async with session.get(url, params=params, cookies=cookies,
                           headers=headers, timeout=BATCH_TIMEOUT) as r:

        if r.status != 200:
            raise RuntimeError("BATCH FAIL")

        return await r.json()


# ================= PROCESS PROPERTY =================

# ================= PROCESS PROPERTY (FIXED ENGINE) =================
async def process_property(P, TF, TT, HF, HT):

    print(f"PROCESSING FAST ASYNC → {P['name']}")

    tf_dt = datetime.strptime(TF, "%Y-%m-%d").date()
    tt_dt = datetime.strptime(TT, "%Y-%m-%d").date()

    async with aiohttp.ClientSession() as session:

        total_rooms = await fetch_total_rooms(session, P)
        property_details = await fetch_property_details(session, P)

        if total_rooms == 0:
            raise RuntimeError("TOTAL ROOMS FETCH FAILED")

        detail_semaphore = asyncio.Semaphore(DETAIL_PARALLEL_LIMIT)

        async def limited_detail_call(booking_no):
            async with detail_semaphore:
                return await fetch_booking_details(session, P, booking_no)

        all_rows = []
        offset = 0

        upcoming_count = cancelled_count = inhouse_count = checkedout_count = 0

        while True:

            data = await fetch_bookings_batch(session, offset, HF, HT, P)

            if not data or not data.get("bookingIds"):
                break

            bookings = data.get("entities", {}).get("bookings", {})

            if not bookings:
                break

            tasks = []
            mapping = []

            for b in bookings.values():

                status = (b.get("status") or "").strip()

                checkin_str = b.get("checkin")
                checkout_str = b.get("checkout")

                if not checkin_str or not checkout_str:
                    continue

                try:
                    ci = datetime.strptime(checkin_str, "%Y-%m-%d")
                    co = datetime.strptime(checkout_str, "%Y-%m-%d")
                except:
                    continue

                # ================= FILTER TARGET RANGE =================
                if not (tf_dt <= ci.date() <= tt_dt):
                    continue

                # ================= STATUS COUNTS =================
                if status == "Checked In":
                    inhouse_count += 1
                elif status == "Checked Out":
                    checkedout_count += 1
                elif status == "Confirm Booking":
                    upcoming_count += 1
                elif status == "Cancelled Booking":
                    cancelled_count += 1

                if status not in ["Checked In", "Checked Out"]:
                    continue

                tasks.append(limited_detail_call(b["booking_no"]))
                mapping.append((b, ci, co))

            # ================= DETAIL CALL =================
            results = await asyncio.gather(*tasks, return_exceptions=True)

            for res, (b, ci, co) in zip(results, mapping):

                if isinstance(res, Exception):
                    continue

                rooms, cash, qr, online, discount, balance = res

                stay = max((co - ci).days, 1)

                paid = float(b.get("get_amount_paid") or 0)
                total_amt = paid + float(balance or 0)

                all_rows.append({
                    "Date": ci.strftime("%Y-%m-%d"),
                    "Booking Id": b["booking_no"],
                    "Guest Name": b.get("guest_name"),
                    "Status": b.get("status"),
                    "Booking Source": get_booking_source(b),
                    "Check In": b["checkin"],
                    "Check Out": b["checkout"],
                    "Rooms": b.get("no_of_rooms", 1),
                    "Room Numbers": ", ".join(rooms),
                    "Amount": round(total_amt / stay, 2),
                    "Cash": round(cash / stay, 2),
                    "QR": round(qr / stay, 2),
                    "Online": round(online / stay, 2),
                    "Discount": round(discount / stay, 2),
                    "Balance": round(balance / stay, 2),
                })

            if len(data.get("bookingIds", [])) < 100:
                break

            offset += 100

        df = pd.DataFrame(all_rows)

        if df.empty:
            print(f"⚠️ NO ROWS → {P['name']}")

            df = pd.DataFrame(columns=[
                "Date","Booking Id","Guest Name","Status","Booking Source",
                "Check In","Check Out","Rooms","Room Numbers",
                "Amount","Cash","QR","Online","Discount","Balance"
            ])

        return (
            P["name"],
            df,
            total_rooms,
            inhouse_count,
            checkedout_count,
            upcoming_count,
            cancelled_count,
            property_details,
            {"cash":0,"qr":0,"online":0}
        )

# ================= RETRY =================

async def run_property_with_retry(P, TF, TT, HF, HT, retries=3):

    for _ in range(retries):
        try:
            return await process_property(P, TF, TT, HF, HT)
        except:
            await asyncio.sleep(2)

    raise RuntimeError("PROPERTY FAILED")


async def run_property_limited(P, TF, TT, HF, HT):
    async with prop_semaphore:
        return await run_property_with_retry(P, TF, TT, HF, HT)


# ================= MAIN =================
async def main():

    now = datetime.now(IST)
    target_date = (now - timedelta(days=1)).date()

    TF = target_date.replace(day=1).strftime("%Y-%m-%d")
    TT = target_date.strftime("%Y-%m-%d")

    HF = (target_date - timedelta(days=120)).strftime("%Y-%m-%d")
    HT = TT

    display_month = target_date.strftime("%B %Y")

    pending = {k: v for k, v in PROPERTIES.items()}
    success = {}

    # ================= FETCH =================
    for _ in range(MAX_FULL_RUN_RETRIES):

        if not pending:
            break

        tasks = [run_property_limited(P, TF, TT, HF, HT) for P in pending.values()]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        new_pending = {}

        for key, (P, res) in zip(list(pending.keys()), zip(pending.values(), results)):

            if isinstance(res, Exception):
                new_pending[key] = P
            else:
                success[key] = res

        pending = new_pending

    valid_results = [success[k] for k in PROPERTIES if k in success]

    # ================= DATE LIST =================
    start_dt = datetime.strptime(TF, "%Y-%m-%d").date()
    end_dt = datetime.strptime(TT, "%Y-%m-%d").date()

    date_list = []
    d = start_dt
    while d <= end_dt:
        date_list.append(d)
        d += timedelta(days=1)

    # ================= EXCEL =================
    wb = Workbook()
    wb.remove(wb.active)

    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    consolidated = {
        d: {"OYO":0,"Walk-in":0,"MMT":0,"BDC":0,"Agoda":0,"CB":0,"TA":0,"OBA":0}
        for d in date_list
    }

    property_totals = []

    # ================= SHEET BUILDER =================
    def create_sheet(ws, date_map):

        headers = [
            "Date",
            "OYO","Walk-in","MMT","BDC","Agoda","CB","TA","OBA","Total"
        ]

        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1F4E78")

        for col in range(1, len(headers)+1):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")

        # widths fixed
        ws.column_dimensions["A"].width = 18
        for col in ["B","C","D","E","F","G","H","I"]:
            ws.column_dimensions[col].width = 12
        ws.column_dimensions["J"].width = 14

        totals = [0]*8

        for idx, d in enumerate(date_list):

            row = date_map[d]

            values = [
                row["OYO"],row["Walk-in"],row["MMT"],row["BDC"],
                row["Agoda"],row["CB"],row["TA"],row["OBA"]
            ]

            total = sum(values)

            for i, v in enumerate(values):
                totals[i] += v

            ws.append([
                d.strftime("%d-%m-%Y"),
                *values,
                total
            ])

            r = ws.max_row
            fill = PatternFill("solid", fgColor=get_hour_color(idx))

            for c in range(1, len(headers)+1):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill
                cell.border = thin
                cell.alignment = Alignment(horizontal="center")

        # total row
        ws.append(["TOTAL", *totals, sum(totals)])

        total_row = ws.max_row

        for col in range(1, len(headers)+1):
            cell = ws.cell(row=total_row, column=col)
            cell.fill = PatternFill("solid", fgColor="000000")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # charts
        start_chart_row = ws.max_row + 3
        chart_gap = 24

        chart_columns = list(range(2,10)) + [10]
        chart_titles = [
            "OYO","Walk-in","MMT","BDC",
            "Agoda","CB","TA","OBA","Total Bookings"
        ]

        for i, col in enumerate(chart_columns):

            chart = BarChart()
            chart.height = 12
            chart.width = 26
            chart.title = chart_titles[i]
            chart.legend = None

            data = Reference(ws, min_col=col, min_row=1, max_row=len(date_list)+1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(date_list)+1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ws.add_chart(chart, f"A{start_chart_row + i*chart_gap}")

    # ================= PROPERTY SHEETS =================
    for name, df, *_ in valid_results:

        # build date map from dataframe
        date_map = {
            d: {"OYO":0,"Walk-in":0,"MMT":0,"BDC":0,"Agoda":0,"CB":0,"TA":0,"OBA":0}
            for d in date_list
        }

        if not df.empty:
            for _, r in df.iterrows():
                try:
                    d_obj = datetime.strptime(r["Date"], "%Y-%m-%d").date()
                except:
                    continue

                if d_obj not in date_map:
                    continue

                src = r.get("Booking Source", "OBA")
                rooms = int(r.get("Rooms", 1))

                if src not in date_map[d_obj]:
                    src = "OBA"

                date_map[d_obj][src] += rooms

        ws = wb.create_sheet(name[:31])
        create_sheet(ws, date_map)

        total_prop = 0

        for d in date_list:

            for k in consolidated[d]:
                consolidated[d][k] += date_map[d][k]

            total_prop += sum(date_map[d].values())

        property_totals.append({
            "name": name,
            "total": total_prop
        })

    # ================= CONSOLIDATED =================
    ws = wb.create_sheet("CONSOLIDATED")
    create_sheet(ws, consolidated)

    # ================= PROPERTY RANKING =================
    ws = wb.create_sheet("PROPERTY RANKING")

    headers = ["Rank","Property","Total Bookings","Badge"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")

    for col in range(1,5):
        c = ws.cell(row=1,column=col)
        c.fill = header_fill
        c.font = Font(bold=True,color="FFFFFF")
        c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20

    def get_medal(rank):
        if rank == 1:
            return "🥇 Gold"
        if rank == 2:
            return "🥈 Silver"
        if rank == 3:
            return "🥉 Bronze"
        return ""

    property_totals.sort(key=lambda x: x["total"], reverse=True)

    rank = 1

    for idx,p in enumerate(property_totals):

        ws.append([
            rank,
            p["name"],
            p["total"],
            get_medal(rank)
        ])

        r = ws.max_row
        fill = PatternFill("solid", fgColor=get_hour_color(idx))

        for c in range(1,5):
            cell = ws.cell(row=r,column=c)
            cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")

        rank += 1

    # ================= SAVE =================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    await send_telegram_excel_buffer(
        buffer,
        filename=f"Booking_Mode_{display_month}.xlsx",
        caption="📊 Date-wise Booking Mode Report"
    )

# ================= RUN =================

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except Exception:
        traceback.print_exc()
