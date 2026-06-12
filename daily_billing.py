from datetime import datetime, timedelta
import pytz

def get_yesterday_ist_range():
    ist = pytz.timezone("Asia/Kolkata")

    now = datetime.now(ist)

    yesterday = now.date() - timedelta(days=1)

    start_dt = ist.localize(
        datetime.combine(yesterday, datetime.min.time())
    )

    end_dt = ist.localize(
        datetime.combine(yesterday, datetime.max.time())
    )

    return start_dt, end_dt




import os

def get_github_headers():
    return {
        "Authorization": f"Bearer {os.environ['GITHUB_TOKEN']}",
        "Accept": "application/vnd.github+json"
    }


import requests

OWNER = "propertiesgrouphyd-manager"
REPO = "reports"

def get_successful_runs(start_dt, end_dt):
    url = f"https://api.github.com/repos/{OWNER}/{REPO}/actions/runs"

    runs = []
    page = 1

    ist = pytz.timezone("Asia/Kolkata")

    while True:
        r = requests.get(
            url,
            headers=get_github_headers(),
            params={
                "per_page": 100,
                "page": page
            },
            timeout=60
        )

        r.raise_for_status()

        data = r.json()

        batch = data.get("workflow_runs", [])

        if not batch:
            break

        for run in batch:

            if run.get("conclusion") != "success":
                continue

            started = datetime.fromisoformat(
                run["run_started_at"].replace(
                    "Z",
                    "+00:00"
                )
            )

            started = started.astimezone(ist)

            if started < start_dt:
                continue

            if start_dt <= started <= end_dt:
                runs.append(run)

        page += 1

    return runs


from datetime import datetime

def get_runtime_seconds(run):
    start = datetime.fromisoformat(
        run["run_started_at"].replace("Z", "+00:00")
    )

    end = datetime.fromisoformat(
        run["updated_at"].replace("Z", "+00:00")
    )

    return int((end - start).total_seconds())

import requests

def download_log_zip(run_id):
    url = (
        f"https://api.github.com/repos/"
        f"{OWNER}/{REPO}/actions/runs/{run_id}/logs"
    )

    r = requests.get(
        url,
        headers=get_github_headers(),
        timeout=120
    )

    r.raise_for_status()

    return r.content



import io
import zipfile

def extract_logs_text(zip_bytes):
    output = []

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
        for name in z.namelist():
            try:
                text = z.read(name).decode(
                    "utf-8",
                    errors="ignore"
                )

                output.append(text)

            except Exception:
                pass

    return "\n".join(output)


import re

def parse_usage(text):
    def get_value(key, default="0"):
        m = re.search(
            rf"{key}=([^\r\n]+)",
            text
        )

        return (
            m.group(1).strip()
            if m else default
        )

    workflow = get_value(
        "USAGE_WORKFLOW",
        ""
    )

    properties = get_value(
        "USAGE_PROPERTIES",
        ""
    )

    if not workflow:
        return None

    return {
        "USAGE_WORKFLOW": workflow,
        "USAGE_PROPERTIES": properties or "0",
        "USAGE_MESSAGES": get_value(
            "USAGE_MESSAGES",
            "0"
        ),
        "USAGE_EARLY_ALERTS": get_value(
            "USAGE_EARLY_ALERTS",
            "0"
        ),
        "USAGE_LATE_ALERTS": get_value(
            "USAGE_LATE_ALERTS",
            "0"
        ),
        "USAGE_FILES": get_value(
            "USAGE_FILES",
            "0"
        ),
    }


def calculate_sheets(properties, files):
    if int(files) <= 0:
        return 0

    return int(properties) + int(files)

def calculate_usage_cost(
    messages,
    early_alerts,
    late_alerts,
    files,
    sheets,
    runtime_sec
):
    return (
        messages * 1.0 +
        early_alerts * 2.0 +
        late_alerts * 2.0 +
        files * 5.0 +
        sheets * 0.5 +
        runtime_sec * 0.5
    )


def get_fixed_costs():
    return {
        "Infrastructure Cost": 100,
        "Technology Cost": 500,
        "Operations Cost": 150,
        "Maintenance Cost": 250,
    }


def build_daily_records():
    start_dt, end_dt = get_yesterday_ist_range()

    runs = get_successful_runs(start_dt, end_dt)

    print(
        f"TOTAL RUNS FOUND = "
        f"{len(runs)}"
    )

    records = []

    for run in runs:

        if run.get("name") in [
            "Daily Property Reports Billing",
            "Monthly Property Reports Billing"
        ]:
            continue

        try:
            run_id = run["id"]

            runtime_sec = get_runtime_seconds(run)

            zip_bytes = download_log_zip(run_id)

            logs_text = extract_logs_text(zip_bytes)

            usage = parse_usage(logs_text)

            if not usage:

                usage = {
                    "USAGE_WORKFLOW": run.get(
                        "name",
                        "Unknown"
                    ),
                    "USAGE_PROPERTIES": "0",
                    "USAGE_MESSAGES": "0",
                    "USAGE_EARLY_ALERTS": "0",
                    "USAGE_LATE_ALERTS": "0",
                    "USAGE_FILES": "0",
                }

                print(
                    f"ASSUMED_USAGE -> "
                    f"{run.get('name')}"
                )
            else:
                print(
                    f"FOUND_USAGE -> "
                    f"{usage['USAGE_WORKFLOW']}"
                )



            properties = int(usage["USAGE_PROPERTIES"])
            messages = int(usage["USAGE_MESSAGES"])
            early_alerts = int(usage["USAGE_EARLY_ALERTS"])
            late_alerts = int(usage["USAGE_LATE_ALERTS"])
            files = int(usage["USAGE_FILES"])

            sheets = calculate_sheets(
                properties,
                files
            )

            cost = calculate_usage_cost(
                messages,
                early_alerts,
                late_alerts,
                files,
                sheets,
                runtime_sec
            )

            records.append({
                "Report Name": usage["USAGE_WORKFLOW"],
                "Properties": properties,
                "Messages": messages,
                "Early Alerts": early_alerts,
                "Late Alerts": late_alerts,
                "Files": files,
                "Sheets": sheets,
                "Runtime Sec": runtime_sec,
                "Cost": round(cost, 2)
            })

        except Exception as e:
            print(
                f"SKIPPED RUN "
                f"{run.get('id')} : {e}"
            )

    return records


from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment

def generate_excel(records):
    start_dt, _ = get_yesterday_ist_range()

    bill_date = start_dt.strftime("%d-%b-%Y")

    filename = (
        f"Property_Reports_Bill_"
        f"{start_dt.strftime('%Y-%m-%d')}.xlsx"
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Property Reports Bill"

    ws.merge_cells("A1:J1")

    ws["A1"] = "PROPERTY REPORTS BILL"

    ws["A1"].font = Font(
        bold=True,
        size=16
    )

    ws["A1"].alignment = Alignment(
        horizontal="center"
    )

    ws["A3"] = f"Billing Date : {bill_date}"

    headers = [
        "S.No",
        "Report Name",
        "Properties",
        "Messages (₹1)",
        "Early Alerts (₹2)",
        "Late Alerts (₹2)",
        "Files (₹5)",
        "Sheets (₹0.5)",
        "Runtime Sec (₹0.5)",
        "Cost"
    ]

    row = 5

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)

        cell.value = header

        cell.font = Font(bold=True)

    row += 1
    ws.freeze_panes = "A6"

    total_properties = 0
    total_messages = 0
    total_early = 0
    total_late = 0
    total_files = 0
    total_sheets = 0
    total_runtime = 0
    usage_cost = 0

    for idx, rec in enumerate(records, start=1):
        cell = ws.cell(row=row, column=1)

        cell.value = idx

        cell.alignment = Alignment(
            horizontal="left"
        )
        ws.cell(row=row, column=2).value = rec["Report Name"]
        ws.cell(row=row, column=3).value = rec["Properties"]
        ws.cell(row=row, column=4).value = rec["Messages"]
        ws.cell(row=row, column=5).value = rec["Early Alerts"]
        ws.cell(row=row, column=6).value = rec["Late Alerts"]
        ws.cell(row=row, column=7).value = rec["Files"]
        ws.cell(row=row, column=8).value = rec["Sheets"]
        ws.cell(row=row, column=9).value = rec["Runtime Sec"]
        ws.cell(row=row, column=10).value = rec["Cost"]

        total_properties += rec["Properties"]
        total_messages += rec["Messages"]
        total_early += rec["Early Alerts"]
        total_late += rec["Late Alerts"]
        total_files += rec["Files"]
        total_sheets += rec["Sheets"]
        total_runtime += rec["Runtime Sec"]
        usage_cost += rec["Cost"]

        row += 1

    row += 2

    ws.cell(row=row, column=1).value = "USAGE SUMMARY"
    ws.cell(row=row, column=1).font = Font(bold=True)

    row += 1

    ws.cell(row=row, column=1).value = "Total Successful Runs"
    ws.cell(row=row, column=2).value = len(records)

    row += 1
    ws.cell(row=row, column=1).value = "Total Properties"
    ws.cell(row=row, column=2).value = total_properties

    row += 1
    ws.cell(row=row, column=1).value = "Total Messages"
    ws.cell(row=row, column=2).value = total_messages

    row += 1
    ws.cell(row=row, column=1).value = "Total Early Alerts"
    ws.cell(row=row, column=2).value = total_early

    row += 1
    ws.cell(row=row, column=1).value = "Total Late Alerts"
    ws.cell(row=row, column=2).value = total_late

    row += 1
    ws.cell(row=row, column=1).value = "Total Files"
    ws.cell(row=row, column=2).value = total_files

    row += 1
    ws.cell(row=row, column=1).value = "Total Sheets"
    ws.cell(row=row, column=2).value = total_sheets

    row += 1
    ws.cell(row=row, column=1).value = "Total Runtime Seconds"
    ws.cell(row=row, column=2).value = total_runtime

    row += 3

    fixed = get_fixed_costs()

    message_cost = total_messages * 1.0
    early_cost = total_early * 2.0
    late_cost = total_late * 2.0
    file_cost = total_files * 5.0
    sheet_cost = total_sheets * 0.5
    runtime_cost = total_runtime * 0.5

    variable_cost_total = (
        message_cost +
        early_cost +
        late_cost +
        file_cost +
        sheet_cost +
        runtime_cost
    )

    infra_cost = fixed["Infrastructure Cost"]
    technology_cost = fixed["Technology Cost"]
    operations_cost = fixed["Operations Cost"]
    maintenance_cost = fixed["Maintenance Cost"]

    fixed_cost_total = (
        infra_cost +
        technology_cost +
        operations_cost +
        maintenance_cost
    )

    total_daily_cost = (
        variable_cost_total +
        fixed_cost_total
    )

    ws.cell(row=row, column=1).value = "COST SUMMARY"
    ws.cell(row=row, column=1).font = Font(
        bold=True,
        size=14
    )

    row += 2

    ws.cell(row=row, column=1).value = "VARIABLE COSTS"
    ws.cell(row=row, column=1).font = Font(
        bold=True
    )

    row += 1
    ws.cell(row=row, column=1).value = (
        f"Messages ({total_messages} × ₹1.00)"
    )
    ws.cell(row=row, column=2).value = round(
        message_cost,
        2
    )

    row += 1
    ws.cell(row=row, column=1).value = (
        f"Early Alerts ({total_early} × ₹2.00)"
    )
    ws.cell(row=row, column=2).value = round(
        early_cost,
        2
    )

    row += 1
    ws.cell(row=row, column=1).value = (
        f"Late Alerts ({total_late} × ₹2.00)"
    )
    ws.cell(row=row, column=2).value = round(
        late_cost,
        2
    )

    row += 1
    ws.cell(row=row, column=1).value = (
        f"Files ({total_files} × ₹5.00)"
    )
    ws.cell(row=row, column=2).value = round(
        file_cost,
        2
    )

    row += 1
    ws.cell(row=row, column=1).value = (
        f"Sheets ({total_sheets} × ₹0.50)"
    )
    ws.cell(row=row, column=2).value = round(
        sheet_cost,
        2
    )

    row += 1
    ws.cell(row=row, column=1).value = (
        f"Runtime ({total_runtime} sec × ₹0.50)"
    )
    ws.cell(row=row, column=2).value = round(
        runtime_cost,
        2
    )

    row += 1
    ws.cell(row=row, column=1).value = (
        "Variable Cost Total"
    )
    ws.cell(row=row, column=2).value = round(
        variable_cost_total,
        2
    )

    row += 2

    ws.cell(row=row, column=1).value = "FIXED COSTS"
    ws.cell(row=row, column=1).font = Font(
        bold=True
    )

    row += 1
    ws.cell(row=row, column=1).value = (
        "Infrastructure Cost"
    )
    ws.cell(row=row, column=2).value = infra_cost

    row += 1
    ws.cell(row=row, column=1).value = (
        "Technology Cost"
    )
    ws.cell(row=row, column=2).value = technology_cost

    row += 1
    ws.cell(row=row, column=1).value = (
        "Operations Cost"
    )
    ws.cell(row=row, column=2).value = operations_cost

    row += 1
    ws.cell(row=row, column=1).value = (
        "Maintenance Cost"
    )
    ws.cell(row=row, column=2).value = maintenance_cost

    row += 1
    ws.cell(row=row, column=1).value = (
        "Fixed Cost Total"
    )
    ws.cell(row=row, column=2).value = fixed_cost_total

    row += 2

    ws.cell(row=row, column=1).value = "FINAL BILL"
    ws.cell(row=row, column=1).font = Font(
        bold=True
    )

    row += 1
    ws.cell(row=row, column=1).value = (
        "Variable Cost Total"
    )
    ws.cell(row=row, column=2).value = round(
        variable_cost_total,
        2
    )

    row += 1
    ws.cell(row=row, column=1).value = (
        "Fixed Cost Total"
    )
    ws.cell(row=row, column=2).value = round(
        fixed_cost_total,
        2
    )

    row += 1

    ws.cell(row=row, column=1).value = (
        "TOTAL DAILY BILL"
    )

    ws.cell(row=row, column=1).font = Font(
        bold=True,
        size=14
    )

    ws.cell(row=row, column=2).value = round(
        total_daily_cost,
        2
    )

    ws.cell(row=row, column=2).font = Font(
        bold=True,
        size=14
    )

    from openpyxl.utils import get_column_letter

    for col_num in range(
        1,
        ws.max_column + 1
    ):
        max_length = 0

        column_letter = get_column_letter(
            col_num
        )

        for row_num in range(
            1,
            ws.max_row + 1
        ):
            try:
                value = str(
                    ws.cell(
                        row=row_num,
                        column=col_num
                    ).value
                )

                if len(value) > max_length:
                    max_length = len(value)

            except:
                pass

        ws.column_dimensions[
            column_letter
        ].width = max_length + 4

    wb.save(filename)

    return filename



import aiohttp
import json

def get_chat_id(key):
    chat_map = json.loads(
        os.environ["TELEGRAM_CHAT_MAP"]
    )

    return str(chat_map[key])

async def send_telegram_excel(filename):
    chat_id = get_chat_id("bill")

    url = (
        f"https://api.telegram.org/bot"
        f"{os.environ['TELEGRAM_BOT_TOKEN']}"
        f"/sendDocument"
    )

    timeout = aiohttp.ClientTimeout(
        total=120
    )

    async with aiohttp.ClientSession(
        timeout=timeout
    ) as session:

        with open(filename, "rb") as f:

            form = aiohttp.FormData()

            form.add_field(
                "chat_id",
                chat_id
            )

            form.add_field(
                "caption",
                f"📊 Daily Property Reports Bill\nDate: {get_yesterday_ist_range()[0].strftime('%d-%b-%Y')}"
            )

            form.add_field(
                "document",
                f,
                filename=filename
            )

            async with session.post(
                url,
                data=form
            ) as resp:

                if resp.status != 200:
                    raise RuntimeError(
                        await resp.text()
                    )


import asyncio

async def main():
    records = build_daily_records()

    if not records:
        print(
            "NO BILLABLE WORKFLOWS FOUND"
        )
        return

    filename = generate_excel(records)

    await send_telegram_excel(
        filename
    )

    print(
        f"DAILY BILL SENT : {filename}"
    )

if __name__ == "__main__":
    asyncio.run(main())



