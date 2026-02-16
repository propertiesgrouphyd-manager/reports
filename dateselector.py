# ==============================
# ULTRA FAST ASYNC MULTI PROPERTY AUTOMATION
# MILLION BOOKING READY
# BEAUTIFUL PREMIUM EXCEL
# ==============================

import asyncio
import aiohttp
import pandas as pd
from datetime import datetime, timedelta
import traceback
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import pytz
IST = pytz.timezone("Asia/Kolkata")
now = datetime.now(IST)


MAX_FULL_RUN_RETRIES = 5
FULL_RUN_RETRY_DELAY = 10  # seconds

# ================= NEW: GLOBAL THROTTLES (NO FEATURE REMOVED) =================
PROP_PARALLEL_LIMIT = 3      # max properties running in parallel
DETAIL_PARALLEL_LIMIT = 10   # max detail calls in parallel per property

prop_semaphore = asyncio.Semaphore(PROP_PARALLEL_LIMIT)

# ================= NEW: TIMEOUTS (MORE STABLE MONTHLY) =================
DETAIL_TIMEOUT = 25
ROOMS_TIMEOUT = 25
BATCH_TIMEOUT = 35

TELEGRAM_BOT_TOKEN = "8457091054:AAHNcJeIpf2-ugHbzaFoImlFuN5lxRbcC5Q"
TELEGRAM_CHAT_ID = -5276067828

# ================= PROPERTIES =================
PROPERTIES = {
    1: {"name":"HYD2857","UIF":"eyJlbWFpbCI6Im1vaGRzdWFpZGFobWVkQGdtYWlsLmNvbSIsImFjY2Vzc190b2tlbiI6Ilo1ZUpSMVJiN3FOb3pNNWY0by10YkEiLCJyb2xlIjoiT3duZXIiLCJpZCI6MjAzMzEzMjUyLCJwaG9uZSI6Ijk5ODUyODMzMDYiLCJjb3VudHJ5X2NvZGUiOiIrOTEiLCJkZXZpc2Vfcm9sZSI6Ik93bmVyX1BvcnRhbF9Vc2VyIiwicGhvbmVfdmVyaWZpZWQiOnRydWUsImVtYWlsX3ZlcmlmaWVkIjp0cnVlLCJ1cGRhdGVkX2F0IjoiMTczMjI2MTE0MiIsImZlYXR1cmVzIjp7fSwic3RhdHVzX2NvZGUiOjEwMCwibWlsbGlzX2xlZnRfZm9yX3Bhc3N3b3JkX2V4cGlyeSI6OTQ3MjgyMDk3NDkzLCJhZGRyZXNzSnNvbiI6e319","UUID":"NDFlNWI1ZTQtODFiZC00MWQ1LWIwODAtM2FmMzcwOGYwYmQz","QID":259690},
    2: {"name":"HYD2728","UIF":"eyJlbWFpbCI6ImNoZWYubml0aW5AZ21haWwuY29tIiwiYWNjZXNzX3Rva2VuIjoicEhfUFRkSEVSa2NDTUxYTi15Qk0ydyIsInJvbGUiOiJPd25lciIsImlkIjoyMDQ3MjI0OTMsInBob25lIjoiNjMwOTMzNjMwOSIsImNvdW50cnlfY29kZSI6Iis5MSIsInNleCI6Ik1hbGUiLCJ0ZWFtIjoiTWFya2V0aW5nIiwiZGV2aXNlX3JvbGUiOiJPd25lcl9Qb3J0YWxfVXNlciIsInBob25lX3ZlcmlmaWVkIjp0cnVlLCJlbWFpbF92ZXJpZmllZCI6dHJ1ZSwidXBkYXRlZF9hdCI6IjE3Njk1MTY0MDYiLCJmZWF0dXJlcyI6e30sInN0YXR1c19jb2RlIjoxMDAsIm1pbGxpc19sZWZ0X2Zvcl9wYXNzd29yZF9leHBpcnkiOjk1MDM4NDk3MzkyOSwiYWRkcmVzc0pzb24iOnt9fQ%3D%3D","UUID":"MzgzZWM2MmUtOGJmOC00MjZiLThhY2ItZGFiYWMwNGU5NDQ5","QID":245844},
    3: {"name":"HYD2927","UIF":"eyJlbWFpbCI6InVwcGFsYXNhaTg4QGdtYWlsLmNvbSIsImFjY2Vzc190b2tlbiI6ImNLeVp6WUVFZTR5SDhIMDk3bGJNUUEiLCJyb2xlIjoiT3duZXIiLCJpZCI6MjE2Mzk4NDcwLCJwaG9uZSI6Ijg2ODYwNjY2NjYiLCJjb3VudHJ5X2NvZGUiOiIrOTEiLCJkZXZpc2Vfcm9sZSI6Ik93bmVyX1BvcnRhbF9Vc2VyIiwicGhvbmVfdmVyaWZpZWQiOnRydWUsImVtYWlsX3ZlcmlmaWVkIjp0cnVlLCJ1cGRhdGVkX2F0IjoiMTczNzc4NTUxNCIsImZlYXR1cmVzIjp7fSwic3RhdHVzX2NvZGUiOjEwMCwibWlsbGlzX2xlZnRfZm9yX3Bhc3N3b3JkX2V4cGlyeSI6OTQ1MDE5Mjk0NzUzLCJhZGRyZXNzSnNvbiI6e319","UUID":"N2Y3ZjdiM2ItNGZiMy00MzlmLTk0MTYtNTlkNzRlZjk3MjA3","QID":292909},
    4: {"name":"HYD3030","UIF":"eyJlbWFpbCI6InN2aG90ZWw5OTlAZ21haWwuY29tIiwiYWNjZXNzX3Rva2VuIjoic2x5Y2Fta0pBbU9uZUt2SXlwOWVsUSIsInJvbGUiOiJPd25lciIsImlkIjoyMzU5NTMyNTAsInBob25lIjoiOTk4NTM0NzQ3NiIsImNvdW50cnlfY29kZSI6Iis5MSIsImZpcnN0X25hbWUiOiJHdWVzdCIsInNleCI6Ik1hbGUiLCJ0ZWFtIjoiTWFya2V0aW5nIiwiZGV2aXNlX3JvbGUiOiJPd25lcl9Qb3J0YWxfVXNlciIsInBob25lX3ZlcmlmaWVkIjp0cnVlLCJlbWFpbF92ZXJpZmllZCI6dHJ1ZSwidXBkYXRlZF9hdCI6IjE3NTM4ODQzOTEiLCJmZWF0dXJlcyI6e30sInN0YXR1c19jb2RlIjoxMDAsIm1pbGxpc19sZWZ0X2Zvcl9wYXNzd29yZF9leHBpcnkiOjk0MDI5NDU3MjI0MywiYWRkcmVzc0pzb24iOnt9fQ%3D%3D","UUID":"ZjNjZmZkMWQtOTJiMS00ZjM3LWE1YWMtZGQ3NGExNGIwN2Q5","QID":304236},
    5: {"name":"HYD1170","UIF":"eyJlbWFpbCI6ImtsLmdyYW5kLmhvdGVsQGdtYWlsLmNvbSIsImFjY2Vzc190b2tlbiI6IlhiTVZVUllmVlNJQUhZSWlRMDRyV0EiLCJyb2xlIjoiT3duZXIiLCJpZCI6MjQ2NTU3NzU4LCJwaG9uZSI6IjkyNDgwMDM3MzgiLCJjb3VudHJ5X2NvZGUiOiIrOTEiLCJmaXJzdF9uYW1lIjoiQW5rZXNoIiwic2V4IjoiTWFsZSIsInRlYW0iOiJNYXJrZXRpbmciLCJkZXZpc2Vfcm9sZSI6Ik93bmVyX1BvcnRhbF9Vc2VyIiwicGhvbmVfdmVyaWZpZWQiOnRydWUsImVtYWlsX3ZlcmlmaWVkIjp0cnVlLCJ1cGRhdGVkX2F0IjoiMTc2Mzk3ODgyMCIsImZlYXR1cmVzIjp7fSwic3RhdHVzX2NvZGUiOjEwMCwibWlsbGlzX2xlZnRfZm9yX3Bhc3N3b3JkX2V4cGlyeSI6OTQ0NjkyNTczODQ5LCJhZGRyZXNzSnNvbiI6e319","UUID":"YzRlZWNmMzUtMTllNS00YjVhLTg4YTgtOGIwNGI2NzlkNWQ0","QID":83460},
    6: {"name":"HYD2984","UIF":"eyJlbWFpbCI6InByYXZlZW5hcHV0bHVyaTIwMDdAZ21haWwuY29tIiwiYWNjZXNzX3Rva2VuIjoiZ3FFMVg3RFhDR0RaeEhfQWdMWVpydyIsInJvbGUiOiJPd25lciIsImlkIjoyMTk1ODcyMjQsInBob25lIjoiODcxMjI5NjIxMiIsImNvdW50cnlfY29kZSI6Iis5MSIsImRldmlzZV9yb2xlIjoiT3duZXJfUG9ydGFsX1VzZXIiLCJwaG9uZV92ZXJpZmllZCI6dHJ1ZSwiZW1haWxfdmVyaWZpZWQiOnRydWUsInVwZGF0ZWRfYXQiOiIxNzQzMjQ1Mjc0IiwiZmVhdHVyZXMiOnt9LCJzdGF0dXNfY29kZSI6MTAwLCJtaWxsaXNfbGVmdF9mb3JfcGFzc3dvcmRfZXhwaXJ5Ijo5MjgzNTcxNDY5MDMsImFkZHJlc3NKc29uIjp7fX0%3D","UUID":"ZDY0ODFkMDgtYmVjZi00ZDU5LTgzZWItMmU1Y2U1NjMyMjEy","QID":299149},
    7: {"name":"HYD495","UIF":"eyJlbWFpbCI6Im1hbm9oYXJqb3NoQGdtYWlsLmNvbSIsImFjY2Vzc190b2tlbiI6IjJQMFVURk9lRElKdzZHejA0WlJMTHciLCJyb2xlIjoiT3duZXIiLCJpZCI6NDc0Mjk5MSwicGhvbmUiOiI5OTg1OTk4NTg4IiwiY291bnRyeV9jb2RlIjoiKzkxIiwiZmlyc3RfbmFtZSI6IlZhcmFwcmFzYWRwbXByYXRhcCIsImxhc3RfbmFtZSI6IjgwOTY5OTQ0MjQiLCJjaXR5IjoiIiwic2V4IjoiTWFsZSIsInRlYW0iOiJPd25lciBFbmdhZ2VtZW50IiwiZGV2aXNlX3JvbGUiOiJPd25lcl9Qb3J0YWxfVXNlciIsInBob25lX3ZlcmlmaWVkIjp0cnVlLCJlbWFpbF92ZXJpZmllZCI6dHJ1ZSwiYWRkcmVzcyI6IiIsInVwZGF0ZWRfYXQiOiIxNzYxOTgzODg1IiwiZmVhdHVyZXMiOnt9LCJzdGF0dXNfY29kZSI6MTAwLCJtaWxsaXNfbGVmdF9mb3JfcGFzc3dvcmRfZXhwaXJ5Ijo5NDYwNjAwMzI2MjgsImFkZHJlc3NKc29uIjp7fX0%3D","UUID":"YjAxMWE2MDgtMDc5Ni00OGZlLTliYjEtNDY0OWJkM2IzNzMx","QID":16711},
    8: {"name":"HYD2963","UIF":"eyJlbWFpbCI6InRoaXJ1cGF0aGlyYW90OEBnbWFpbC5jb20iLCJhY2Nlc3NfdG9rZW4iOiJCM1l1U1k1cy0wZE1aeXB1M1l4b2R3Iiwicm9sZSI6Ik93bmVyIiwiaWQiOjExMTEyMjI2MywicGhvbmUiOiI5NTAyMzIzNTEzIiwiY291bnRyeV9jb2RlIjoiKzkxIiwiZmlyc3RfbmFtZSI6InRhbmRyYSIsImxhc3RfbmFtZSI6InRpcnVwYXRoaXJhbyIsImNpdHkiOiIiLCJzZXgiOiJNYWxlIiwidGVhbSI6IlRyYXZlbCBBZ2VudCIsImRldmlzZV9yb2xlIjoiT3duZXJfUG9ydGFsX1VzZXIiLCJwaG9uZV92ZXJpZmllZCI6dHJ1ZSwiZW1haWxfdmVyaWZpZWQiOnRydWUsImFkZHJlc3MiOiIiLCJ1cGRhdGVkX2F0IjoiMTY2NjA5OTMzMyIsImZlYXR1cmVzIjp7fSwic3RhdHVzX2NvZGUiOjEwMCwibWlsbGlzX2xlZnRfZm9yX3Bhc3N3b3JkX2V4cGlyeSI6OTUwMzcxODg3NDAyLCJhZGRyZXNzSnNvbiI6e319","UUID":"YjdlYTZhNDItZGNlNi00NGFhLWI5YzgtODkzZjIyYmM2N2Ri","QID":296969},
    9: {"name":"HYD2012","UIF":"eyJlbWFpbCI6InRoaXJ1cGF0aGlyYW90OEBnbWFpbC5jb20iLCJhY2Nlc3NfdG9rZW4iOiJCM1l1U1k1cy0wZE1aeXB1M1l4b2R3Iiwicm9sZSI6Ik93bmVyIiwiaWQiOjExMTEyMjI2MywicGhvbmUiOiI5NTAyMzIzNTEzIiwiY291bnRyeV9jb2RlIjoiKzkxIiwiZmlyc3RfbmFtZSI6InRhbmRyYSIsImxhc3RfbmFtZSI6InRpcnVwYXRoaXJhbyIsImNpdHkiOiIiLCJzZXgiOiJNYWxlIiwidGVhbSI6IlRyYXZlbCBBZ2VudCIsImRldmlzZV9yb2xlIjoiT3duZXJfUG9ydGFsX1VzZXIiLCJwaG9uZV92ZXJpZmllZCI6dHJ1ZSwiZW1haWxfdmVyaWZpZWQiOnRydWUsImFkZHJlc3MiOiIiLCJ1cGRhdGVkX2F0IjoiMTY2NjA5OTMzMyIsImZlYXR1cmVzIjp7fSwic3RhdHVzX2NvZGUiOjEwMCwibWlsbGlzX2xlZnRfZm9yX3Bhc3N3b3JkX2V4cGlyeSI6OTUwMzcxODg3NDAyLCJhZGRyZXNzSnNvbiI6e319","UUID":"YjdlYTZhNDItZGNlNi00NGFhLWI5YzgtODkzZjIyYmM2N2Ri","QID":196450},
    10: {"name":"HYD1498","UIF":"eyJlbWFpbCI6InRoaXJ1cGF0aGlyYW90OEBnbWFpbC5jb20iLCJhY2Nlc3NfdG9rZW4iOiJCM1l1U1k1cy0wZE1aeXB1M1l4b2R3Iiwicm9sZSI6Ik93bmVyIiwiaWQiOjExMTEyMjI2MywicGhvbmUiOiI5NTAyMzIzNTEzIiwiY291bnRyeV9jb2RlIjoiKzkxIiwiZmlyc3RfbmFtZSI6InRhbmRyYSIsImxhc3RfbmFtZSI6InRpcnVwYXRoaXJhbyIsImNpdHkiOiIiLCJzZXgiOiJNYWxlIiwidGVhbSI6IlRyYXZlbCBBZ2VudCIsImRldmlzZV9yb2xlIjoiT3duZXJfUG9ydGFsX1VzZXIiLCJwaG9uZV92ZXJpZmllZCI6dHJ1ZSwiZW1haWxfdmVyaWZpZWQiOnRydWUsImFkZHJlc3MiOiIiLCJ1cGRhdGVkX2F0IjoiMTY2NjA5OTMzMyIsImZlYXR1cmVzIjp7fSwic3RhdHVzX2NvZGUiOjEwMCwibWlsbGlzX2xlZnRfZm9yX3Bhc3N3b3JkX2V4cGlyeSI6OTUwMzcxODg3NDAyLCJhZGRyZXNzSnNvbiI6e319","UUID":"dlYTZhNDItZGNlNi00NGFhLWI5YzgtODkzZjIyYmM2N2Ri","QID":105249},
    11: {"name":"HYD3183","UIF":"eyJlbWFpbCI6ImthbWFsYWFjaGFAZ21haWwuY29tIiwiYWNjZXNzX3Rva2VuIjoia2RQTVZhV3ZVaGg1cTVaeTMxN3pKUSIsInJvbGUiOiJPd25lciIsImlkIjoyMTg0ODczNjEsInBob25lIjoiOTM5MTA0NDA3MSIsImNvdW50cnlfY29kZSI6Iis5MSIsImRldmlzZV9yb2xlIjoiT3duZXJfUG9ydGFsX1VzZXIiLCJwaG9uZV92ZXJpZmllZCI6dHJ1ZSwiZW1haWxfdmVyaWZpZWQiOnRydWUsInVwZGF0ZWRfYXQiOiIxNzQwNjUyMjIwIiwiZmVhdHVyZXMiOnt9LCJzdGF0dXNfY29kZSI6MTAwLCJtaWxsaXNfbGVmdF9mb3JfcGFzc3dvcmRfZXhwaXJ5Ijo5NDA0NjU5NjU0MjYsImFkZHJlc3NKc29uIjp7fX0%3D","UUID":"YzA1YmE5ODItY2RhMy00MDhiLTk1NzQtNzMzMDA0NTZiM2Yw","QID":328327},
    12: {"name":"HYD1090","UIF":"eyJlbWFpbCI6InNoYW50aGFyZXNpZGVuY3lsb2RnZUBnbWFpbC5jb20iLCJhY2Nlc3NfdG9rZW4iOiJMV1d3VmxHOFhwRHVZQnBySXpkQkhnIiwicm9sZSI6Ik93bmVyIiwiaWQiOjIyMzI4MjUzNCwicGhvbmUiOiI4NTIwMDA1NDc5IiwiY291bnRyeV9jb2RlIjoiKzkxIiwiZmlyc3RfbmFtZSI6Ikd1ZXN0Iiwic2V4IjoiTWFsZSIsInRlYW0iOiJNYXJrZXRpbmciLCJkZXZpc2Vfcm9sZSI6Ik93bmVyX1BvcnRhbF9Vc2VyIiwicGhvbmVfdmVyaWZpZWQiOnRydWUsImVtYWlsX3ZlcmlmaWVkIjp0cnVlLCJ1cGRhdGVkX2F0IjoiMTc0OTIxMjg3NyIsImZlYXR1cmVzIjp7fSwic3RhdHVzX2NvZGUiOjEwMCwibWlsbGlzX2xlZnRfZm9yX3Bhc3N3b3JkX2V4cGlyeSI6OTMzNzUwNTgwMzk5LCJhZGRyZXNzSnNvbiI6e319","UUID":"Zjg4NDc3ZjgtMzM5Zi00ZmYwLWE2OGItYjdkMDEyOGQzNWJk","QID":78637},
    13: {"name":"HYD1762","UIF":"eyJlbWFpbCI6ImtlZXJ0aGljaGFuZHJhOTJAeWFob28uY29tIiwiYWNjZXNzX3Rva2VuIjoiUVF2QURDVmY3R3ZrUFB3Q3Q4SldNQSIsInJvbGUiOiJPd25lciIsImlkIjoxMTA1NjkzOTUsInBob25lIjoiOTk1OTY2NjYwMiIsImNvdW50cnlfY29kZSI6Iis5MSIsImZpcnN0X25hbWUiOiJCYW5kYXJ1IiwibGFzdF9uYW1lIjoiVmVua2F0YXNhdHlha2VlcnRoaSIsImNpdHkiOiIiLCJzZXgiOiJNYWxlIiwidGVhbSI6Ik93bmVyIEVuZ2FnZW1lbnQiLCJkZXZpc2Vfcm9sZSI6Ik93bmVyX1BvcnRhbF9Vc2VyIiwicGhvbmVfdmVyaWZpZWQiOnRydWUsImVtYWlsX3ZlcmlmaWVkIjp0cnVlLCJhZGRyZXNzIjoiIiwidXBkYXRlZF9hdCI6IjE3MDczOTk3NTIiLCJmZWF0dXJlcyI6e30sInN0YXR1c19jb2RlIjoxMDAsIm1pbGxpc19sZWZ0X2Zvcl9wYXNzd29yZF9leHBpcnkiOjk1MDM2MDg0MjM0NiwiYWRkcmVzc0pzb24iOnt9fQ%3D%3D","UUID":"M2Q4MzgxMmYtYzlhMS00NDVlLTk3MzUtZmFjMmQ3ODc0YTEx","QID":115451},
    14: {"name":"HYD588","UIF":"eyJlbWFpbCI6ImtlZXJ0aGljaGFuZHJhOTJAeWFob28uY29tIiwiYWNjZXNzX3Rva2VuIjoiUVF2QURDVmY3R3ZrUFB3Q3Q4SldNQSIsInJvbGUiOiJPd25lciIsImlkIjoxMTA1NjkzOTUsInBob25lIjoiOTk1OTY2NjYwMiIsImNvdW50cnlfY29kZSI6Iis5MSIsImZpcnN0X25hbWUiOiJCYW5kYXJ1IiwibGFzdF9uYW1lIjoiVmVua2F0YXNhdHlha2VlcnRoaSIsImNpdHkiOiIiLCJzZXgiOiJNYWxlIiwidGVhbSI6Ik93bmVyIEVuZ2FnZW1lbnQiLCJkZXZpc2Vfcm9sZSI6Ik93bmVyX1BvcnRhbF9Vc2VyIiwicGhvbmVfdmVyaWZpZWQiOnRydWUsImVtYWlsX3ZlcmlmaWVkIjp0cnVlLCJhZGRyZXNzIjoiIiwidXBkYXRlZF9hdCI6IjE3MDczOTk3NTIiLCJmZWF0dXJlcyI6e30sInN0YXR1c19jb2RlIjoxMDAsIm1pbGxpc19sZWZ0X2Zvcl9wYXNzd29yZF9leHBpcnkiOjk1MDM2MDg0MjM0NiwiYWRkcmVzc0pzb24iOnt9fQ%3D%3D","UUID":"M2Q4MzgxMmYtYzlhMS00NDVlLTk3MzUtZmFjMmQ3ODc0YTEx","QID":37182},
    15: {"name":"WAR144","UIF":"eyJlbWFpbCI6InZpc2hudWdyYW5kLmhhbmFta29uZGFAZ21haWwuY29tIiwiYWNjZXNzX3Rva2VuIjoiSUp5Q2dScWVBUHRrT1czMWRRcTJpZyIsInJvbGUiOiJPd25lciIsImlkIjoyMzcwNDQ0MjgsInBob25lIjoiNjMwMTg4ODg0MyIsImNvdW50cnlfY29kZSI6Iis5MSIsImRldmlzZV9yb2xlIjoiT3duZXJfUG9ydGFsX1VzZXIiLCJwaG9uZV92ZXJpZmllZCI6dHJ1ZSwiZW1haWxfdmVyaWZpZWQiOnRydWUsInVwZGF0ZWRfYXQiOiIxNzU0NTQ5MjEyIiwiZmVhdHVyZXMiOnt9LCJzdGF0dXNfY29kZSI6MTAwLCJtaWxsaXNfbGVmdF9mb3JfcGFzc3dvcmRfZXhwaXJ5Ijo5Mzg3MTc2NDI1MjgsImFkZHJlc3NKc29uIjp7fX0%3D","UUID":"OWRhOTk1MjItNzZlMy00ZjkwLWFhODMtN2U3NTM1YzE4YzZi","QID":326437},
    16: {"name":"KMM030","UIF":"eyJlbWFpbCI6ImJsdWVtb29uaG90ZWwyNEBnbWFpbC5jb20iLCJhY2Nlc3NfdG9rZW4iOiJaRUtKbzBGWUpUNWROYWplOS1ocV9nIiwicm9sZSI6Ik93bmVyIiwiaWQiOjIwMzc1ODk1MywicGhvbmUiOiI5MTAwNzE4Mzg3IiwiY291bnRyeV9jb2RlIjoiKzkxIiwiZGV2aXNlX3JvbGUiOiJPd25lcl9Qb3J0YWxfVXNlciIsInBob25lX3ZlcmlmaWVkIjp0cnVlLCJlbWFpbF92ZXJpZmllZCI6dHJ1ZSwidXBkYXRlZF9hdCI6IjE3MjEzOTEzMzkiLCJmZWF0dXJlcyI6e30sInN0YXR1c19jb2RlIjoxMDAsIm1pbGxpc19sZWZ0X2Zvcl9wYXNzd29yZF9leHBpcnkiOjkyODMzNzE4MDMzMywiYWRkcmVzc0pzb24iOnt9fQ%3D%3D","UUID":"NzE2MGQxMDctNDliNS00YWE5LWI4MGMtY2E0ODQ1ZmZmNGIx","QID":244631},
    17: {'name':"NGA028","UIF":"eyJlbWFpbCI6ImtzYW5qZWV2YTlAZ21haWwuY29tIiwiYWNjZXNzX3Rva2VuIjoiX3FQZFdWSjNTeHNINVE3ZGs0S05xdyIsInJvbGUiOiJPd25lciIsImlkIjo3MjA4MjY4OCwicGhvbmUiOiI4NDk5ODgzMzExIiwiY291bnRyeV9jb2RlIjoiKzkxIiwiZmlyc3RfbmFtZSI6IkthbXNhbmkiLCJsYXN0X25hbWUiOiJTYW5qZWV2YSIsInRlYW0iOiJPcGVyYXRpb25zIiwiZGV2aXNlX3JvbGUiOiJPd25lcl9Qb3J0YWxfVXNlciIsInBob25lX3ZlcmlmaWVkIjp0cnVlLCJlbWFpbF92ZXJpZmllZCI6dHJ1ZSwidXBkYXRlZF9hdCI6IjE3NjQ3NTc5NjIiLCJmZWF0dXJlcyI6e30sInN0YXR1c19jb2RlIjoxMDAsIm1pbGxpc19sZWZ0X2Zvcl9wYXNzd29yZF9leHBpcnkiOjk0NzQyMTczMzQzMSwiYWRkcmVzc0pzb24iOnt9fQ%3D%3D","UUID":"NzRkNjcyMmEtNTU5Ni00NWM0LTk3NjQtNmFkZTVjODE5YjQ2","QID": 353264},
}
# ================= TELEGRAM =================
# NEW FEATURE: REUSE SESSION (but keeps same capability)
async def send_telegram_message(text, retries=3, session=None):
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {"chat_id": TELEGRAM_CHAT_ID, "text": text, "parse_mode": "HTML"}

    async def _post(sess):
        async with sess.post(url, json=payload, timeout=25) as resp:
            if resp.status == 200:
                return True
            raise RuntimeError(f"Telegram HTTP {resp.status}")

    # if session not provided, keep old behavior (no feature removed)
    if session is None:
        for attempt in range(1, retries + 1):
            try:
                async with aiohttp.ClientSession() as s:
                    if await _post(s):
                        return
            except Exception as e:
                if attempt == retries:
                    print("❌ TELEGRAM FAILED AFTER RETRIES")
                    print(e)
                await asyncio.sleep(2)
        async with aiohttp.ClientSession() as s:
            if not await _post(s):
                raise RuntimeError("Telegram send failed")
        return

    # session provided (stable + fast)
    for attempt in range(1, retries + 1):
        try:
            if await _post(session):
                return
        except Exception as e:
            if attempt == retries:
                print("❌ TELEGRAM FAILED AFTER RETRIES")
                print(e)
            await asyncio.sleep(2)

    raise RuntimeError("Telegram send failed")

# ================= BEAUTIFY EXCEL =================
def beautify(ws):
    blue = PatternFill("solid", fgColor="1F4E78")
    light1 = PatternFill("solid", fgColor="DDEBF7")
    light2 = PatternFill("solid", fgColor="F2F2F2")
    yellow = PatternFill("solid", fgColor="FFF4CC")

    bold_white = Font(color="FFFFFF", bold=True, size=12)
    bold_black = Font(color="000000", bold=True, size=12)

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    max_row = ws.max_row
    max_col = ws.max_column
    ws.freeze_panes = "A2"

    for col in range(1, max_col + 1):
        c = ws.cell(row=1, column=col)
        c.fill = blue
        c.font = bold_white
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin

    for r in range(2, max_row + 1):
        fill = light1 if r % 2 == 0 else light2
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue

            # ✅ DO NOT override custom styled rows (headings/tables/boxes)
            if cell.fill is not None and cell.fill.patternType is not None:
                # already styled (blue headings, yellow labels etc.) → keep as-is
                cell.border = thin
                continue

            cell.fill = fill
            cell.border = thin


    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 5

    for r in range(2, max_row + 1):
        text = str(ws.cell(row=r, column=1).value or "")
        if text.strip() == "":
            continue
        if "Booking" in text or "Amount" in text or "Total" in text or "OYO" in text:
            ws.cell(row=r, column=1).fill = yellow
            ws.cell(row=r, column=1).font = bold_black

# ================= NEW: PREMIUM PROPERTY DETAILS BOX =================
def add_property_details_box(ws, prop):
    """
    Adds a fixed premium 'box' at bottom:
    Name, Alternate Name, Address, Google Map
    """

    blue = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="DDEBF7")
    white = PatternFill("solid", fgColor="FFFFFF")

    bold_white = Font(color="FFFFFF", bold=True, size=12)
    bold_black = Font(color="000000", bold=True, size=11)
    normal = Font(color="000000", size=11)
    link_font = Font(color="0563C1", underline="single", bold=True, size=11)

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # fixed box width A:H (8 columns) looks premium always
    start_col = 1
    end_col = 8

    def _border_range(r1, c1, r2, c2):
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(row=rr, column=cc).border = thin

    def _merge(row, c1, c2, value, fill=None, font=None, center=False, wrap=False):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1)
        cell.value = value
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center",
            wrap_text=wrap
        )
        return cell

    # build address line
    plot = (prop.get("plot_number") or "").strip()
    street = (prop.get("street") or "").strip()
    pincode = (prop.get("pincode") or "").strip()
    city = (prop.get("city") or "").strip()
    country = (prop.get("country") or "").strip()

    address_parts = []
    if plot: address_parts.append(plot)
    if street: address_parts.append(street)
    city_pin = " ".join([x for x in [city, pincode] if x]).strip()
    if city_pin: address_parts.append(city_pin)
    if country: address_parts.append(country)

    address = ", ".join(address_parts).strip()

    # spacing
    ws.append([])
    ws.append([])

    top = ws.max_row + 1

    # Header
    _merge(top, start_col, end_col, "PROPERTY DETAILS", fill=blue, font=bold_white, center=True)
    ws.row_dimensions[top].height = 22

    # Row 1: Name
    _merge(top + 1, 1, 2, "Name", fill=light, font=bold_black, wrap=True)
    _merge(top + 1, 3, end_col, prop.get("name", "") or "", fill=white, font=normal, wrap=True)
    ws.row_dimensions[top + 1].height = 20

    # Row 2: Alternate Name
    _merge(top + 2, 1, 2, "Alternative Name", fill=light, font=bold_black, wrap=True)
    _merge(top + 2, 3, end_col, prop.get("alternate_name", "") or "", fill=white, font=normal, wrap=True)
    ws.row_dimensions[top + 2].height = 20

    # Row 3: Address
    _merge(top + 3, 1, 2, "Address", fill=light, font=bold_black, wrap=True)
    _merge(top + 3, 3, end_col, address, fill=white, font=normal, wrap=True)
    ws.row_dimensions[top + 3].height = 45

    # Row 4: Google Map (hyperlink + fixed look)
    _merge(top + 4, 1, 2, "Google Map", fill=light, font=bold_black, wrap=True)

    map_link = (prop.get("map_link") or "").strip()
    if not map_link:
        map_link = ""

    link_cell = _merge(top + 4, 3, end_col, "OPEN IN GOOGLE MAPS" if map_link else "", fill=white, font=link_font, center=True)
    if map_link:
        link_cell.hyperlink = map_link

    ws.row_dimensions[top + 4].height = 22

    # apply border to whole box
    _border_range(top, start_col, top + 4, end_col)


# ================= NEW: PREMIUM PAYMENT TABLES =================
def add_payment_tables(ws, df, today_collect, title_prefix=""):
    """
    Adds 2 premium tables BEFORE property details:
    1) Booking Source x Payment Split
    2) Target Date Collected (NOT divided by stay days)

    today_collect = {"cash": x, "qr": x, "online": x}
    """

    blue = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="DDEBF7")
    white = PatternFill("solid", fgColor="FFFFFF")
    yellow = PatternFill("solid", fgColor="FFF4CC")

    bold_white = Font(color="FFFFFF", bold=True, size=12)
    bold_black = Font(color="000000", bold=True, size=11)
    normal = Font(color="000000", size=11)

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def _style_row(row, start_col, end_col, fill=None, font=None, center=True):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=c)
            if fill: cell.fill = fill
            if font: cell.font = font
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")

    def _merge(row, c1, c2, value, fill=None, font=None, center=True):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1)
        cell.value = value
        if fill: cell.fill = fill
        if font: cell.font = font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
        return cell

    # ================= PREMIUM WIDTH SETTINGS =================
    start_col = 1
    end_col = 7  # A:G premium fixed table width

    premium_widths = [18, 14, 14, 14, 14, 14, 16]
    for i, w in enumerate(premium_widths, start=1):
        col_letter = get_column_letter(i)
        current = ws.column_dimensions[col_letter].width
        ws.column_dimensions[col_letter].width = max(current or 0, w)

    # ✅ guaranteed 1 row space before payment tables
    ws.append([])

    # ================= TABLE 1: Booking Source vs Payment Mode =================
    top = ws.max_row + 1
    heading = f"{title_prefix}BOOKING SOURCE × PAYMENT MODE".strip()
    _merge(top, start_col, end_col, heading, fill=blue, font=bold_white, center=True)
    ws.row_dimensions[top].height = 20

    headers = ["Source", "Cash", "QR", "Online", "Discount", "Balance", "Total"]
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=top + 1, column=idx).value = h
    _style_row(top + 1, start_col, end_col, fill=light, font=bold_black)

    sources = ["OYO", "Walk-in", "MMT", "BDC", "Agoda", "CB", "TA", "OBA"]
    r = top + 2

    for src in sources:
        part = df[df["Booking Source"] == src] if (not df.empty and "Booking Source" in df.columns) else df

        cash = round(float(part["Cash"].sum()), 2) if (not part.empty and "Cash" in part.columns) else 0
        qr = round(float(part["QR"].sum()), 2) if (not part.empty and "QR" in part.columns) else 0
        online = round(float(part["Online"].sum()), 2) if (not part.empty and "Online" in part.columns) else 0
        disc = round(float(part["Discount"].sum()), 2) if (not part.empty and "Discount" in part.columns) else 0
        bal = round(float(part["Balance"].sum()), 2) if (not part.empty and "Balance" in part.columns) else 0
        total = round(float(part["Amount"].sum()), 2) if (not part.empty and "Amount" in part.columns) else 0

        ws.cell(row=r, column=1).value = src
        ws.cell(row=r, column=2).value = cash
        ws.cell(row=r, column=3).value = qr
        ws.cell(row=r, column=4).value = online
        ws.cell(row=r, column=5).value = disc
        ws.cell(row=r, column=6).value = bal
        ws.cell(row=r, column=7).value = total

        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin
            cell.font = normal
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = white

        ws.cell(row=r, column=1).fill = yellow
        ws.cell(row=r, column=1).font = bold_black
        r += 1
    # ================= TOTAL ROW (TABLE 1) =================
    total_cash = round(float(df["Cash"].sum()), 2) if (not df.empty and "Cash" in df.columns) else 0
    total_qr = round(float(df["QR"].sum()), 2) if (not df.empty and "QR" in df.columns) else 0
    total_online = round(float(df["Online"].sum()), 2) if (not df.empty and "Online" in df.columns) else 0
    total_disc = round(float(df["Discount"].sum()), 2) if (not df.empty and "Discount" in df.columns) else 0
    total_bal = round(float(df["Balance"].sum()), 2) if (not df.empty and "Balance" in df.columns) else 0
    total_amt = round(float(df["Amount"].sum()), 2) if (not df.empty and "Amount" in df.columns) else 0

    ws.cell(row=r, column=1).value = "TOTAL"
    ws.cell(row=r, column=2).value = total_cash
    ws.cell(row=r, column=3).value = total_qr
    ws.cell(row=r, column=4).value = total_online
    ws.cell(row=r, column=5).value = total_disc
    ws.cell(row=r, column=6).value = total_bal
    ws.cell(row=r, column=7).value = total_amt

    # premium styling for total row
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = thin
        cell.font = bold_black
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = light  # light header style

    ws.cell(row=r, column=1).fill = yellow
    ws.cell(row=r, column=1).font = bold_black

    r += 1


    # ✅ exactly 1 blank row between tables
    ws.append([])



# ================= BOOKING SOURCE =================
def get_booking_source(b):
    source = str(b.get("source", "") or "").strip()
    ota = str(b.get("ota_source", "") or "").strip()
    sub = str(b.get("sub_source", "") or "").strip()
    corp = bool(b.get("is_corporate", False))

    booking_identifier = str(b.get("booking_identifier", "") or "").strip()

    # ✅ strongest indicator first
    if booking_identifier == "TA":
        return "TA"

    # ✅ walk-in
    if source == "Walk In":
        return "Walk-in"

    # ✅ corporate
    if corp or sub == "corporate":
        return "CB"

    # ✅ OTA
    if "Booking.com" in ota:
        return "BDC"
    if "GoMMT" in ota:
        return "MMT"
    if "Agoda" in ota:
        return "Agoda"

    # ✅ travel agent / TPO (THIS fixes your sub_source issue)
    if source == "Travel Agent" or sub == "TPO":
        return "TA"

    # ✅ OYO direct
    if source in [
        "Android App",
        "IOS App",
        "Web Booking",
        "Mobile Web Booking",
        "Website Booking",
        "Direct"
    ]:
        return "OYO"

    # ✅ fallback
    return "OBA"

# ================= FETCH DETAILS =================
async def fetch_booking_details(session, P, booking_no):
    url = "https://www.oyoos.com/hms_ms/api/v1/visibility/booking_details_with_entities"
    params = {
        "qid": P["QID"],
        "booking_id": booking_no,
        "role": 0,
        "platform": "OYOOS",
        "country_code": 1
    }
    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}
    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    for attempt in range(1, 4):
        try:
            async with session.get(
                url, params=params, headers=headers, cookies=cookies, timeout=DETAIL_TIMEOUT
            ) as r:
                if r.status != 200:
                    raise RuntimeError("DETAIL API FAILED")

                data = await r.json()

                rooms = []
                stay = data.get("entities", {}).get("stayDetails", {})
                for s in stay.values():
                    rn = s.get("room_number")
                    if rn:
                        rooms.append(rn)

                booking = next(iter(data.get("entities", {}).get("bookings", {}).values()), {})
                payments = booking.get("payments", [])

                cash = qr = online = discount = 0
                for p in payments:
                    mode = p.get("mode", "")
                    amt = float(p.get("amount", 0))
                    if mode == "oyo_wizard_discount":
                        discount += amt
                    elif mode == "Cash at Hotel":
                        cash += amt
                    elif mode == "UPI QR":
                        qr += amt
                    else:
                        online += amt

                balance = booking.get("payable_amount", 0)
                return rooms, cash, qr, online, discount, balance

        except Exception:
            # NEW FEATURE: Better backoff (reduces rate limits)
            await asyncio.sleep(2 + attempt)

    raise RuntimeError("DETAIL FETCH FAILED")

# ================= BATCH FETCH =================
async def fetch_bookings_batch(session, offset, f, t, P):
    url = "https://www.oyoos.com/hms_ms/api/v1/get_booking_with_ids"
    params = {
        "qid": P["QID"],
        "checkin_from": f,
        "checkin_till": t,
        "batch_count": 100,
        "batch_offset": offset,
        "visibility_required": "true",
        "additionalParams": "payment_hold_transaction,guest,stay_details",
        "decimal_price": "true",
        "ascending": "true",
        "sort_on": "checkin_date"
    }
    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}
    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    async with session.get(
        url, params=params, cookies=cookies, headers=headers, timeout=BATCH_TIMEOUT
    ) as r:
        if r.status != 200:
            raise RuntimeError("BATCH API FAILED")
        return await r.json()


# ================= NEW: PROPERTY DETAILS API =================
async def fetch_property_details(session, P):
    """
    Fetches:
    name, alternate_name, address fields, map_link
    """
    url = "https://www.oyoos.com/hms_ms/api/v1/location/property-details"
    params = {"qid": P["QID"]}
    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}
    headers = {
        "accept": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    for attempt in range(1, 4):
        try:
            async with session.get(url, params=params, cookies=cookies, headers=headers, timeout=20) as r:
                if r.status != 200:
                    raise RuntimeError(f"PROPERTY DETAILS API FAILED ({r.status})")
                data = await r.json()
                return {
                    "name": str(data.get("name", "") or "").strip(),
                    "alternate_name": str(data.get("alternate_name", "") or "").strip(),
                    "plot_number": str(data.get("plot_number", "") or "").strip(),
                    "street": str(data.get("street", "") or "").strip(),
                    "pincode": str(data.get("pincode", "") or "").strip(),
                    "city": str(data.get("city", "") or "").strip(),
                    "country": str(data.get("country", "") or "").strip(),
                    "map_link": str(data.get("map_link", "") or "").strip(),
                    "latitude": data.get("latitude", None),
                    "longitude": data.get("longitude", None),
                }
        except Exception:
            await asyncio.sleep(2 + attempt)

    # fallback safe empty
    return {
        "name": "",
        "alternate_name": "",
        "plot_number": "",
        "street": "",
        "pincode": "",
        "city": "",
        "country": "",
        "map_link": "",
        "latitude": None,
        "longitude": None,
    }


# ================= FETCH TOTAL ROOMS =================
async def fetch_total_rooms(session, P):
    url = "https://www.oyoos.com/hms_ms/api/v1/hotels/roomsNew"
    params = {"qid": P["QID"]}
    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}
    headers = {
        "accept": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    for attempt in range(1, 4):
        try:
            async with session.get(
                url, params=params, cookies=cookies, headers=headers, timeout=ROOMS_TIMEOUT
            ) as r:
                if r.status != 200:
                    raise RuntimeError("ROOM API FAILED")

                data = await r.json()
                rooms = data.get("rooms", {})
                return len(rooms)

        except Exception:
            await asyncio.sleep(2 + attempt)

    return 0

# ================= PROCESS PROPERTY =================
async def process_property(P, TF, TT, HF, HT):
    print(f"PROCESSING FAST ASYNC → {P['name']}")

    async with aiohttp.ClientSession() as session:
        total_rooms = await fetch_total_rooms(session, P)
        property_details = await fetch_property_details(session, P)


        if total_rooms == 0:
            raise RuntimeError("TOTAL ROOMS FETCH FAILED")

        # NEW FEATURE: limit detail calls per property
        detail_semaphore = asyncio.Semaphore(DETAIL_PARALLEL_LIMIT)
        # ================= NEW: TARGET DATE COLLECTION (TF) =================
        target_collect_date = str(TF).strip()  # ✅ today means TARGET FROM date (TF)
        target_collect = {"cash": 0.0, "qr": 0.0, "online": 0.0}
        target_seen_bookings = set()


        async def limited_detail_call(booking_no):
            async with detail_semaphore:
                return await fetch_booking_details(session, P, booking_no)

        all_rows = []
        offset = 0
        upcoming_count = cancelled_count = inhouse_count = checkedout_count = 0

        while True:
            data = await fetch_bookings_batch(session, offset, HF, HT, P)

            if not data or not data.get("bookingIds"):
                break

            bookings = data.get("entities", {}).get("bookings", {})

            if not bookings:
                raise RuntimeError("BOOKING ENTITY EMPTY")

            curr = datetime.strptime(TF, "%Y-%m-%d")
            end = datetime.strptime(TT, "%Y-%m-%d")

            while curr <= end:
                target = curr.strftime("%Y-%m-%d")
                target_dt = curr

                tasks, mapping = [], []

                for b in bookings.values():
                    status = (b.get("status") or "").strip()
                    ci = datetime.strptime(b["checkin"], "%Y-%m-%d")
                    co = datetime.strptime(b["checkout"], "%Y-%m-%d")
                    tf_date = datetime.strptime(TF, "%Y-%m-%d")

                    # ---- STATUS COUNTS (UNCHANGED) ----
                    if status == "Checked In" and ci <= tf_date:
                        inhouse_count += 1
                    elif status == "Checked Out" and co.date() == now.date():
                        checkedout_count += 1
                    elif status == "Confirm Booking" and ci.date() == now.date():
                        upcoming_count += 1
                    elif status == "Cancelled Booking" and ci <= tf_date:
                        cancelled_count += 1

                    if status not in ["Checked In", "Checked Out"]:
                        continue

                    if not (target_dt >= ci and target_dt < co):
                        continue

                    tasks.append(limited_detail_call(b["booking_no"]))
                    mapping.append((b, target, ci, co))

                if tasks:
                    results = await asyncio.gather(*tasks, return_exceptions=True)
                else:
                    results = []

                for res, (b, target, ci, co) in zip(results, mapping):
                    if isinstance(res, Exception):
                        # detail fail shouldn't kill property
                        continue

                    rooms, cash, qr, online, discount, balance = res
                    # ================= NEW: TARGET DATE COLLECTION (NOT DIVIDED) =================
                    # Rule: Collect only bookings whose CHECK-IN is exactly TF
                    # Add once per booking_id (not repeated for each stay date)
                    if str(b.get("checkin", "")).strip() == target_collect_date and b["booking_no"] not in target_seen_bookings:
                        target_seen_bookings.add(b["booking_no"])
                        target_collect["cash"] += float(cash or 0)
                        target_collect["qr"] += float(qr or 0)
                        target_collect["online"] += float(online or 0)

                    stay = max((co - ci).days, 1)
                    paid = float(b.get("get_amount_paid") or 0)
                    total_amt = paid + float(balance or 0)

                    all_rows.append({
                        "Date": target,
                        "Booking Id": b["booking_no"],
                        "Guest Name": b["guest_name"],
                        "Status": b.get("status"),
                        "Booking Source": get_booking_source(b),
                        "Check In": b["checkin"],
                        "Check Out": b["checkout"],
                        "Rooms": b.get("no_of_rooms", 1),
                        "Room Numbers": ", ".join(rooms),
                        "Amount": round(total_amt / stay, 2),
                        "Cash": round(cash / stay, 2),
                        "QR": round(qr / stay, 2),
                        "Online": round(online / stay, 2),
                        "Discount": round(discount / stay, 2),
                        "Balance": round(balance / stay, 2),
                    })

                curr += timedelta(days=1)

            if len(data["bookingIds"]) < 100:
                break

            offset += 100

        df = pd.DataFrame(all_rows)

        # NEW FEATURE: DO NOT FAIL PROPERTY IF NO ROWS
        if df.empty:
            print(f"⚠️ NO ROWS → {P['name']} (month has no stays)")
            df = pd.DataFrame(columns=[
                "Date", "Booking Id", "Guest Name", "Status", "Booking Source",
                "Check In", "Check Out", "Rooms", "Room Numbers",
                "Amount", "Cash", "QR", "Online", "Discount", "Balance"
            ])

        df.columns = [str(c).strip() for c in df.columns]

        return (
            P["name"],
            df,
            total_rooms,
            inhouse_count,
            checkedout_count,
            upcoming_count,
            cancelled_count,
            property_details,
            target_collect
        )

# ================= RELIABILITY WRAPPER =================
async def run_property_with_retry(P, TF, TT, HF, HT, retries=3):
    last_error = None
    for attempt in range(1, retries + 1):
        try:
            return await process_property(P, TF, TT, HF, HT)
        except Exception as e:
            last_error = e
            print(f"RETRY {attempt}/{retries} → {P['name']} :: {e}")
            await asyncio.sleep(2 + attempt * 2)
    raise RuntimeError(f"PROPERTY FAILED → {P['name']}") from last_error

# NEW FEATURE: PROPERTY PARALLEL LIMITER (reduces failures)
async def run_property_limited(P, TF, TT, HF, HT):
    async with prop_semaphore:
        return await run_property_with_retry(P, TF, TT, HF, HT)

# ================= COUNT / AMOUNT =================
def count(df, src):
    if df.empty:
        return 0
    return int(df[df["Booking Source"] == src]["Rooms"].sum())

def amt(df, src):
    if df.empty:
        return 0
    return round(df[df["Booking Source"] == src]["Amount"].sum(), 2)

def count_upcoming(df, tf):
    tf_date = datetime.strptime(tf, "%Y-%m-%d")
    next_date = tf_date + timedelta(days=1)

    c = 0
    for _, r in df.iterrows():
        if r.get("Status") != "Confirm Booking":
            continue

        d = datetime.strptime(r["Date"], "%Y-%m-%d")

        if d == tf_date and now.hour >= 12:
            c += 1
        elif d == next_date and now.hour < 12:
            c += 1

    return c

def count_cancelled(df, tf):
    tf_date = datetime.strptime(tf, "%Y-%m-%d")
    next_date = tf_date + timedelta(days=1)

    c = 0
    for _, r in df.iterrows():
        if r.get("Status") != "Cancelled Booking":
            continue

        d = datetime.strptime(r["Date"], "%Y-%m-%d")

        if d == tf_date or d == next_date:
            c += 1

    return c






async def send_telegram_excel_buffer(buffer, filename, caption=None):
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"

    data = aiohttp.FormData()
    data.add_field("chat_id", str(TELEGRAM_CHAT_ID))
    if caption:
        data.add_field("caption", caption)

    data.add_field(
        "document",
        buffer,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    async with aiohttp.ClientSession() as session:
        async with session.post(url, data=data, timeout=120) as resp:
            if resp.status != 200:
                text = await resp.text()
                raise RuntimeError(f"Telegram send failed: {text}")

# ================= MAIN =================
# ================= MAIN (MONTH MODE — BASE STRENGTH GUARANTEE) =================
async def main():
    print("========================================")
    print(" OYO DAILY TELEGRAM AUTOMATION")
    print("========================================")

    global now
    now = datetime.now(IST)


    # ================= BUSINESS DATE CUTOVER (12 PM RULE) =================
    if now.hour < 12:
        target_date = (now - timedelta(days=1)).date()
    else:
        target_date = now.date()

    # ================= PREVIOUS MONTH (BASED ON TARGET_DATE) =================
 
    TF = input("Enter TARGET FROM date (YYYY-MM-DD): ").strip()
    TT = input("Enter TARGET TO   date (YYYY-MM-DD): ").strip()

    tf_date = datetime.strptime(TF, "%Y-%m-%d")
    tt_date = datetime.strptime(TT, "%Y-%m-%d")

    if tt_date < tf_date:
        raise ValueError("TARGET TO date cannot be before TARGET FROM date")

    target_days = (tt_date - tf_date).days + 1  # 🔒 BASE EXPECTS THIS

    # ------------------------------------------------
    # HISTORY RANGE (MANDATORY)
    # ------------------------------------------------
    HF = input("Enter HISTORY FROM date (YYYY-MM-DD): ").strip()
    HT = input("Enter HISTORY TO   date (YYYY-MM-DD): ").strip()

    print("\nTARGET RANGE  :", TF, "→", TT)
    print("HISTORY RANGE :", HF, "→", HT)




    # ================= SMART RETRY (ONLY FAILED PROPERTIES) =================
    pending = {k: v for k, v in PROPERTIES.items()}
    success_results = {}

    for run_attempt in range(1, MAX_FULL_RUN_RETRIES + 1):
        if not pending:
            break

        print(f"\n🔁 PARTIAL RUN ATTEMPT {run_attempt}/{MAX_FULL_RUN_RETRIES}")
        print(f"⏳ Pending Properties: {len(pending)}")

        tasks = [run_property_limited(P, TF, TT, HF, HT) for P in pending.values()]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        new_pending = {}
        for key, (P, result) in zip(list(pending.keys()), zip(pending.values(), results)):

            if isinstance(result, Exception):
                print(f"❌ FAILED → {P['name']} :: {result}")
                new_pending[key] = P
                continue

            name, df, *_ = result

            # strict verify (df always exists now; empty allowed as valid)
            if df is None:
                print(f"❌ EMPTY DATA → {name}")
                new_pending[key] = P
                continue

            success_results[key] = result
            print(f"✅ OK → {name} rows={len(df)}")

        pending = new_pending

        if pending:
            if run_attempt == MAX_FULL_RUN_RETRIES:
                failed_names = [p["name"] for p in pending.values()]
                raise RuntimeError(f"FINAL FAILURE: Properties failed after retries: {failed_names}")

            print(f"🔁 RETRYING ONLY FAILED PROPERTIES after {FULL_RUN_RETRY_DELAY}s...")
            await asyncio.sleep(FULL_RUN_RETRY_DELAY)

    # ================= FINAL VERIFICATION =================
    valid_results = [success_results[k] for k in PROPERTIES.keys() if k in success_results]

    if len(valid_results) != len(PROPERTIES):
        missing = [PROPERTIES[k]["name"] for k in PROPERTIES.keys() if k not in success_results]
        raise RuntimeError(f"DATA INCOMPLETE: Missing properties: {missing}")

    print("✅ DATA VERIFIED — ALL PROPERTIES PRESENT")

    async with aiohttp.ClientSession() as tg_session:
        # ================= PER-PROPERTY MONTHLY REPORTS =================
            # ================= EXCEL CREATION (UNCHANGED) =================
            wb = Workbook()
            wb.remove(wb.active)
            all_dfs = []

            for name, df, total_rooms, inhouse, checkedout, upcoming, cancelled, prop_details, target_collect in valid_results:


                all_dfs.append(df)
                ws = wb.create_sheet(name)

                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)

                beautify(ws)

                ws.append([])
                stats = [
                    ["Total Bookings", int(df["Rooms"].sum())],
                    ["Total OYO Bookings", count(df, "OYO")],
                    ["Walk-in Bookings", count(df, "Walk-in")],
                    ["MMT Bookings", count(df, "MMT")],
                    ["BDC Bookings", count(df, "BDC")],
                    ["Agoda Bookings", count(df, "Agoda")],
                    ["CB Bookings", count(df, "CB")],
                    ["TA Bookings", count(df, "TA")],
                    ["OBA Bookings", count(df, "OBA")],
                    [],
                    ["Total Amount", round(df["Amount"].sum(), 2)],
                    ["Cash Amount", round(df["Cash"].sum(), 2)],
                    ["QR Amount", round(df["QR"].sum(), 2)],
                    ["Online Amount", round(df["Online"].sum(), 2)],
                    ["Discount Amount", round(df["Discount"].sum(), 2)],
                    ["Ballance Amount", round(df["Balance"].sum(), 2)],
                    [],
                    ["OYO Amount", amt(df, "OYO")],
                    ["Walk-in Amount", amt(df, "Walk-in")],
                    ["MMT Amount", amt(df, "MMT")],
                    ["BDC Amount", amt(df, "BDC")],
                    ["Agoda Amount", amt(df, "Agoda")],
                    ["CB Amount", amt(df, "CB")],
                    ["TA Amount", amt(df, "TA")],
                    ["OBA Amount", amt(df, "OBA")]
                ]

                for s in stats:
                    ws.append(s)

                ws.append([])

                total_booked_rooms = int(df["Rooms"].sum())
                total_amt = float(df["Amount"].sum())
                arr = round(total_amt / total_booked_rooms, 2) if total_booked_rooms else 0

                oyo_df = df[df["Booking Source"] == "OYO"]
                oyo_rooms = int(oyo_df["Rooms"].sum())
                oyo_amount = float(oyo_df["Amount"].sum())
                app_arr = round(oyo_amount / oyo_rooms, 2) if oyo_rooms else 0

                effective_total_rooms = total_rooms * target_days

                available_rooms = effective_total_rooms - total_booked_rooms
                occupancy = round((total_booked_rooms / effective_total_rooms) * 100, 2) if effective_total_rooms else 0

        

                ws.append(["Total Rooms", effective_total_rooms])
                ws.append(["Booked Rooms", total_booked_rooms])
                ws.append(["Available Rooms", available_rooms])
                ws.append(["Occupancy", f"{occupancy}%"])
                ws.append(["ARR", arr])
                ws.append(["App ARR", app_arr])

                ws.append([])
                ws.append([])
                add_payment_tables(ws, df, target_collect)
                add_property_details_box(ws, prop_details)

            # ================= CONSOLIDATED =================
            big = pd.concat(all_dfs)
            ws = wb.create_sheet("CONSOLIDATED STATISTICS")
                        # ================= NEW: CONSOLIDATED TARGET COLLECTION =================
            consolidated_target_collect = {"cash": 0.0, "qr": 0.0, "online": 0.0}
            for res in valid_results:
                tc = res[-1]  # target_collect
                consolidated_target_collect["cash"] += float(tc.get("cash", 0))
                consolidated_target_collect["qr"] += float(tc.get("qr", 0))
                consolidated_target_collect["online"] += float(tc.get("online", 0))


            rows = [
                ["Total Bookings", int(big["Rooms"].sum())],
                ["Total OYO Bookings", count(big, "OYO")],
                ["Walk-in Bookings", count(big, "Walk-in")],
                ["MMT Bookings", count(big, "MMT")],
                ["BDC Bookings", count(big, "BDC")],
                ["Agoda Bookings", count(big, "Agoda")],
                ["CB Bookings", count(big, "CB")],
                ["TA Bookings", count(big, "TA")],
                ["OBA Bookings", count(big, "OBA")],
                [],
                ["Total Amount", round(big["Amount"].sum(), 2)],
                ["Cash Amount", round(big["Cash"].sum(), 2)],
                ["QR Amount", round(big["QR"].sum(), 2)],
                ["Online Amount", round(big["Online"].sum(), 2)],
                ["Discount Amount", round(big["Discount"].sum(), 2)],
                ["Ballance Amount", round(big["Balance"].sum(), 2)],
                [],
                ["OYO Amount", amt(big, "OYO")],
                ["Walk-in Amount", amt(big, "Walk-in")],
                ["MMT Amount", amt(big, "MMT")],
                ["BDC Amount", amt(big, "BDC")],
                ["Agoda Amount", amt(big, "Agoda")],
                ["CB Amount", amt(big, "CB")],
                ["TA Amount", amt(big, "TA")],
                ["OBA Amount", amt(big, "OBA")]
            ]

            for r in rows:
                ws.append(r)

            ws.append([])

            grand_total_rooms = sum(r[2] for r in valid_results) * target_days
            total_rooms_booked = int(big["Rooms"].sum())
            total_amt = float(big["Amount"].sum())
            arr = round(total_amt / total_rooms_booked, 2) if total_rooms_booked else 0

            oyo_big = big[big["Booking Source"] == "OYO"]
            oyo_rooms = int(oyo_big["Rooms"].sum())
            oyo_amt = float(oyo_big["Amount"].sum())
            app_arr = round(oyo_amt / oyo_rooms, 2) if oyo_rooms else 0

            available_rooms = grand_total_rooms - total_rooms_booked
            occupancy = round((total_rooms_booked / grand_total_rooms) * 100, 2) if grand_total_rooms else 0

            ws.append(["Total Rooms", grand_total_rooms])
            ws.append(["Booked Rooms", total_rooms_booked])
            ws.append(["Available Rooms", available_rooms])
            ws.append(["Occupancy", f"{occupancy}%"])
            ws.append(["ARR", arr])
            ws.append(["App ARR", app_arr])
            
            blue = PatternFill("solid", fgColor="1F4E78")
            light = PatternFill("solid", fgColor="DDEBF7")
            green = PatternFill("solid", fgColor="C6EFCE")
            yellow = PatternFill("solid", fgColor="FFF4CC")
            red = PatternFill("solid", fgColor="FFC7CE")

            bold_white = Font(color="FFFFFF", bold=True, size=12)
            bold_black = Font(color="000000", bold=True, size=11)
            normal = Font(color="000000", size=11)

            thin = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
             

                        # ================= PROPERTY WISE AMOUNT TABLE =================
            ws.append([])

            start_amt_row = ws.max_row + 1
            ws.merge_cells(start_row=start_amt_row, start_column=1, end_row=start_amt_row, end_column=7)
            cell = ws.cell(row=start_amt_row, column=1)
            cell.value = "PROPERTY WISE AMOUNT"
            cell.fill = blue
            cell.font = bold_white
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin

            amt_headers = ["Property Code", "Cash", "QR", "Online", "Discount", "Balance", "Total Amount"]
            for i, h in enumerate(amt_headers, 1):
                ws.cell(row=start_amt_row + 1, column=i).value = h
                ws.cell(row=start_amt_row + 1, column=i).fill = light
                ws.cell(row=start_amt_row + 1, column=i).font = bold_black
                ws.cell(row=start_amt_row + 1, column=i).border = thin
                ws.cell(row=start_amt_row + 1, column=i).alignment = Alignment(horizontal="center")

            rr = start_amt_row + 2

            tot_cash = tot_qr = tot_online = tot_disc = tot_bal = tot_amt = 0

            for name, df_prop, *_ in valid_results:
                cash = round(df_prop["Cash"].sum(), 2)
                qr = round(df_prop["QR"].sum(), 2)
                online = round(df_prop["Online"].sum(), 2)
                disc = round(df_prop["Discount"].sum(), 2)
                bal = round(df_prop["Balance"].sum(), 2)
                amt_total = round(df_prop["Amount"].sum(), 2)

                tot_cash += cash
                tot_qr += qr
                tot_online += online
                tot_disc += disc
                tot_bal += bal
                tot_amt += amt_total

                vals = [name, cash, qr, online, disc, bal, amt_total]
                for c, v in enumerate(vals, 1):
                    ws.cell(row=rr, column=c).value = v
                    ws.cell(row=rr, column=c).border = thin
                    ws.cell(row=rr, column=c).alignment = Alignment(horizontal="center")
                rr += 1

            # totals row
            totals = ["TOTAL", round(tot_cash,2), round(tot_qr,2), round(tot_online,2),
                      round(tot_disc,2), round(tot_bal,2), round(tot_amt,2)]

            for c, v in enumerate(totals, 1):
                ws.cell(row=rr, column=c).value = v
                ws.cell(row=rr, column=c).fill = yellow
                ws.cell(row=rr, column=c).font = bold_black
                ws.cell(row=rr, column=c).border = thin
                ws.cell(row=rr, column=c).alignment = Alignment(horizontal="center")

            # ================= PROPERTY WISE BOOKINGS TABLE =================
            ws.append([])

            start_book_row = ws.max_row + 1
            ws.merge_cells(start_row=start_book_row, start_column=1, end_row=start_book_row, end_column=10)
            cell = ws.cell(row=start_book_row, column=1)
            cell.value = "PROPERTY WISE BOOKINGS"
            cell.fill = blue
            cell.font = bold_white
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin

            book_headers = ["Property Code","OYO","Walk-in","MMT","BDC","Agoda","CB","TA","OBA","Total"]
            for i, h in enumerate(book_headers, 1):
                ws.cell(row=start_book_row + 1, column=i).value = h
                ws.cell(row=start_book_row + 1, column=i).fill = light
                ws.cell(row=start_book_row + 1, column=i).font = bold_black
                ws.cell(row=start_book_row + 1, column=i).border = thin
                ws.cell(row=start_book_row + 1, column=i).alignment = Alignment(horizontal="center")

            rr = start_book_row + 2

            col_totals = [0]*8

            for name, df_prop, *_ in valid_results:
                vals = [
                    name,
                    count(df_prop,"OYO"),
                    count(df_prop,"Walk-in"),
                    count(df_prop,"MMT"),
                    count(df_prop,"BDC"),
                    count(df_prop,"Agoda"),
                    count(df_prop,"CB"),
                    count(df_prop,"TA"),
                    count(df_prop,"OBA")
                ]

                total_row = sum(vals[1:])
                vals.append(total_row)

                for i in range(1,9):
                    col_totals[i-1] += vals[i]

                for c, v in enumerate(vals, 1):
                    ws.cell(row=rr, column=c).value = v
                    ws.cell(row=rr, column=c).border = thin
                    ws.cell(row=rr, column=c).alignment = Alignment(horizontal="center")
                rr += 1

            total_vals = ["TOTAL"] + col_totals + [sum(col_totals)]

            for c, v in enumerate(total_vals, 1):
                ws.cell(row=rr, column=c).value = v
                ws.cell(row=rr, column=c).fill = yellow
                ws.cell(row=rr, column=c).font = bold_black
                ws.cell(row=rr, column=c).border = thin
                ws.cell(row=rr, column=c).alignment = Alignment(horizontal="center")
          

            # ================= NEW: PROPERTY SCORE TABLE (CONSOLIDATED ONLY) =================
            ws.append([])  # ✅ 1 row space after App ARR

            

            start_row = ws.max_row + 1

            # Header merged A:C
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
            hcell = ws.cell(row=start_row, column=1)
            hcell.value = "PROPERTY SCORE"
            hcell.fill = blue
            hcell.font = bold_white
            hcell.alignment = Alignment(horizontal="center", vertical="center")
            hcell.border = thin
            ws.cell(row=start_row, column=2).border = thin
            ws.cell(row=start_row, column=3).border = thin
            ws.row_dimensions[start_row].height = 20

            # Column headers
            headers = ["Property Code", "Revenue", "Score", "Revenue Loss"]
            for idx, h in enumerate(headers, start=1):
                ws.cell(row=start_row + 1, column=idx).value = h

            for c in [1, 2, 3, 4]:
                cell = ws.cell(row=start_row + 1, column=c)
                cell.fill = light
                cell.font = bold_black
                cell.border = thin
                cell.alignment = Alignment(horizontal="center", vertical="center")

            r = start_row + 2

            # Each property score + revenue loss
            for name, df_prop, total_rooms_prop, *_ in valid_results:
                booked_rooms_prop = int(df_prop["Rooms"].sum()) if not df_prop.empty else 0
                total_rooms_effective = (total_rooms_prop * target_days) if total_rooms_prop else 0
                available_rooms_prop = total_rooms_effective - booked_rooms_prop

                total_amount_prop = float(df_prop["Amount"].sum()) if not df_prop.empty else 0.0

                score = round((booked_rooms_prop / total_rooms_effective) * 100, 2) if total_rooms_effective else 0.0

                # ✅ Revenue Loss = Total Amount * (Available Rooms / Booked Rooms)
                if booked_rooms_prop > 0:
                    revenue_loss = round(total_amount_prop * (available_rooms_prop / booked_rooms_prop), 2)
                else:
                    revenue_loss = 0.0

                ws.cell(row=r, column=1).value = name
                ws.cell(row=r, column=2).value = round(total_amount_prop, 2)   # Revenue
                ws.cell(row=r, column=3).value = f"{score}%"
                ws.cell(row=r, column=4).value = revenue_loss


                # Conditional fill based on score
                if score > 80:
                    row_fill = green
                elif score >= 60:
                    row_fill = yellow
                else:
                    row_fill = red

                for c in [1, 2, 3, 4]:
                    cell = ws.cell(row=r, column=c)
                    cell.fill = row_fill
                    cell.font = normal
                    cell.border = thin
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                r += 1

            ws.append([])  # ✅ 1 row space after table

            
                        # ================= NEW: CONSOLIDATED PAYMENT TABLES =================
            add_payment_tables(ws, big, consolidated_target_collect, title_prefix="CONSOLIDATED — ")


            beautify(ws)

            # ================= SEND EXCEL TO TELEGRAM (NO LOCAL SAVE) =================
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            await send_telegram_excel_buffer(
                buffer,
                filename=f"Bookings_{TF} to {TF}.xlsx",
                caption=f"📊 Daily Bookings Report"
            )


            print("✅ EXCEL SENT TO TELEGRAM (NO LOCAL FILE)")
            return




# ================= RUN =================
if __name__ == "__main__":
    try:
        asyncio.run(main())
    except Exception as e:
        print("SCRIPT CRASHED")
        print(e)
        traceback.print_exc()
        print("SCRIPT CRASHED", e, flush=True)
