# ==============================
# ULTRA FAST ASYNC MULTI PROPERTY AUTOMATION
# DATE-WISE COLLECTION BASED ON PAYMENT CREATED_AT
# FINAL EXCEL: ONLY PAID BOOKINGS (NO PER DAY STAY CALC)
# ==============================

import asyncio
import aiohttp
import pandas as pd
from datetime import datetime, timedelta
import traceback
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import pytz
IST = pytz.timezone("Asia/Kolkata")
now = datetime.now(IST)

MAX_FULL_RUN_RETRIES = 5
FULL_RUN_RETRY_DELAY = 10

PROP_PARALLEL_LIMIT = 3
DETAIL_PARALLEL_LIMIT = 10

prop_semaphore = asyncio.Semaphore(PROP_PARALLEL_LIMIT)

DETAIL_TIMEOUT = 25
ROOMS_TIMEOUT = 25
BATCH_TIMEOUT = 35

TELEGRAM_BOT_TOKEN = "8457091054:AAHNcJeIpf2-ugHbzaFoImlFuN5lxRbcC5Q"
TELEGRAM_CHAT_ID = -5276067828

# ================= PROPERTIES =================
PROPERTIES = {
    1: {"name":"HYD2857","UIF":"eyJlbWFpbCI6Im1vaGRzdWFpZGFobWVkQGdtYWlsLmNvbSIsImFjY2Vzc190b2tlbiI6Ilo1ZUpSMVJiN3FOb3pNNWY0by10YkEiLCJyb2xlIjoiT3duZXIiLCJpZCI6MjAzMzEzMjUyLCJwaG9uZSI6Ijk5ODUyODMzMDYiLCJjb3VudHJ5X2NvZGUiOiIrOTEiLCJkZXZpc2Vfcm9sZSI6Ik93bmVyX1BvcnRhbF9Vc2VyIiwicGhvbmVfdmVyaWZpZWQiOnRydWUsImVtYWlsX3ZlcmlmaWVkIjp0cnVlLCJ1cGRhdGVkX2F0IjoiMTczMjI2MTE0MiIsImZlYXR1cmVzIjp7fSwic3RhdHVzX2NvZGUiOjEwMCwibWlsbGlzX2xlZnRfZm9yX3Bhc3N3b3JkX2V4cGlyeSI6OTQ3MjgyMDk3NDkzLCJhZGRyZXNzSnNvbiI6e319","UUID":"NDFlNWI1ZTQtODFiZC00MWQ1LWIwODAtM2FmMzcwOGYwYmQz","QID":259690},
    2: {"name":"HYD2728","UIF":"eyJlbWFpbCI6ImNoZWYubml0aW5AZ21haWwuY29tIiwiYWNjZXNzX3Rva2VuIjoicEhfUFRkSEVSa2NDTUxYTi15Qk0ydyIsInJvbGUiOiJPd25lciIsImlkIjoyMDQ3MjI0OTMsInBob25lIjoiNjMwOTMzNjMwOSIsImNvdW50cnlfY29kZSI6Iis5MSIsInNleCI6Ik1hbGUiLCJ0ZWFtIjoiTWFya2V0aW5nIiwiZGV2aXNlX3JvbGUiOiJPd25lcl9Qb3J0YWxfVXNlciIsInBob25lX3ZlcmlmaWVkIjp0cnVlLCJlbWFpbF92ZXJpZmllZCI6dHJ1ZSwidXBkYXRlZF9hdCI6IjE3Njk1MTY0MDYiLCJmZWF0dXJlcyI6e30sInN0YXR1c19jb2RlIjoxMDAsIm1pbGxpc19sZWZ0X2Zvcl9wYXNzd29yZF9leHBpcnkiOjk1MDM4NDk3MzkyOSwiYWRkcmVzc0pzb24iOnt9fQ%3D%3D","UUID":"MzgzZWM2MmUtOGJmOC00MjZiLThhY2ItZGFiYWMwNGU5NDQ5","QID":245844},
    3: {"name":"HYD2927","UIF":"eyJlbWFpbCI6InVwcGFsYXNhaTg4QGdtYWlsLmNvbSIsImFjY2Vzc190b2tlbiI6ImNLeVp6WUVFZTR5SDhIMDk3bGJNUUEiLCJyb2xlIjoiT3duZXIiLCJpZCI6MjE2Mzk4NDcwLCJwaG9uZSI6Ijg2ODYwNjY2NjYiLCJjb3VudHJ5X2NvZGUiOiIrOTEiLCJkZXZpc2Vfcm9sZSI6Ik93bmVyX1BvcnRhbF9Vc2VyIiwicGhvbmVfdmVyaWZpZWQiOnRydWUsImVtYWlsX3ZlcmlmaWVkIjp0cnVlLCJ1cGRhdGVkX2F0IjoiMTczNzc4NTUxNCIsImZlYXR1cmVzIjp7fSwic3RhdHVzX2NvZGUiOjEwMCwibWlsbGlzX2xlZnRfZm9yX3Bhc3N3b3JkX2V4cGlyeSI6OTQ1MDE5Mjk0NzUzLCJhZGRyZXNzSnNvbiI6e319","UUID":"N2Y3ZjdiM2ItNGZiMy00MzlmLTk0MTYtNTlkNzRlZjk3MjA3","QID":292909},
    4: {"name":"HYD3030","UIF":"eyJlbWFpbCI6InN2aG90ZWw5OTlAZ21haWwuY29tIiwiYWNjZXNzX3Rva2VuIjoic2x5Y2Fta0pBbU9uZUt2SXlwOWVsUSIsInJvbGUiOiJPd25lciIsImlkIjoyMzU5NTMyNTAsInBob25lIjoiOTk4NTM0NzQ3NiIsImNvdW50cnlfY29kZSI6Iis5MSIsImZpcnN0X25hbWUiOiJHdWVzdCIsInNleCI6Ik1hbGUiLCJ0ZWFtIjoiTWFya2V0aW5nIiwiZGV2aXNlX3JvbGUiOiJPd25lcl9Qb3J0YWxfVXNlciIsInBob25lX3ZlcmlmaWVkIjp0cnVlLCJlbWFpbF92ZXJpZmllZCI6dHJ1ZSwidXBkYXRlZF9hdCI6IjE3NTM4ODQzOTEiLCJmZWF0dXJlcyI6e30sInN0YXR1c19jb2RlIjoxMDAsIm1pbGxpc19sZWZ0X2Zvcl9wYXNzd29yZF9leHBpcnkiOjk0MDI5NDU3MjI0MywiYWRkcmVzc0pzb24iOnt9fQ%3D%3D","UUID":"ZjNjZmZkMWQtOTJiMS00ZjM3LWE1YWMtZGQ3NGExNGIwN2Q5","QID":304236},
    5: {"name":"HYD1170","UIF":"eyJlbWFpbCI6ImtsLmdyYW5kLmhvdGVsQGdtYWlsLmNvbSIsImFjY2Vzc190b2tlbiI6IlhiTVZVUllmVlNJQUhZSWlRMDRyV0EiLCJyb2xlIjoiT3duZXIiLCJpZCI6MjQ2NTU3NzU4LCJwaG9uZSI6IjkyNDgwMDM3MzgiLCJjb3VudHJ5X2NvZGUiOiIrOTEiLCJmaXJzdF9uYW1lIjoiQW5rZXNoIiwic2V4IjoiTWFsZSIsInRlYW0iOiJNYXJrZXRpbmciLCJkZXZpc2Vfcm9sZSI6Ik93bmVyX1BvcnRhbF9Vc2VyIiwicGhvbmVfdmVyaWZpZWQiOnRydWUsImVtYWlsX3ZlcmlmaWVkIjp0cnVlLCJ1cGRhdGVkX2F0IjoiMTc2Mzk3ODgyMCIsImZlYXR1cmVzIjp7fSwic3RhdHVzX2NvZGUiOjEwMCwibWlsbGlzX2xlZnRfZm9yX3Bhc3N3b3JkX2V4cGlyeSI6OTQ0NjkyNTczODQ5LCJhZGRyZXNzSnNvbiI6e319","UUID":"YzRlZWNmMzUtMTllNS00YjVhLTg4YTgtOGIwNGI2NzlkNWQ0","QID":83460},
    6: {"name":"HYD2984","UIF":"eyJlbWFpbCI6InByYXZlZW5hcHV0bHVyaTIwMDdAZ21haWwuY29tIiwiYWNjZXNzX3Rva2VuIjoiZ3FFMVg3RFhDR0RaeEhfQWdMWVpydyIsInJvbGUiOiJPd25lciIsImlkIjoyMTk1ODcyMjQsInBob25lIjoiODcxMjI5NjIxMiIsImNvdW50cnlfY29kZSI6Iis5MSIsImRldmlzZV9yb2xlIjoiT3duZXJfUG9ydGFsX1VzZXIiLCJwaG9uZV92ZXJpZmllZCI6dHJ1ZSwiZW1haWxfdmVyaWZpZWQiOnRydWUsInVwZGF0ZWRfYXQiOiIxNzQzMjQ1Mjc0IiwiZmVhdHVyZXMiOnt9LCJzdGF0dXNfY29kZSI6MTAwLCJtaWxsaXNfbGVmdF9mb3JfcGFzc3dvcmRfZXhwaXJ5Ijo5MjgzNTcxNDY5MDMsImFkZHJlc3NKc29uIjp7fX0%3D","UUID":"ZDY0ODFkMDgtYmVjZi00ZDU5LTgzZWItMmU1Y2U1NjMyMjEy","QID":299149},
    7: {"name":"HYD495","UIF":"eyJlbWFpbCI6Im1hbm9oYXJqb3NoQGdtYWlsLmNvbSIsImFjY2Vzc190b2tlbiI6IjJQMFVURk9lRElKdzZHejA0WlJMTHciLCJyb2xlIjoiT3duZXIiLCJpZCI6NDc0Mjk5MSwicGhvbmUiOiI5OTg1OTk4NTg4IiwiY291bnRyeV9jb2RlIjoiKzkxIiwiZmlyc3RfbmFtZSI6IlZhcmFwcmFzYWRwbXByYXRhcCIsImxhc3RfbmFtZSI6IjgwOTY5OTQ0MjQiLCJjaXR5IjoiIiwic2V4IjoiTWFsZSIsInRlYW0iOiJPd25lciBFbmdhZ2VtZW50IiwiZGV2aXNlX3JvbGUiOiJPd25lcl9Qb3J0YWxfVXNlciIsInBob25lX3ZlcmlmaWVkIjp0cnVlLCJlbWFpbF92ZXJpZmllZCI6dHJ1ZSwiYWRkcmVzcyI6IiIsInVwZGF0ZWRfYXQiOiIxNzYxOTgzODg1IiwiZmVhdHVyZXMiOnt9LCJzdGF0dXNfY29kZSI6MTAwLCJtaWxsaXNfbGVmdF9mb3JfcGFzc3dvcmRfZXhwaXJ5Ijo5NDYwNjAwMzI2MjgsImFkZHJlc3NKc29uIjp7fX0%3D","UUID":"YjAxMWE2MDgtMDc5Ni00OGZlLTliYjEtNDY0OWJkM2IzNzMx","QID":16711},
    8: {"name":"HYD2963","UIF":"eyJlbWFpbCI6InRoaXJ1cGF0aGlyYW90OEBnbWFpbC5jb20iLCJhY2Nlc3NfdG9rZW4iOiJCM1l1U1k1cy0wZE1aeXB1M1l4b2R3Iiwicm9sZSI6Ik93bmVyIiwiaWQiOjExMTEyMjI2MywicGhvbmUiOiI5NTAyMzIzNTEzIiwiY291bnRyeV9jb2RlIjoiKzkxIiwiZmlyc3RfbmFtZSI6InRhbmRyYSIsImxhc3RfbmFtZSI6InRpcnVwYXRoaXJhbyIsImNpdHkiOiIiLCJzZXgiOiJNYWxlIiwidGVhbSI6IlRyYXZlbCBBZ2VudCIsImRldmlzZV9yb2xlIjoiT3duZXJfUG9ydGFsX1VzZXIiLCJwaG9uZV92ZXJpZmllZCI6dHJ1ZSwiZW1haWxfdmVyaWZpZWQiOnRydWUsImFkZHJlc3MiOiIiLCJ1cGRhdGVkX2F0IjoiMTY2NjA5OTMzMyIsImZlYXR1cmVzIjp7fSwic3RhdHVzX2NvZGUiOjEwMCwibWlsbGlzX2xlZnRfZm9yX3Bhc3N3b3JkX2V4cGlyeSI6OTUwMzcxODg3NDAyLCJhZGRyZXNzSnNvbiI6e319","UUID":"YjdlYTZhNDItZGNlNi00NGFhLWI5YzgtODkzZjIyYmM2N2Ri","QID":296969},
    9: {"name":"HYD2012","UIF":"eyJlbWFpbCI6InRoaXJ1cGF0aGlyYW90OEBnbWFpbC5jb20iLCJhY2Nlc3NfdG9rZW4iOiJCM1l1U1k1cy0wZE1aeXB1M1l4b2R3Iiwicm9sZSI6Ik93bmVyIiwiaWQiOjExMTEyMjI2MywicGhvbmUiOiI5NTAyMzIzNTEzIiwiY291bnRyeV9jb2RlIjoiKzkxIiwiZmlyc3RfbmFtZSI6InRhbmRyYSIsImxhc3RfbmFtZSI6InRpcnVwYXRoaXJhbyIsImNpdHkiOiIiLCJzZXgiOiJNYWxlIiwidGVhbSI6IlRyYXZlbCBBZ2VudCIsImRldmlzZV9yb2xlIjoiT3duZXJfUG9ydGFsX1VzZXIiLCJwaG9uZV92ZXJpZmllZCI6dHJ1ZSwiZW1haWxfdmVyaWZpZWQiOnRydWUsImFkZHJlc3MiOiIiLCJ1cGRhdGVkX2F0IjoiMTY2NjA5OTMzMyIsImZlYXR1cmVzIjp7fSwic3RhdHVzX2NvZGUiOjEwMCwibWlsbGlzX2xlZnRfZm9yX3Bhc3N3b3JkX2V4cGlyeSI6OTUwMzcxODg3NDAyLCJhZGRyZXNzSnNvbiI6e319","UUID":"YjdlYTZhNDItZGNlNi00NGFhLWI5YzgtODkzZjIyYmM2N2Ri","QID":196450},
    10: {"name":"HYD1498","UIF":"eyJlbWFpbCI6InRoaXJ1cGF0aGlyYW90OEBnbWFpbC5jb20iLCJhY2Nlc3NfdG9rZW4iOiJCM1l1U1k1cy0wZE1aeXB1M1l4b2R3Iiwicm9sZSI6Ik93bmVyIiwiaWQiOjExMTEyMjI2MywicGhvbmUiOiI5NTAyMzIzNTEzIiwiY291bnRyeV9jb2RlIjoiKzkxIiwiZmlyc3RfbmFtZSI6InRhbmRyYSIsImxhc3RfbmFtZSI6InRpcnVwYXRoaXJhbyIsImNpdHkiOiIiLCJzZXgiOiJNYWxlIiwidGVhbSI6IlRyYXZlbCBBZ2VudCIsImRldmlzZV9yb2xlIjoiT3duZXJfUG9ydGFsX1VzZXIiLCJwaG9uZV92ZXJpZmllZCI6dHJ1ZSwiZW1haWxfdmVyaWZpZWQiOnRydWUsImFkZHJlc3MiOiIiLCJ1cGRhdGVkX2F0IjoiMTY2NjA5OTMzMyIsImZlYXR1cmVzIjp7fSwic3RhdHVzX2NvZGUiOjEwMCwibWlsbGlzX2xlZnRfZm9yX3Bhc3N3b3JkX2V4cGlyeSI6OTUwMzcxODg3NDAyLCJhZGRyZXNzSnNvbiI6e319","UUID":"dlYTZhNDItZGNlNi00NGFhLWI5YzgtODkzZjIyYmM2N2Ri","QID":105249},
    11: {"name":"HYD3183","UIF":"eyJlbWFpbCI6ImthbWFsYWFjaGFAZ21haWwuY29tIiwiYWNjZXNzX3Rva2VuIjoia2RQTVZhV3ZVaGg1cTVaeTMxN3pKUSIsInJvbGUiOiJPd25lciIsImlkIjoyMTg0ODczNjEsInBob25lIjoiOTM5MTA0NDA3MSIsImNvdW50cnlfY29kZSI6Iis5MSIsImRldmlzZV9yb2xlIjoiT3duZXJfUG9ydGFsX1VzZXIiLCJwaG9uZV92ZXJpZmllZCI6dHJ1ZSwiZW1haWxfdmVyaWZpZWQiOnRydWUsInVwZGF0ZWRfYXQiOiIxNzQwNjUyMjIwIiwiZmVhdHVyZXMiOnt9LCJzdGF0dXNfY29kZSI6MTAwLCJtaWxsaXNfbGVmdF9mb3JfcGFzc3dvcmRfZXhwaXJ5Ijo5NDA0NjU5NjU0MjYsImFkZHJlc3NKc29uIjp7fX0%3D","UUID":"YzA1YmE5ODItY2RhMy00MDhiLTk1NzQtNzMzMDA0NTZiM2Yw","QID":328327},
    12: {"name":"HYD1090","UIF":"eyJlbWFpbCI6InNoYW50aGFyZXNpZGVuY3lsb2RnZUBnbWFpbC5jb20iLCJhY2Nlc3NfdG9rZW4iOiJMV1d3VmxHOFhwRHVZQnBySXpkQkhnIiwicm9sZSI6Ik93bmVyIiwiaWQiOjIyMzI4MjUzNCwicGhvbmUiOiI4NTIwMDA1NDc5IiwiY291bnRyeV9jb2RlIjoiKzkxIiwiZmlyc3RfbmFtZSI6Ikd1ZXN0Iiwic2V4IjoiTWFsZSIsInRlYW0iOiJNYXJrZXRpbmciLCJkZXZpc2Vfcm9sZSI6Ik93bmVyX1BvcnRhbF9Vc2VyIiwicGhvbmVfdmVyaWZpZWQiOnRydWUsImVtYWlsX3ZlcmlmaWVkIjp0cnVlLCJ1cGRhdGVkX2F0IjoiMTc0OTIxMjg3NyIsImZlYXR1cmVzIjp7fSwic3RhdHVzX2NvZGUiOjEwMCwibWlsbGlzX2xlZnRfZm9yX3Bhc3N3b3JkX2V4cGlyeSI6OTMzNzUwNTgwMzk5LCJhZGRyZXNzSnNvbiI6e319","UUID":"Zjg4NDc3ZjgtMzM5Zi00ZmYwLWE2OGItYjdkMDEyOGQzNWJk","QID":78637},
    13: {"name":"HYD1762","UIF":"eyJlbWFpbCI6ImtlZXJ0aGljaGFuZHJhOTJAeWFob28uY29tIiwiYWNjZXNzX3Rva2VuIjoiUVF2QURDVmY3R3ZrUFB3Q3Q4SldNQSIsInJvbGUiOiJPd25lciIsImlkIjoxMTA1NjkzOTUsInBob25lIjoiOTk1OTY2NjYwMiIsImNvdW50cnlfY29kZSI6Iis5MSIsImZpcnN0X25hbWUiOiJCYW5kYXJ1IiwibGFzdF9uYW1lIjoiVmVua2F0YXNhdHlha2VlcnRoaSIsImNpdHkiOiIiLCJzZXgiOiJNYWxlIiwidGVhbSI6Ik93bmVyIEVuZ2FnZW1lbnQiLCJkZXZpc2Vfcm9sZSI6Ik93bmVyX1BvcnRhbF9Vc2VyIiwicGhvbmVfdmVyaWZpZWQiOnRydWUsImVtYWlsX3ZlcmlmaWVkIjp0cnVlLCJhZGRyZXNzIjoiIiwidXBkYXRlZF9hdCI6IjE3MDczOTk3NTIiLCJmZWF0dXJlcyI6e30sInN0YXR1c19jb2RlIjoxMDAsIm1pbGxpc19sZWZ0X2Zvcl9wYXNzd29yZF9leHBpcnkiOjk1MDM2MDg0MjM0NiwiYWRkcmVzc0pzb24iOnt9fQ%3D%3D","UUID":"M2Q4MzgxMmYtYzlhMS00NDVlLTk3MzUtZmFjMmQ3ODc0YTEx","QID":115451},
    14: {"name":"HYD588","UIF":"eyJlbWFpbCI6ImtlZXJ0aGljaGFuZHJhOTJAeWFob28uY29tIiwiYWNjZXNzX3Rva2VuIjoiUVF2QURDVmY3R3ZrUFB3Q3Q4SldNQSIsInJvbGUiOiJPd25lciIsImlkIjoxMTA1NjkzOTUsInBob25lIjoiOTk1OTY2NjYwMiIsImNvdW50cnlfY29kZSI6Iis5MSIsImZpcnN0X25hbWUiOiJCYW5kYXJ1IiwibGFzdF9uYW1lIjoiVmVua2F0YXNhdHlha2VlcnRoaSIsImNpdHkiOiIiLCJzZXgiOiJNYWxlIiwidGVhbSI6Ik93bmVyIEVuZ2FnZW1lbnQiLCJkZXZpc2Vfcm9sZSI6Ik93bmVyX1BvcnRhbF9Vc2VyIiwicGhvbmVfdmVyaWZpZWQiOnRydWUsImVtYWlsX3ZlcmlmaWVkIjp0cnVlLCJhZGRyZXNzIjoiIiwidXBkYXRlZF9hdCI6IjE3MDczOTk3NTIiLCJmZWF0dXJlcyI6e30sInN0YXR1c19jb2RlIjoxMDAsIm1pbGxpc19sZWZ0X2Zvcl9wYXNzd29yZF9leHBpcnkiOjk1MDM2MDg0MjM0NiwiYWRkcmVzc0pzb24iOnt9fQ%3D%3D","UUID":"M2Q4MzgxMmYtYzlhMS00NDVlLTk3MzUtZmFjMmQ3ODc0YTEx","QID":37182},
    15: {"name":"WAR144","UIF":"eyJlbWFpbCI6InZpc2hudWdyYW5kLmhhbmFta29uZGFAZ21haWwuY29tIiwiYWNjZXNzX3Rva2VuIjoiSUp5Q2dScWVBUHRrT1czMWRRcTJpZyIsInJvbGUiOiJPd25lciIsImlkIjoyMzcwNDQ0MjgsInBob25lIjoiNjMwMTg4ODg0MyIsImNvdW50cnlfY29kZSI6Iis5MSIsImRldmlzZV9yb2xlIjoiT3duZXJfUG9ydGFsX1VzZXIiLCJwaG9uZV92ZXJpZmllZCI6dHJ1ZSwiZW1haWxfdmVyaWZpZWQiOnRydWUsInVwZGF0ZWRfYXQiOiIxNzU0NTQ5MjEyIiwiZmVhdHVyZXMiOnt9LCJzdGF0dXNfY29kZSI6MTAwLCJtaWxsaXNfbGVmdF9mb3JfcGFzc3dvcmRfZXhwaXJ5Ijo5Mzg3MTc2NDI1MjgsImFkZHJlc3NKc29uIjp7fX0%3D","UUID":"OWRhOTk1MjItNzZlMy00ZjkwLWFhODMtN2U3NTM1YzE4YzZi","QID":326437},
    16: {"name":"KMM030","UIF":"eyJlbWFpbCI6ImJsdWVtb29uaG90ZWwyNEBnbWFpbC5jb20iLCJhY2Nlc3NfdG9rZW4iOiJaRUtKbzBGWUpUNWROYWplOS1ocV9nIiwicm9sZSI6Ik93bmVyIiwiaWQiOjIwMzc1ODk1MywicGhvbmUiOiI5MTAwNzE4Mzg3IiwiY291bnRyeV9jb2RlIjoiKzkxIiwiZGV2aXNlX3JvbGUiOiJPd25lcl9Qb3J0YWxfVXNlciIsInBob25lX3ZlcmlmaWVkIjp0cnVlLCJlbWFpbF92ZXJpZmllZCI6dHJ1ZSwidXBkYXRlZF9hdCI6IjE3MjEzOTEzMzkiLCJmZWF0dXJlcyI6e30sInN0YXR1c19jb2RlIjoxMDAsIm1pbGxpc19sZWZ0X2Zvcl9wYXNzd29yZF9leHBpcnkiOjkyODMzNzE4MDMzMywiYWRkcmVzc0pzb24iOnt9fQ%3D%3D","UUID":"NzE2MGQxMDctNDliNS00YWE5LWI4MGMtY2E0ODQ1ZmZmNGIx","QID":244631},
    17: {'name':"NGA028","UIF":"eyJlbWFpbCI6ImtzYW5qZWV2YTlAZ21haWwuY29tIiwiYWNjZXNzX3Rva2VuIjoiX3FQZFdWSjNTeHNINVE3ZGs0S05xdyIsInJvbGUiOiJPd25lciIsImlkIjo3MjA4MjY4OCwicGhvbmUiOiI4NDk5ODgzMzExIiwiY291bnRyeV9jb2RlIjoiKzkxIiwiZmlyc3RfbmFtZSI6IkthbXNhbmkiLCJsYXN0X25hbWUiOiJTYW5qZWV2YSIsInRlYW0iOiJPcGVyYXRpb25zIiwiZGV2aXNlX3JvbGUiOiJPd25lcl9Qb3J0YWxfVXNlciIsInBob25lX3ZlcmlmaWVkIjp0cnVlLCJlbWFpbF92ZXJpZmllZCI6dHJ1ZSwidXBkYXRlZF9hdCI6IjE3NjQ3NTc5NjIiLCJmZWF0dXJlcyI6e30sInN0YXR1c19jb2RlIjoxMDAsIm1pbGxpc19sZWZ0X2Zvcl9wYXNzd29yZF9leHBpcnkiOjk0NzQyMTczMzQzMSwiYWRkcmVzc0pzb24iOnt9fQ%3D%3D","UUID":"NzRkNjcyMmEtNTU5Ni00NWM0LTk3NjQtNmFkZTVjODE5YjQ2","QID": 353264},
}
# ================= TELEGRAM =================
async def send_telegram_excel_buffer(buffer, filename, caption=None):
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"

    data = aiohttp.FormData()
    data.add_field("chat_id", str(TELEGRAM_CHAT_ID))
    if caption:
        data.add_field("caption", caption)

    data.add_field(
        "document",
        buffer,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    async with aiohttp.ClientSession() as session:
        async with session.post(url, data=data, timeout=120) as resp:
            if resp.status != 200:
                text = await resp.text()
                raise RuntimeError(f"Telegram send failed: {text}")

# ================= BEAUTIFY EXCEL =================
def beautify(ws):
    blue = PatternFill("solid", fgColor="1F4E78")
    light1 = PatternFill("solid", fgColor="DDEBF7")
    light2 = PatternFill("solid", fgColor="F2F2F2")
    yellow = PatternFill("solid", fgColor="FFF4CC")

    bold_white = Font(color="FFFFFF", bold=True, size=12)
    bold_black = Font(color="000000", bold=True, size=12)

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    max_row = ws.max_row
    max_col = ws.max_column
    ws.freeze_panes = "A2"

    for col in range(1, max_col + 1):
        c = ws.cell(row=1, column=col)
        c.fill = blue
        c.font = bold_white
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin

    for r in range(2, max_row + 1):
        fill = light1 if r % 2 == 0 else light2
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue

            if cell.fill is not None and cell.fill.patternType is not None:
                cell.border = thin
                continue

            cell.fill = fill
            cell.border = thin

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 5

    for r in range(2, max_row + 1):
        text = str(ws.cell(row=r, column=1).value or "")
        if text.strip() == "":
            continue
        if "Booking" in text or "Amount" in text or "Total" in text or "OYO" in text:
            ws.cell(row=r, column=1).fill = yellow
            ws.cell(row=r, column=1).font = bold_black

# ================= BOOKING SOURCE =================
def get_booking_source(b):
    source = str(b.get("source", "") or "").strip()
    ota = str(b.get("ota_source", "") or "").strip()
    sub = str(b.get("sub_source", "") or "").strip()
    corp = bool(b.get("is_corporate", False))
    booking_identifier = str(b.get("booking_identifier", "") or "").strip()

    if booking_identifier == "TA":
        return "TA"
    if source == "Walk In":
        return "Walk-in"
    if corp or sub == "corporate":
        return "CB"
    if "Booking.com" in ota:
        return "BDC"
    if "GoMMT" in ota:
        return "MMT"
    if "Agoda" in ota:
        return "Agoda"
    if source == "Travel Agent" or sub == "TPO":
        return "TA"
    if source in ["Android App","IOS App","Web Booking","Mobile Web Booking","Website Booking","Direct"]:
        return "OYO"
    return "OBA"

# ================= PREMIUM PROPERTY DETAILS BOX =================
def add_property_details_box(ws, prop):
    blue = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="DDEBF7")
    white = PatternFill("solid", fgColor="FFFFFF")

    bold_white = Font(color="FFFFFF", bold=True, size=12)
    bold_black = Font(color="000000", bold=True, size=11)
    normal = Font(color="000000", size=11)
    link_font = Font(color="0563C1", underline="single", bold=True, size=11)

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    start_col = 1
    end_col = 8

    def _border_range(r1, c1, r2, c2):
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(row=rr, column=cc).border = thin

    def _merge(row, c1, c2, value, fill=None, font=None, center=False, wrap=False):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1)
        cell.value = value
        if fill: cell.fill = fill
        if font: cell.font = font
        cell.alignment = Alignment(horizontal="center" if center else "left",
                                   vertical="center", wrap_text=wrap)
        return cell

    plot = (prop.get("plot_number") or "").strip()
    street = (prop.get("street") or "").strip()
    pincode = (prop.get("pincode") or "").strip()
    city = (prop.get("city") or "").strip()
    country = (prop.get("country") or "").strip()

    address_parts = []
    if plot: address_parts.append(plot)
    if street: address_parts.append(street)
    city_pin = " ".join([x for x in [city, pincode] if x]).strip()
    if city_pin: address_parts.append(city_pin)
    if country: address_parts.append(country)
    address = ", ".join(address_parts).strip()

    ws.append([])
    ws.append([])
    top = ws.max_row + 1

    _merge(top, start_col, end_col, "PROPERTY DETAILS", fill=blue, font=bold_white, center=True)
    ws.row_dimensions[top].height = 22

    _merge(top + 1, 1, 2, "Name", fill=light, font=bold_black, wrap=True)
    _merge(top + 1, 3, end_col, prop.get("name", "") or "", fill=white, font=normal, wrap=True)

    _merge(top + 2, 1, 2, "Alternative Name", fill=light, font=bold_black, wrap=True)
    _merge(top + 2, 3, end_col, prop.get("alternate_name", "") or "", fill=white, font=normal, wrap=True)

    _merge(top + 3, 1, 2, "Address", fill=light, font=bold_black, wrap=True)
    _merge(top + 3, 3, end_col, address, fill=white, font=normal, wrap=True)
    ws.row_dimensions[top + 3].height = 45

    _merge(top + 4, 1, 2, "Google Map", fill=light, font=bold_black, wrap=True)
    map_link = (prop.get("map_link") or "").strip() or ""
    link_cell = _merge(top + 4, 3, end_col,
                       "OPEN IN GOOGLE MAPS" if map_link else "",
                       fill=white, font=link_font, center=True)
    if map_link:
        link_cell.hyperlink = map_link

    _border_range(top, start_col, top + 4, end_col)

# ================= PREMIUM PAYMENT TABLES =================
def add_payment_tables(ws, df, daily_collect, TF, TT, title_prefix=""):
    blue = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="DDEBF7")
    white = PatternFill("solid", fgColor="FFFFFF")
    yellow = PatternFill("solid", fgColor="FFF4CC")

    bold_white = Font(color="FFFFFF", bold=True, size=12)
    bold_black = Font(color="000000", bold=True, size=11)
    normal = Font(color="000000", size=11)

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def _style_row(row, start_col, end_col, fill=None, font=None, center=True):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=c)
            if fill: cell.fill = fill
            if font: cell.font = font
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if center else "left",
                                       vertical="center")

    def _merge(row, c1, c2, value, fill=None, font=None, center=True):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1)
        cell.value = value
        if fill: cell.fill = fill
        if font: cell.font = font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center" if center else "left",
                                   vertical="center")
        return cell

    start_col = 1
    end_col = 7

    premium_widths = [18, 14, 14, 14, 14, 14, 16]
    for i, w in enumerate(premium_widths, start=1):
        col_letter = get_column_letter(i)
        current = ws.column_dimensions[col_letter].width
        ws.column_dimensions[col_letter].width = max(current or 0, w)

    ws.append([])

    # ================= TABLE 1: Booking Source vs Payment Mode =================
    top = ws.max_row + 1
    heading = f"{title_prefix}BOOKING SOURCE × PAYMENT MODE".strip()
    _merge(top, start_col, end_col, heading, fill=blue, font=bold_white, center=True)

    headers = ["Source", "Cash", "QR", "Online", "Total Paid", "", ""]
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=top + 1, column=idx).value = h
    _style_row(top + 1, start_col, end_col, fill=light, font=bold_black)

    sources = ["OYO", "Walk-in", "MMT", "BDC", "Agoda", "CB", "TA", "OBA"]
    r = top + 2

    for src in sources:
        part = df[df["Booking Source"] == src] if (not df.empty and "Booking Source" in df.columns) else df

        cash = round(float(part["Cash"].sum()), 2) if (not part.empty and "Cash" in part.columns) else 0
        qr = round(float(part["QR"].sum()), 2) if (not part.empty and "QR" in part.columns) else 0
        online = round(float(part["Online"].sum()), 2) if (not part.empty and "Online" in part.columns) else 0
        total_paid = round(cash + qr + online, 2)

        ws.cell(row=r, column=1).value = src
        ws.cell(row=r, column=2).value = cash
        ws.cell(row=r, column=3).value = qr
        ws.cell(row=r, column=4).value = online
        ws.cell(row=r, column=5).value = total_paid

        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin
            cell.font = normal
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = white

        ws.cell(row=r, column=1).fill = yellow
        ws.cell(row=r, column=1).font = bold_black
        r += 1

    # TOTAL row
    tot_cash = round(float(df["Cash"].sum()), 2) if not df.empty else 0
    tot_qr = round(float(df["QR"].sum()), 2) if not df.empty else 0
    tot_online = round(float(df["Online"].sum()), 2) if not df.empty else 0
    tot_paid = round(tot_cash + tot_qr + tot_online, 2)

    ws.cell(row=r, column=1).value = "TOTAL"
    ws.cell(row=r, column=2).value = tot_cash
    ws.cell(row=r, column=3).value = tot_qr
    ws.cell(row=r, column=4).value = tot_online
    ws.cell(row=r, column=5).value = tot_paid

    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = thin
        cell.font = bold_black
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = light

    ws.cell(row=r, column=1).fill = yellow
    ws.cell(row=r, column=1).font = bold_black

    ws.append([])

    # ================= TABLE 2: DATE WISE COLLECTION SUMMARY =================
    top2 = ws.max_row + 1
    heading2 = f"{title_prefix}DATE WISE COLLECTION SUMMARY".strip()
    _merge(top2, start_col, end_col, heading2, fill=blue, font=bold_white, center=True)

    headers2 = ["Date", "Cash", "QR", "Online", "Total Paid", "", ""]
    for idx, h in enumerate(headers2, start=1):
        ws.cell(row=top2 + 1, column=idx).value = h
    _style_row(top2 + 1, start_col, end_col, fill=light, font=bold_black)

    rr = top2 + 2
    tf_dt = datetime.strptime(TF, "%Y-%m-%d").date()
    tt_dt = datetime.strptime(TT, "%Y-%m-%d").date()

    grand_cash = grand_qr = grand_online = 0.0

    cur = tf_dt
    while cur <= tt_dt:
        dkey = cur.strftime("%Y-%m-%d")
        vals = daily_collect.get(dkey, {"cash": 0.0, "qr": 0.0, "online": 0.0})

        cash = round(float(vals.get("cash", 0)), 2)
        qr = round(float(vals.get("qr", 0)), 2)
        online = round(float(vals.get("online", 0)), 2)
        total_paid = round(cash + qr + online, 2)

        ws.cell(row=rr, column=1).value = f"{dkey}"
        ws.cell(row=rr, column=2).value = cash
        ws.cell(row=rr, column=3).value = qr
        ws.cell(row=rr, column=4).value = online
        ws.cell(row=rr, column=5).value = total_paid

        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=rr, column=c)
            cell.border = thin
            cell.font = normal
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = white

        ws.cell(row=rr, column=1).fill = yellow
        ws.cell(row=rr, column=1).font = bold_black

        grand_cash += cash
        grand_qr += qr
        grand_online += online

        rr += 1
        cur += timedelta(days=1)

    grand_total = round(grand_cash + grand_qr + grand_online, 2)

    ws.cell(row=rr, column=1).value = "TOTAL"
    ws.cell(row=rr, column=2).value = round(grand_cash, 2)
    ws.cell(row=rr, column=3).value = round(grand_qr, 2)
    ws.cell(row=rr, column=4).value = round(grand_online, 2)
    ws.cell(row=rr, column=5).value = grand_total

    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=rr, column=c)
        cell.border = thin
        cell.font = bold_black
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = light

    ws.cell(row=rr, column=1).fill = yellow
    ws.cell(row=rr, column=1).font = bold_black

    ws.append([])

# ================= FETCH DETAILS =================
async def fetch_booking_details(session, P, booking_no):
    url = "https://www.oyoos.com/hms_ms/api/v1/visibility/booking_details_with_entities"
    params = {
        "qid": P["QID"],
        "booking_id": booking_no,
        "role": 0,
        "platform": "OYOOS",
        "country_code": 1
    }
    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}
    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    for attempt in range(1, 4):
        try:
            async with session.get(url, params=params, headers=headers, cookies=cookies, timeout=DETAIL_TIMEOUT) as r:
                if r.status != 200:
                    raise RuntimeError("DETAIL API FAILED")

                data = await r.json()
                booking = next(iter(data.get("entities", {}).get("bookings", {}).values()), {})
                payments = booking.get("payments", [])

                payment_events = []  # [{"date":"YYYY-MM-DD","mode":"cash/qr/online","amt":x}]

                for p in payments:
                    mode = p.get("mode", "")
                    amt = float(p.get("amount", 0) or 0)

                    created_at = str(p.get("created_at") or "").strip()
                    pay_date = created_at.split("T")[0].strip() if created_at else ""

                    if mode == "oyo_wizard_discount":
                        continue

                    if mode == "Cash at Hotel":
                        if pay_date:
                            payment_events.append({"date": pay_date, "mode": "cash", "amt": amt})
                    elif mode == "UPI QR":
                        if pay_date:
                            payment_events.append({"date": pay_date, "mode": "qr", "amt": amt})
                    else:
                        if pay_date:
                            payment_events.append({"date": pay_date, "mode": "online", "amt": amt})

                return payment_events

        except Exception:
            await asyncio.sleep(2 + attempt)

    raise RuntimeError("DETAIL FETCH FAILED")

# ================= BATCH FETCH =================
async def fetch_bookings_batch(session, offset, f, t, P):
    url = "https://www.oyoos.com/hms_ms/api/v1/get_booking_with_ids"
    params = {
        "qid": P["QID"],
        "checkin_from": f,
        "checkin_till": t,
        "batch_count": 100,
        "batch_offset": offset,
        "visibility_required": "true",
        "additionalParams": "payment_hold_transaction,guest,stay_details",
        "decimal_price": "true",
        "ascending": "true",
        "sort_on": "checkin_date"
    }
    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}
    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    async with session.get(url, params=params, cookies=cookies, headers=headers, timeout=BATCH_TIMEOUT) as r:
        if r.status != 200:
            raise RuntimeError("BATCH API FAILED")
        return await r.json()

# ================= PROPERTY DETAILS API =================
async def fetch_property_details(session, P):
    url = "https://www.oyoos.com/hms_ms/api/v1/location/property-details"
    params = {"qid": P["QID"]}
    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}
    headers = {"accept": "application/json", "x-qid": str(P["QID"]), "x-source-client": "merchant"}

    for attempt in range(1, 4):
        try:
            async with session.get(url, params=params, cookies=cookies, headers=headers, timeout=20) as r:
                if r.status != 200:
                    raise RuntimeError(f"PROPERTY DETAILS API FAILED ({r.status})")
                data = await r.json()
                return {
                    "name": str(data.get("name", "") or "").strip(),
                    "alternate_name": str(data.get("alternate_name", "") or "").strip(),
                    "plot_number": str(data.get("plot_number", "") or "").strip(),
                    "street": str(data.get("street", "") or "").strip(),
                    "pincode": str(data.get("pincode", "") or "").strip(),
                    "city": str(data.get("city", "") or "").strip(),
                    "country": str(data.get("country", "") or "").strip(),
                    "map_link": str(data.get("map_link", "") or "").strip(),
                }
        except Exception:
            await asyncio.sleep(2 + attempt)

    return {"name":"","alternate_name":"","plot_number":"","street":"","pincode":"","city":"","country":"","map_link":""}

# ================= FETCH TOTAL ROOMS =================
async def fetch_total_rooms(session, P):
    url = "https://www.oyoos.com/hms_ms/api/v1/hotels/roomsNew"
    params = {"qid": P["QID"]}
    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}
    headers = {"accept": "application/json", "x-qid": str(P["QID"]), "x-source-client": "merchant"}

    for attempt in range(1, 4):
        try:
            async with session.get(url, params=params, cookies=cookies, headers=headers, timeout=ROOMS_TIMEOUT) as r:
                if r.status != 200:
                    raise RuntimeError("ROOM API FAILED")
                data = await r.json()
                rooms = data.get("rooms", {})
                return len(rooms)
        except Exception:
            await asyncio.sleep(2 + attempt)

    return 0

# ================= PROCESS PROPERTY =================
async def process_property(P, TF, TT, HF, HT):
    print(f"PROCESSING → {P['name']}")

    tf_dt = datetime.strptime(TF, "%Y-%m-%d").date()
    tt_dt = datetime.strptime(TT, "%Y-%m-%d").date()

    async with aiohttp.ClientSession() as session:
        total_rooms = await fetch_total_rooms(session, P)
        prop_details = await fetch_property_details(session, P)

        detail_semaphore = asyncio.Semaphore(DETAIL_PARALLEL_LIMIT)
        detail_cache = {}

        async def limited_detail_call(booking_no):
            async with detail_semaphore:
                if booking_no in detail_cache:
                    return detail_cache[booking_no]
                res = await fetch_booking_details(session, P, booking_no)
                detail_cache[booking_no] = res
                return res

        daily_collect = {}
        booking_date_mode_map = {}  # (date, booking_no) -> {"cash":x,"qr":x,"online":x}

        offset = 0
        while True:
            data = await fetch_bookings_batch(session, offset, HF, HT, P)
            if not data or not data.get("bookingIds"):
                break

            bookings = data.get("entities", {}).get("bookings", {})
            if not bookings:
                break

            tasks = []
            mapping = []

            for b in bookings.values():
                status = (b.get("status") or "").strip()
                if status not in ["Checked In", "Checked Out"]:
                    continue

                booking_no = b.get("booking_no")
                if not booking_no:
                    continue

                tasks.append(limited_detail_call(booking_no))
                mapping.append(b)

            results = await asyncio.gather(*tasks, return_exceptions=True)

            for res, b in zip(results, mapping):
                if isinstance(res, Exception):
                    continue

                payment_events = res or []
                booking_no = b.get("booking_no")
                source = get_booking_source(b)

                for ev in payment_events:
                    d = ev.get("date")
                    if not d:
                        continue

                    try:
                        d_dt = datetime.strptime(d, "%Y-%m-%d").date()
                    except Exception:
                        continue

                    if not (tf_dt <= d_dt <= tt_dt):
                        continue

                    if d not in daily_collect:
                        daily_collect[d] = {"cash": 0.0, "qr": 0.0, "online": 0.0}

                    if (d, booking_no) not in booking_date_mode_map:
                        booking_date_mode_map[(d, booking_no)] = {"cash": 0.0, "qr": 0.0, "online": 0.0, "b": b, "source": source}

                    mode = ev.get("mode")
                    amt = float(ev.get("amt", 0) or 0)

                    if mode == "cash":
                        daily_collect[d]["cash"] += amt
                        booking_date_mode_map[(d, booking_no)]["cash"] += amt
                    elif mode == "qr":
                        daily_collect[d]["qr"] += amt
                        booking_date_mode_map[(d, booking_no)]["qr"] += amt
                    else:
                        daily_collect[d]["online"] += amt
                        booking_date_mode_map[(d, booking_no)]["online"] += amt

            if len(data["bookingIds"]) < 100:
                break
            offset += 100

        # Build DF rows
        all_rows = []
        for (d, booking_no), vals in booking_date_mode_map.items():
            b = vals["b"]
            cash = vals["cash"]
            qr = vals["qr"]
            online = vals["online"]
            total_paid = cash + qr + online

            all_rows.append({
                "Date": d,
                "Booking Id": booking_no,
                "Guest Name": b.get("guest_name"),
                "Status": b.get("status"),
                "Booking Source": vals["source"],
                "Check In": b.get("checkin"),
                "Check Out": b.get("checkout"),
                "Cash": round(cash, 2),
                "QR": round(qr, 2),
                "Online": round(online, 2),
                "Total Paid": round(total_paid, 2),
            })

        df = pd.DataFrame(all_rows)

        if df.empty:
            df = pd.DataFrame(columns=[
                "Date", "Booking Id", "Guest Name", "Status", "Booking Source",
                "Check In", "Check Out", "Cash", "QR", "Online", "Total Paid"
            ])

        df = df.sort_values(["Date", "Booking Id"], ascending=True)

        return (P["name"], df, total_rooms, prop_details, daily_collect)

# ================= RETRY =================
async def run_property_with_retry(P, TF, TT, HF, HT, retries=3):
    last_error = None
    for attempt in range(1, retries + 1):
        try:
            return await process_property(P, TF, TT, HF, HT)
        except Exception as e:
            last_error = e
            print(f"RETRY {attempt}/{retries} → {P['name']} :: {e}")
            await asyncio.sleep(2 + attempt * 2)
    raise RuntimeError(f"PROPERTY FAILED → {P['name']}") from last_error

async def run_property_limited(P, TF, TT, HF, HT):
    async with prop_semaphore:
        return await run_property_with_retry(P, TF, TT, HF, HT)

# ================= MAIN =================
async def main():
    print("========================================")
    print(" OYO MONTHLY TELEGRAM AUTOMATION")
    print("========================================")

    global now
    now = datetime.now(IST)


    # ================= BUSINESS DATE CUTOVER (12 PM RULE) =================
    # ================= ALWAYS YESTERDAY =================
    target_date = (now - timedelta(days=1)).date()


    # ================= PREVIOUS MONTH (BASED ON TARGET_DATE) =================
    # ================= CURRENT MONTH TO YESTERDAY =================
    TF = target_date.replace(day=1).strftime("%Y-%m-%d")
    TT = target_date.strftime("%Y-%m-%d")

    # NEW FEATURE: total days in range
    target_days = (datetime.strptime(TT, "%Y-%m-%d") - datetime.strptime(TF, "%Y-%m-%d")).days + 1


    # ================= HISTORY RANGE (120 DAYS BEFORE → TARGET_DATE) =================
    HF = (target_date - timedelta(days=120)).strftime("%Y-%m-%d")
    HT = target_date.strftime("%Y-%m-%d")

    # ================= MONTH LABEL (PREVIOUS MONTH) =================
    MONTH_LABEL = datetime.strptime(TF, "%Y-%m-%d").strftime("%B %Y")

    print("\nMONTHLY MODE (BUSINESS DATE CUTOVER ENABLED)")
    print("BUSINESS DATE :", target_date.strftime("%Y-%m-%d"))
    print("MONTH         :", MONTH_LABEL)
    print("TARGET RANGE  :", TF, "→", TT)
    print("HISTORY RANGE :", HF, "→", HT)


    tf_date = datetime.strptime(TF, "%Y-%m-%d")
    tt_date = datetime.strptime(TT, "%Y-%m-%d")

    if tt_date < tf_date:
        raise ValueError("TARGET TO date cannot be before TARGET FROM date")

    target_days = (tt_date - tf_date).days + 1  # always 1 day now

    pending = {k: v for k, v in PROPERTIES.items()}
    success_results = {}

    for run_attempt in range(1, MAX_FULL_RUN_RETRIES + 1):
        if not pending:
            break

        print(f"\n🔁 PARTIAL RUN {run_attempt}/{MAX_FULL_RUN_RETRIES}")
        tasks = [run_property_limited(P, TF, TT, HF, HT) for P in pending.values()]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        new_pending = {}
        for key, (P, result) in zip(list(pending.keys()), zip(pending.values(), results)):
            if isinstance(result, Exception):
                print(f"❌ FAILED → {P['name']} :: {result}")
                new_pending[key] = P
                continue
            success_results[key] = result
            print(f"✅ OK → {P['name']}")

        pending = new_pending

        if pending:
            if run_attempt == MAX_FULL_RUN_RETRIES:
                raise RuntimeError(f"FINAL FAILURE: Properties failed after retries: {[p['name'] for p in pending.values()]}")
            await asyncio.sleep(FULL_RUN_RETRY_DELAY)

    valid_results = [success_results[k] for k in PROPERTIES.keys() if k in success_results]
    if len(valid_results) != len(PROPERTIES):
        missing = [PROPERTIES[k]["name"] for k in PROPERTIES.keys() if k not in success_results]
        raise RuntimeError(f"DATA INCOMPLETE: Missing properties: {missing}")

    # ================= EXCEL =================
    wb = Workbook()
    wb.remove(wb.active)

    all_dfs = []
    consolidated_daily_collect = {}

    for name, df, total_rooms, prop_details, daily_collect in valid_results:
        all_dfs.append(df)

        # consolidate daily collection
        for dkey, vals in (daily_collect or {}).items():
            if dkey not in consolidated_daily_collect:
                consolidated_daily_collect[dkey] = {"cash": 0.0, "qr": 0.0, "online": 0.0}
            consolidated_daily_collect[dkey]["cash"] += float(vals.get("cash", 0) or 0)
            consolidated_daily_collect[dkey]["qr"] += float(vals.get("qr", 0) or 0)
            consolidated_daily_collect[dkey]["online"] += float(vals.get("online", 0) or 0)

        ws = wb.create_sheet(name)

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        beautify(ws)

        # ✅ ONLY 3 bottom tables
        ws.append([])
        ws.append([])
        add_payment_tables(ws, df, daily_collect, TF, TT)
        add_property_details_box(ws, prop_details)

    # ================= CONSOLIDATED SHEET =================
    big = pd.concat(all_dfs) if all_dfs else pd.DataFrame(columns=[
        "Date", "Booking Id", "Guest Name", "Status", "Booking Source",
        "Check In", "Check Out", "Cash", "QR", "Online", "Total Paid"
    ])

    ws = wb.create_sheet("CONSOLIDATED STATISTICS")

    # ✅ ONLY 2 tables in consolidated
    add_payment_tables(ws, big, consolidated_daily_collect, TF, TT, title_prefix="CONSOLIDATED — ")
    beautify(ws)

    # ================= SEND EXCEL =================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    await send_telegram_excel_buffer(
        buffer,
        filename=f"Collection_{MONTH_LABEL}.xlsx",
        caption="📊 Date Wise Collection Report (Paid Only)"
    )

    print("✅ EXCEL SENT TO TELEGRAM")
    return

# ================= RUN =================
if __name__ == "__main__":
    try:
        asyncio.run(main())
    except Exception as e:
        print("SCRIPT CRASHED")
        print(e)
        traceback.print_exc()
        print("SCRIPT CRASHED", e, flush=True)
