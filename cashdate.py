# ==============================
# ULTRA FAST ASYNC MULTI PROPERTY AUTOMATION
# HOURLY CASH REPORT
# ==============================

import os
import json
import asyncio
import aiohttp
from datetime import datetime, timedelta
import traceback
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from io import BytesIO
import pytz


IST = pytz.timezone("Asia/Kolkata")
now = datetime.now(IST)

MAX_FULL_RUN_RETRIES = 5
FULL_RUN_RETRY_DELAY = 10

PROP_PARALLEL_LIMIT = 3
DETAIL_PARALLEL_LIMIT = 10

prop_semaphore = asyncio.Semaphore(PROP_PARALLEL_LIMIT)

DETAIL_TIMEOUT = 25
BATCH_TIMEOUT = 35


# ================= TELEGRAM =================

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
if not TELEGRAM_BOT_TOKEN:
    raise RuntimeError("❌ TELEGRAM_BOT_TOKEN missing")

CHAT_MAP = json.loads(os.getenv("TELEGRAM_CHAT_MAP", "{}"))

def get_chat_id(name: str):
    if name not in CHAT_MAP:
        raise RuntimeError(f"❌ Chat ID not configured: {name}")
    return int(CHAT_MAP[name])

TELEGRAM_CHAT_ID = get_chat_id("6am")


async def send_telegram_excel_buffer(buffer, filename, caption=None):

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"

    data = aiohttp.FormData()
    data.add_field("chat_id", str(TELEGRAM_CHAT_ID))

    if caption:
        data.add_field("caption", caption)

    data.add_field(
        "document",
        buffer,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    async with aiohttp.ClientSession() as session:
        async with session.post(url, data=data, timeout=120) as resp:
            if resp.status != 200:
                text = await resp.text()
                raise RuntimeError(f"Telegram send failed: {text}")


# ================= PROPERTIES =================

PROPERTIES_RAW = json.loads(os.getenv("OYO_PROPERTIES", "{}"))
PROPERTIES = {int(k): v for k, v in PROPERTIES_RAW.items()}

if not PROPERTIES:
    raise RuntimeError("❌ OYO_PROPERTIES secret missing or empty")


# ================= COLOR =================

def get_hour_color(day_index: int, total_days: int = 31):

    if total_days <= 1:
        total_days = 31

    t = day_index / (total_days - 1)

    colors = [
        (238, 243, 251),
        (217, 242, 255),
        (255, 244, 204),
        (255, 221, 128),
        (255, 204, 153),
        (255, 193, 193),
        (232, 234, 246),
        (227, 242, 253)
    ]

    seg = t * (len(colors) - 1)
    i = int(seg)
    frac = seg - i

    if i >= len(colors) - 1:
        r, g, b = colors[-1]
    else:
        r1, g1, b1 = colors[i]
        r2, g2, b2 = colors[i + 1]

        r = int(r1 + (r2 - r1) * frac)
        g = int(g1 + (g2 - g1) * frac)
        b = int(b1 + (b2 - b1) * frac)

    return f"{r:02X}{g:02X}{b:02X}"


# ================= FETCH DETAILS =================

async def fetch_booking_details(session, P, booking_no):

    url = "https://www.oyoos.com/hms_ms/api/v1/visibility/booking_details_with_entities"

    params = {
        "qid": P["QID"],
        "booking_id": booking_no,
        "role": 0,
        "platform": "OYOOS",
        "country_code": 1
    }

    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}

    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    for attempt in range(1, 4):

        try:
            async with session.get(
                url,
                params=params,
                headers=headers,
                cookies=cookies,
                timeout=DETAIL_TIMEOUT
            ) as r:

                if r.status != 200:
                    raise RuntimeError("DETAIL API FAILED")

                data = await r.json()

                booking = next(
                    iter(data.get("entities", {}).get("bookings", {}).values()),
                    {}
                )

                payments = booking.get("payments", [])

                events = []

                for p in payments:

                    amt = float(p.get("amount", 0) or 0)
                    if amt <= 0:
                        continue

                    created_at = str(p.get("created_at") or "").strip()
                    if not created_at:
                        continue

                    try:
                        dt = datetime.fromisoformat(created_at.replace("Z", ""))
                        dt = dt.astimezone(IST)
                    except Exception:
                        continue

                    mode_raw = p.get("mode", "")

                    if mode_raw != "Cash at Hotel":
                        continue

                    events.append({
                        "date": dt.strftime("%Y-%m-%d"),
                        "hour": dt.hour,
                        "mode": "cash",
                        "amt": amt
                    })

                return events

        except Exception:
            await asyncio.sleep(2 + attempt)

    raise RuntimeError("DETAIL FETCH FAILED")


# ================= FETCH BOOKINGS =================

async def fetch_bookings_batch(session, offset, f, t, P):

    url = "https://www.oyoos.com/hms_ms/api/v1/get_booking_with_ids"

    params = {
        "qid": P["QID"],
        "checkin_from": f,
        "checkin_till": t,
        "batch_count": 100,
        "batch_offset": offset,
        "visibility_required": "true",
        "additionalParams": "payment_hold_transaction,guest,stay_details",
        "decimal_price": "true",
        "ascending": "true",
        "sort_on": "checkin_date"
    }

    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}

    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    async with session.get(
        url,
        params=params,
        cookies=cookies,
        headers=headers,
        timeout=BATCH_TIMEOUT
    ) as r:

        if r.status != 200:
            raise RuntimeError("BATCH API FAILED")

        return await r.json()


# ================= PROCESS PROPERTY =================

async def process_property(P, TF, TT, HF, HT):

    print(f"PROCESSING → {P['name']}")

    tf_dt = datetime.strptime(TF, "%Y-%m-%d").date()
    tt_dt = datetime.strptime(TT, "%Y-%m-%d").date()

    async with aiohttp.ClientSession() as session:

        detail_semaphore = asyncio.Semaphore(DETAIL_PARALLEL_LIMIT)
        detail_cache = {}

        async def limited_detail_call(booking_no):
            async with detail_semaphore:
                if booking_no in detail_cache:
                    return detail_cache[booking_no]
                res = await fetch_booking_details(session, P, booking_no)
                detail_cache[booking_no] = res
                return res

        date_map = {}

        d = tf_dt
        while d <= tt_dt:
            date_map[d] = {"cash": 0.0}
            d += timedelta(days=1)

        offset = 0

        while True:

            data = await fetch_bookings_batch(session, offset, HF, HT, P)

            if not data or not data.get("bookingIds"):
                break

            bookings = data.get("entities", {}).get("bookings", {})
            if not bookings:
                break

            tasks = []

            for b in bookings.values():

                status = (b.get("status") or "").strip()
                if status not in ["Checked In", "Checked Out"]:
                    continue

                booking_no = b.get("booking_no")
                if not booking_no:
                    continue

                tasks.append(limited_detail_call(booking_no))

            results = await asyncio.gather(*tasks, return_exceptions=True)

            for res in results:

                if isinstance(res, Exception):
                    continue

                for ev in res or []:

                    try:
                        d_dt = datetime.strptime(ev["date"], "%Y-%m-%d").date()
                    except:
                        continue

                    if not (tf_dt <= d_dt <= tt_dt):
                        continue

                    amt = float(ev["amt"])
                    date_map[d_dt]["cash"] += amt

            if len(data.get("bookingIds", [])) < 100:
                break

            offset += 100

        return (P["name"], date_map)


# ================= MAIN =================

async def main():

    global now
    now = datetime.now(IST)

    target_date = (now - timedelta(days=1)).date()

    TF = target_date.replace(day=1).strftime("%Y-%m-%d")
    TT = target_date.strftime("%Y-%m-%d")

    HF = (target_date - timedelta(days=120)).strftime("%Y-%m-%d")
    HT = TT

    pending = {k: v for k, v in PROPERTIES.items()}
    success_results = {}

    for run_attempt in range(1, MAX_FULL_RUN_RETRIES + 1):

        if not pending:
            break

        tasks = [run_property_limited(P, TF, TT, HF, HT) for P in pending.values()]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        new_pending = {}

        for key, (P, result) in zip(list(pending.keys()), zip(pending.values(), results)):

            if isinstance(result, Exception):
                new_pending[key] = P
                continue

            success_results[key] = result

        pending = new_pending

        if pending:

            if run_attempt == MAX_FULL_RUN_RETRIES:
                failed_names = [p["name"] for p in pending.values()]
                raise RuntimeError(
                    f"FINAL FAILURE: Properties failed after retries: {failed_names}"
                )

            await asyncio.sleep(FULL_RUN_RETRY_DELAY)

    valid_results = [
        success_results[k]
        for k in PROPERTIES.keys()
        if k in success_results
    ]

    wb = Workbook()
    wb.remove(wb.active)

    start_dt = datetime.strptime(TF, "%Y-%m-%d").date()
    end_dt = datetime.strptime(TT, "%Y-%m-%d").date()

    date_list = []
    d = start_dt
    while d <= end_dt:
        date_list.append(d)
        d += timedelta(days=1)

    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    def create_sheet(ws, date_map):

        ws.append(["Date", "Cash"])

        header_fill = PatternFill("solid", fgColor="1F4E78")

        for col in range(1, 3):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")

        sum_cash = 0

        for idx, d in enumerate(date_list):

            row = date_map.get(d, {"cash": 0})
            cash = round(row["cash"], 2)
            sum_cash += cash

            ws.append([d.strftime("%d-%m-%Y"), cash])

            r = ws.max_row
            fill = PatternFill("solid", fgColor=get_hour_color(idx, len(date_list)))

            for c in range(1, 3):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill
                cell.border = thin
                cell.alignment = Alignment(horizontal="center")

        ws.append(["TOTAL", round(sum_cash, 2)])

        total_row = ws.max_row

        for c in range(1, 3):
            cell = ws.cell(row=total_row, column=c)
            cell.fill = PatternFill("solid", fgColor="000000")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 18


    for name, date_map in valid_results:

        ws = wb.create_sheet(name[:31])
        create_sheet(ws, date_map)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    await send_telegram_excel_buffer(
        buffer,
        filename=f"Cash_Monthly.xlsx",
        caption="📊 Date-wise Cash Collection Report"
    )


# ================= RUN =================

async def run_property_with_retry(P, TF, TT, HF, HT, retries=5):

    for attempt in range(1, retries + 1):

        try:
            return await process_property(P, TF, TT, HF, HT)

        except Exception as e:

            print(f"RETRY {attempt}/{retries} → {P['name']} :: {e}")

            await asyncio.sleep(2 + attempt * 2)

    raise RuntimeError(f"PROPERTY FAILED → {P['name']}")


async def run_property_limited(P, TF, TT, HF, HT):

    async with prop_semaphore:
        return await run_property_with_retry(P, TF, TT, HF, HT)


if __name__ == "__main__":

    try:
        asyncio.run(main())

    except Exception:

        print("SCRIPT CRASHED")
        traceback.print_exc()
