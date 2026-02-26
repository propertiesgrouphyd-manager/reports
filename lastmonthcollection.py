# ==============================
# ULTRA FAST ASYNC MULTI PROPERTY AUTOMATION
# DATE-WISE COLLECTION BASED ON PAYMENT CREATED_AT
# FINAL EXCEL: ONLY PAID BOOKINGS (NO PER DAY STAY CALC)
# ==============================

import os
import json
import asyncio
import aiohttp
import pandas as pd
from datetime import datetime, timedelta
import traceback
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import pytz
IST = pytz.timezone("Asia/Kolkata")
now = datetime.now(IST)

MAX_FULL_RUN_RETRIES = 5
FULL_RUN_RETRY_DELAY = 10

PROP_PARALLEL_LIMIT = 3
DETAIL_PARALLEL_LIMIT = 10

prop_semaphore = asyncio.Semaphore(PROP_PARALLEL_LIMIT)

DETAIL_TIMEOUT = 25
ROOMS_TIMEOUT = 25
BATCH_TIMEOUT = 35


# ================= TELEGRAM =================

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")

if not TELEGRAM_BOT_TOKEN:
    raise RuntimeError("❌ TELEGRAM_BOT_TOKEN missing")


CHAT_MAP = json.loads(os.getenv("TELEGRAM_CHAT_MAP", "{}"))

def get_chat_id(name: str):
    if name not in CHAT_MAP:
        raise RuntimeError(f"❌ Chat ID not configured: {name}")
    return int(CHAT_MAP[name])


# choose chat key from secret map
TELEGRAM_CHAT_ID = get_chat_id("6am")   # change if needed


# ================= PROPERTIES =================

PROPERTIES_RAW = json.loads(os.getenv("OYO_PROPERTIES", "{}"))

PROPERTIES = {int(k): v for k, v in PROPERTIES_RAW.items()}

if not PROPERTIES:
    raise RuntimeError("❌ OYO_PROPERTIES secret missing or empty")

# ================= TELEGRAM =================
async def send_telegram_excel_buffer(buffer, filename, caption=None):
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"

    data = aiohttp.FormData()
    data.add_field("chat_id", str(TELEGRAM_CHAT_ID))
    if caption:
        data.add_field("caption", caption)

    data.add_field(
        "document",
        buffer,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    async with aiohttp.ClientSession() as session:
        async with session.post(url, data=data, timeout=120) as resp:
            if resp.status != 200:
                text = await resp.text()
                raise RuntimeError(f"Telegram send failed: {text}")

# ================= BEAUTIFY EXCEL =================
def beautify(ws):
    blue = PatternFill("solid", fgColor="1F4E78")
    light1 = PatternFill("solid", fgColor="DDEBF7")
    light2 = PatternFill("solid", fgColor="F2F2F2")
    yellow = PatternFill("solid", fgColor="FFF4CC")

    bold_white = Font(color="FFFFFF", bold=True, size=12)
    bold_black = Font(color="000000", bold=True, size=12)

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    max_row = ws.max_row
    max_col = ws.max_column
    ws.freeze_panes = "A2"

    for col in range(1, max_col + 1):
        c = ws.cell(row=1, column=col)
        c.fill = blue
        c.font = bold_white
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin

    for r in range(2, max_row + 1):
        fill = light1 if r % 2 == 0 else light2
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue

            if cell.fill is not None and cell.fill.patternType is not None:
                cell.border = thin
                continue

            cell.fill = fill
            cell.border = thin

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 5

    for r in range(2, max_row + 1):
        text = str(ws.cell(row=r, column=1).value or "")
        if text.strip() == "":
            continue
        if "Booking" in text or "Amount" in text or "Total" in text or "OYO" in text:
            ws.cell(row=r, column=1).fill = yellow
            ws.cell(row=r, column=1).font = bold_black

# ================= BOOKING SOURCE =================
def get_booking_source(b):
    source = str(b.get("source", "") or "").strip()
    ota = str(b.get("ota_source", "") or "").strip()
    sub = str(b.get("sub_source", "") or "").strip()
    corp = bool(b.get("is_corporate", False))
    booking_identifier = str(b.get("booking_identifier", "") or "").strip()

    if booking_identifier == "TA":
        return "TA"
    if source == "Walk In":
        return "Walk-in"
    if corp or sub == "corporate":
        return "CB"
    if "Booking.com" in ota:
        return "BDC"
    if "GoMMT" in ota:
        return "MMT"
    if "Agoda" in ota:
        return "Agoda"
    if source == "Travel Agent" or sub == "TPO":
        return "TA"
    if source in ["Android App","IOS App","Web Booking","Mobile Web Booking","Website Booking","Direct"]:
        return "OYO"
    return "OBA"

# ================= PREMIUM PROPERTY DETAILS BOX =================
def add_property_details_box(ws, prop):
    blue = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="DDEBF7")
    white = PatternFill("solid", fgColor="FFFFFF")

    bold_white = Font(color="FFFFFF", bold=True, size=12)
    bold_black = Font(color="000000", bold=True, size=11)
    normal = Font(color="000000", size=11)
    link_font = Font(color="0563C1", underline="single", bold=True, size=11)

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    start_col = 1
    end_col = 8

    def _border_range(r1, c1, r2, c2):
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(row=rr, column=cc).border = thin

    def _merge(row, c1, c2, value, fill=None, font=None, center=False, wrap=False):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1)
        cell.value = value
        if fill: cell.fill = fill
        if font: cell.font = font
        cell.alignment = Alignment(horizontal="center" if center else "left",
                                   vertical="center", wrap_text=wrap)
        return cell

    plot = (prop.get("plot_number") or "").strip()
    street = (prop.get("street") or "").strip()
    pincode = (prop.get("pincode") or "").strip()
    city = (prop.get("city") or "").strip()
    country = (prop.get("country") or "").strip()

    address_parts = []
    if plot: address_parts.append(plot)
    if street: address_parts.append(street)
    city_pin = " ".join([x for x in [city, pincode] if x]).strip()
    if city_pin: address_parts.append(city_pin)
    if country: address_parts.append(country)
    address = ", ".join(address_parts).strip()

    ws.append([])
    ws.append([])
    top = ws.max_row + 1

    _merge(top, start_col, end_col, "PROPERTY DETAILS", fill=blue, font=bold_white, center=True)
    ws.row_dimensions[top].height = 22

    _merge(top + 1, 1, 2, "Name", fill=light, font=bold_black, wrap=True)
    _merge(top + 1, 3, end_col, prop.get("name", "") or "", fill=white, font=normal, wrap=True)

    _merge(top + 2, 1, 2, "Alternative Name", fill=light, font=bold_black, wrap=True)
    _merge(top + 2, 3, end_col, prop.get("alternate_name", "") or "", fill=white, font=normal, wrap=True)

    _merge(top + 3, 1, 2, "Address", fill=light, font=bold_black, wrap=True)
    _merge(top + 3, 3, end_col, address, fill=white, font=normal, wrap=True)
    ws.row_dimensions[top + 3].height = 45

    _merge(top + 4, 1, 2, "Google Map", fill=light, font=bold_black, wrap=True)
    map_link = (prop.get("map_link") or "").strip() or ""
    link_cell = _merge(top + 4, 3, end_col,
                       "OPEN IN GOOGLE MAPS" if map_link else "",
                       fill=white, font=link_font, center=True)
    if map_link:
        link_cell.hyperlink = map_link

    _border_range(top, start_col, top + 4, end_col)

# ================= PREMIUM PAYMENT TABLES =================
# ================= PREMIUM PAYMENT TABLES =================
def add_payment_tables(ws, df, daily_collect, TF, TT, title_prefix=""):
    blue = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="DDEBF7")
    white = PatternFill("solid", fgColor="FFFFFF")
    yellow = PatternFill("solid", fgColor="FFF4CC")

    bold_white = Font(color="FFFFFF", bold=True, size=12)
    bold_black = Font(color="000000", bold=True, size=11)
    normal = Font(color="000000", size=11)

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def _style_row(row, start_col, end_col, fill=None, font=None, center=True):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=c)
            if fill: cell.fill = fill
            if font: cell.font = font
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if center else "left",
                                       vertical="center")

    def _merge(row, c1, c2, value, fill=None, font=None, center=True):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1)
        cell.value = value
        if fill: cell.fill = fill
        if font: cell.font = font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center" if center else "left",
                                   vertical="center")
        return cell

    start_col = 1
    end_col = 7

    premium_widths = [18, 14, 14, 14, 14, 14, 16]
    for i, w in enumerate(premium_widths, start=1):
        col_letter = get_column_letter(i)
        current = ws.column_dimensions[col_letter].width
        ws.column_dimensions[col_letter].width = max(current or 0, w)

    ws.append([])

    # ================= TABLE 1 =================
    top = ws.max_row + 1
    heading = f"{title_prefix}BOOKING SOURCE × PAYMENT MODE".strip()
    _merge(top, start_col, end_col, heading, fill=blue, font=bold_white, center=True)

    headers = ["Source", "Cash", "QR", "Online", "Discount", "Total Paid", ""]
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=top + 1, column=idx).value = h
    _style_row(top + 1, start_col, end_col, fill=light, font=bold_black)

    sources = ["OYO", "Walk-in", "MMT", "BDC", "Agoda", "CB", "TA", "OBA"]
    r = top + 2

    for src in sources:
        part = df[df["Booking Source"] == src] if (not df.empty and "Booking Source" in df.columns) else df

        cash = round(float(part["Cash"].sum()), 2) if (not part.empty and "Cash" in part.columns) else 0
        qr = round(float(part["QR"].sum()), 2) if (not part.empty and "QR" in part.columns) else 0
        online = round(float(part["Online"].sum()), 2) if (not part.empty and "Online" in part.columns) else 0
        discount = round(float(part["Discount"].sum()), 2) if (not part.empty and "Discount" in part.columns) else 0

        total_paid = round(cash + qr + online + discount, 2)

        ws.cell(row=r, column=1).value = src
        ws.cell(row=r, column=2).value = cash
        ws.cell(row=r, column=3).value = qr
        ws.cell(row=r, column=4).value = online
        ws.cell(row=r, column=5).value = discount
        ws.cell(row=r, column=6).value = total_paid

        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin
            cell.font = normal
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = white

        ws.cell(row=r, column=1).fill = yellow
        ws.cell(row=r, column=1).font = bold_black
        r += 1

    # TOTAL row
    tot_cash = round(float(df["Cash"].sum()), 2) if not df.empty else 0
    tot_qr = round(float(df["QR"].sum()), 2) if not df.empty else 0
    tot_online = round(float(df["Online"].sum()), 2) if not df.empty else 0
    tot_discount = round(float(df["Discount"].sum()), 2) if not df.empty else 0
    tot_paid = round(tot_cash + tot_qr + tot_online + tot_discount, 2)

    ws.cell(row=r, column=1).value = "TOTAL"
    ws.cell(row=r, column=2).value = tot_cash
    ws.cell(row=r, column=3).value = tot_qr
    ws.cell(row=r, column=4).value = tot_online
    ws.cell(row=r, column=5).value = tot_discount
    ws.cell(row=r, column=6).value = tot_paid

    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = thin
        cell.font = bold_black
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = light

    ws.cell(row=r, column=1).fill = yellow
    ws.cell(row=r, column=1).font = bold_black

    ws.append([])

    # ================= TABLE 2 =================
    top2 = ws.max_row + 1
    heading2 = f"{title_prefix}DATE WISE COLLECTION SUMMARY".strip()
    _merge(top2, start_col, end_col, heading2, fill=blue, font=bold_white, center=True)

    headers2 = ["Date", "Cash", "QR", "Online", "Discount", "Total Paid", ""]
    for idx, h in enumerate(headers2, start=1):
        ws.cell(row=top2 + 1, column=idx).value = h
    _style_row(top2 + 1, start_col, end_col, fill=light, font=bold_black)

    rr = top2 + 2
    tf_dt = datetime.strptime(TF, "%Y-%m-%d").date()
    tt_dt = datetime.strptime(TT, "%Y-%m-%d").date()

    grand_cash = grand_qr = grand_online = grand_discount = 0.0

    cur = tf_dt
    while cur <= tt_dt:
        dkey = cur.strftime("%Y-%m-%d")
        vals = daily_collect.get(dkey, {
            "cash": 0.0,
            "qr": 0.0,
            "online": 0.0,
            "discount": 0.0
        })

        cash = round(float(vals.get("cash", 0)), 2)
        qr = round(float(vals.get("qr", 0)), 2)
        online = round(float(vals.get("online", 0)), 2)
        discount = round(float(vals.get("discount", 0)), 2)

        total_paid = round(cash + qr + online + discount, 2)

        ws.cell(row=rr, column=1).value = dkey
        ws.cell(row=rr, column=2).value = cash
        ws.cell(row=rr, column=3).value = qr
        ws.cell(row=rr, column=4).value = online
        ws.cell(row=rr, column=5).value = discount
        ws.cell(row=rr, column=6).value = total_paid

        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=rr, column=c)
            cell.border = thin
            cell.font = normal
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = white

        ws.cell(row=rr, column=1).fill = yellow
        ws.cell(row=rr, column=1).font = bold_black

        grand_cash += cash
        grand_qr += qr
        grand_online += online
        grand_discount += discount

        rr += 1
        cur += timedelta(days=1)

    grand_total = round(grand_cash + grand_qr + grand_online + grand_discount, 2)

    ws.cell(row=rr, column=1).value = "TOTAL"
    ws.cell(row=rr, column=2).value = round(grand_cash, 2)
    ws.cell(row=rr, column=3).value = round(grand_qr, 2)
    ws.cell(row=rr, column=4).value = round(grand_online, 2)
    ws.cell(row=rr, column=5).value = round(grand_discount, 2)
    ws.cell(row=rr, column=6).value = grand_total

    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=rr, column=c)
        cell.border = thin
        cell.font = bold_black
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = light

    ws.cell(row=rr, column=1).fill = yellow
    ws.cell(row=rr, column=1).font = bold_black

    ws.append([])
# ================= FETCH DETAILS =================
# ================= FETCH DETAILS =================
async def fetch_booking_details(session, P, booking_no):
    url = "https://www.oyoos.com/hms_ms/api/v1/visibility/booking_details_with_entities"
    params = {
        "qid": P["QID"],
        "booking_id": booking_no,
        "role": 0,
        "platform": "OYOOS",
        "country_code": 1
    }
    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}
    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    for attempt in range(1, 4):
        try:
            async with session.get(
                url,
                params=params,
                headers=headers,
                cookies=cookies,
                timeout=DETAIL_TIMEOUT
            ) as r:
                if r.status != 200:
                    raise RuntimeError("DETAIL API FAILED")

                data = await r.json()
                booking = next(
                    iter(data.get("entities", {}).get("bookings", {}).values()),
                    {}
                )
                payments = booking.get("payments", [])

                payment_events = []

                for p in payments:
                    mode = p.get("mode", "")
                    amt = float(p.get("amount", 0) or 0)

                    # 🔥 SKIP ZERO AMOUNT ENTRIES
                    if amt <= 0:
                        continue


                    created_at = str(p.get("created_at") or "").strip()
                    if not created_at:
                        continue

                    try:
                        dt = datetime.fromisoformat(created_at.replace("Z", ""))
                        dt = dt.astimezone(IST)
                    except Exception:
                        continue

                    pay_date = dt.strftime("%Y-%m-%d")
                    pay_time = dt.strftime("%H:%M")

                    # ================= CLASSIFICATION =================
                    if mode == "Cash at Hotel":
                        bucket = "cash"
                    elif mode == "UPI QR":
                        bucket = "qr"
                    elif mode == "oyo_wizard_discount":
                        bucket = "discount"
                    else:
                        bucket = "online"

                    payment_events.append({
                        "date": pay_date,
                        "time": pay_time,
                        "mode": bucket,
                        "amt": amt
                    })

                return payment_events

        except Exception:
            await asyncio.sleep(2 + attempt)

    raise RuntimeError("DETAIL FETCH FAILED")


# ================= BATCH FETCH =================
async def fetch_bookings_batch(session, offset, f, t, P):
    url = "https://www.oyoos.com/hms_ms/api/v1/get_booking_with_ids"
    params = {
        "qid": P["QID"],
        "checkin_from": f,
        "checkin_till": t,
        "batch_count": 100,
        "batch_offset": offset,
        "visibility_required": "true",
        "additionalParams": "payment_hold_transaction,guest,stay_details",
        "decimal_price": "true",
        "ascending": "true",
        "sort_on": "checkin_date"
    }
    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}
    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    async with session.get(url, params=params, cookies=cookies, headers=headers, timeout=BATCH_TIMEOUT) as r:
        if r.status != 200:
            raise RuntimeError("BATCH API FAILED")
        return await r.json()

# ================= PROPERTY DETAILS API =================
async def fetch_property_details(session, P):
    url = "https://www.oyoos.com/hms_ms/api/v1/location/property-details"
    params = {"qid": P["QID"]}
    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}
    headers = {"accept": "application/json", "x-qid": str(P["QID"]), "x-source-client": "merchant"}

    for attempt in range(1, 4):
        try:
            async with session.get(url, params=params, cookies=cookies, headers=headers, timeout=20) as r:
                if r.status != 200:
                    raise RuntimeError(f"PROPERTY DETAILS API FAILED ({r.status})")
                data = await r.json()
                return {
                    "name": str(data.get("name", "") or "").strip(),
                    "alternate_name": str(data.get("alternate_name", "") or "").strip(),
                    "plot_number": str(data.get("plot_number", "") or "").strip(),
                    "street": str(data.get("street", "") or "").strip(),
                    "pincode": str(data.get("pincode", "") or "").strip(),
                    "city": str(data.get("city", "") or "").strip(),
                    "country": str(data.get("country", "") or "").strip(),
                    "map_link": str(data.get("map_link", "") or "").strip(),
                }
        except Exception:
            await asyncio.sleep(2 + attempt)

    return {"name":"","alternate_name":"","plot_number":"","street":"","pincode":"","city":"","country":"","map_link":""}

# ================= FETCH TOTAL ROOMS =================
async def fetch_total_rooms(session, P):
    url = "https://www.oyoos.com/hms_ms/api/v1/hotels/roomsNew"
    params = {"qid": P["QID"]}
    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}
    headers = {"accept": "application/json", "x-qid": str(P["QID"]), "x-source-client": "merchant"}

    for attempt in range(1, 4):
        try:
            async with session.get(url, params=params, cookies=cookies, headers=headers, timeout=ROOMS_TIMEOUT) as r:
                if r.status != 200:
                    raise RuntimeError("ROOM API FAILED")
                data = await r.json()
                rooms = data.get("rooms", {})
                return len(rooms)
        except Exception:
            await asyncio.sleep(2 + attempt)

    return 0

# ================= PROCESS PROPERTY =================
# ================= PROCESS PROPERTY =================
async def process_property(P, TF, TT, HF, HT):
    print(f"PROCESSING → {P['name']}")

    tf_dt = datetime.strptime(TF, "%Y-%m-%d").date()
    tt_dt = datetime.strptime(TT, "%Y-%m-%d").date()

    async with aiohttp.ClientSession() as session:
        total_rooms = await fetch_total_rooms(session, P)
        prop_details = await fetch_property_details(session, P)

        detail_semaphore = asyncio.Semaphore(DETAIL_PARALLEL_LIMIT)
        detail_cache = {}

        async def limited_detail_call(booking_no):
            async with detail_semaphore:
                if booking_no in detail_cache:
                    return detail_cache[booking_no]
                res = await fetch_booking_details(session, P, booking_no)
                detail_cache[booking_no] = res
                return res

        daily_collect = {}
        booking_date_mode_map = {}

        offset = 0
        while True:
            data = await fetch_bookings_batch(session, offset, HF, HT, P)
            if not data or not data.get("bookingIds"):
                break

            bookings = data.get("entities", {}).get("bookings", {})
            if not bookings:
                break

            tasks = []
            mapping = []

            for b in bookings.values():
                status = (b.get("status") or "").strip()
                if status not in ["Checked In", "Checked Out"]:
                    continue

                booking_no = b.get("booking_no")
                if not booking_no:
                    continue

                tasks.append(limited_detail_call(booking_no))
                mapping.append(b)

            results = await asyncio.gather(*tasks, return_exceptions=True)

            for res, b in zip(results, mapping):
                if isinstance(res, Exception):
                    continue

                payment_events = res or []
                booking_no = b.get("booking_no")
                source = get_booking_source(b)

                for ev in payment_events:
                    d = ev.get("date")
                    if not d:
                        continue

                    try:
                        d_dt = datetime.strptime(d, "%Y-%m-%d").date()
                    except Exception:
                        continue

                    if not (tf_dt <= d_dt <= tt_dt):
                        continue

                    if d not in daily_collect:
                        daily_collect[d] = {
                            "cash": 0.0,
                            "qr": 0.0,
                            "online": 0.0,
                            "discount": 0.0
                        }

                    if (d, booking_no) not in booking_date_mode_map:
                        booking_date_mode_map[(d, booking_no)] = {
                            "cash": 0.0,
                            "qr": 0.0,
                            "online": 0.0,
                            "discount": 0.0,
                            "times": set(),
                            "b": b,
                            "source": source
                        }

                    mode = ev.get("mode")
                    amt = float(ev.get("amt", 0) or 0)
                    ptime = ev.get("time")

                    if ptime:
                        booking_date_mode_map[(d, booking_no)]["times"].add(ptime)

                    if mode == "cash":
                        daily_collect[d]["cash"] += amt
                        booking_date_mode_map[(d, booking_no)]["cash"] += amt
                    elif mode == "qr":
                        daily_collect[d]["qr"] += amt
                        booking_date_mode_map[(d, booking_no)]["qr"] += amt
                    elif mode == "discount":
                        daily_collect[d]["discount"] += amt
                        booking_date_mode_map[(d, booking_no)]["discount"] += amt
                    else:
                        daily_collect[d]["online"] += amt
                        booking_date_mode_map[(d, booking_no)]["online"] += amt

            if len(data.get("bookingIds", [])) < 100:
                break
            offset += 100

        # ================= BUILD DATAFRAME =================
        all_rows = []

        for (d, booking_no), vals in booking_date_mode_map.items():
            b = vals["b"]

            cash = vals["cash"]
            qr = vals["qr"]
            online = vals["online"]
            discount = vals["discount"]

            total_paid = cash + qr + online + discount
            times_str = ", ".join(sorted(vals["times"]))

            all_rows.append({
                "Date": d,
                "Booking Id": booking_no,
                "Guest Name": b.get("guest_name"),
                "Status": b.get("status"),
                "Booking Source": vals["source"],
                "Check In": b.get("checkin"),
                "Check Out": b.get("checkout"),
                "Cash": round(cash, 2),
                "QR": round(qr, 2),
                "Online": round(online, 2),
                "Discount": round(discount, 2),
                "Total Paid": round(total_paid, 2),
                "Time": times_str
            })

        df = pd.DataFrame(all_rows)

        if df.empty:
            df = pd.DataFrame(columns=[
                "Date", "Booking Id", "Guest Name", "Status",
                "Booking Source", "Check In", "Check Out",
                "Cash", "QR", "Online", "Discount",
                "Total Paid", "Time"
            ])

        df = df.sort_values(["Date", "Booking Id"], ascending=True)

        return (P["name"], df, total_rooms, prop_details, daily_collect)


# ================= RETRY =================
async def run_property_with_retry(P, TF, TT, HF, HT, retries=3):
    last_error = None
    for attempt in range(1, retries + 1):
        try:
            return await process_property(P, TF, TT, HF, HT)
        except Exception as e:
            last_error = e
            print(f"RETRY {attempt}/{retries} → {P['name']} :: {e}")
            await asyncio.sleep(2 + attempt * 2)
    raise RuntimeError(f"PROPERTY FAILED → {P['name']}") from last_error

async def run_property_limited(P, TF, TT, HF, HT):
    async with prop_semaphore:
        return await run_property_with_retry(P, TF, TT, HF, HT)

# ================= MAIN =================
async def main():
    print("========================================")
    print(" OYO MONTHLY TELEGRAM AUTOMATION")
    print("========================================")

    global now
    now = datetime.now(IST)

    # ================= BUSINESS DATE CUTOVER (12 PM RULE) =================
    if now.hour < 12:
        target_date = (now - timedelta(days=1)).date()
    else:
        target_date = now.date()

    # ================= PREVIOUS MONTH (BASED ON TARGET_DATE) =================
    first_of_current_month = target_date.replace(day=1)
    last_day_previous_month = first_of_current_month - timedelta(days=1)
    first_day_previous_month = last_day_previous_month.replace(day=1)

    TF = first_day_previous_month.strftime("%Y-%m-%d")
    TT = last_day_previous_month.strftime("%Y-%m-%d")

    # NEW FEATURE: total days in month
    target_days = (datetime.strptime(TT, "%Y-%m-%d") - datetime.strptime(TF, "%Y-%m-%d")).days + 1

    # ================= HISTORY RANGE (120 DAYS BEFORE → TARGET_DATE) =================
    HF = (target_date - timedelta(days=120)).strftime("%Y-%m-%d")
    HT = target_date.strftime("%Y-%m-%d")

    # ================= MONTH LABEL (PREVIOUS MONTH) =================
    MONTH_LABEL = first_day_previous_month.strftime("%B %Y")

    print("\nMONTHLY MODE (BUSINESS DATE CUTOVER ENABLED)")
    print("BUSINESS DATE :", target_date.strftime("%Y-%m-%d"))
    print("MONTH         :", MONTH_LABEL)
    print("TARGET RANGE  :", TF, "→", TT)
    print("HISTORY RANGE :", HF, "→", HT)

    tf_date = datetime.strptime(TF, "%Y-%m-%d")
    tt_date = datetime.strptime(TT, "%Y-%m-%d")

    if tt_date < tf_date:
        raise ValueError("TARGET TO date cannot be before TARGET FROM date")

    pending = {k: v for k, v in PROPERTIES.items()}
    success_results = {}

    for run_attempt in range(1, MAX_FULL_RUN_RETRIES + 1):
        if not pending:
            break

        print(f"\n🔁 PARTIAL RUN {run_attempt}/{MAX_FULL_RUN_RETRIES}")
        tasks = [run_property_limited(P, TF, TT, HF, HT) for P in pending.values()]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        new_pending = {}
        for key, (P, result) in zip(list(pending.keys()), zip(pending.values(), results)):
            if isinstance(result, Exception):
                print(f"❌ FAILED → {P['name']} :: {result}")
                new_pending[key] = P
                continue
            success_results[key] = result
            print(f"✅ OK → {P['name']}")

        pending = new_pending

        if pending:
            if run_attempt == MAX_FULL_RUN_RETRIES:
                raise RuntimeError(
                    f"FINAL FAILURE: Properties failed after retries: {[p['name'] for p in pending.values()]}"
                )
            await asyncio.sleep(FULL_RUN_RETRY_DELAY)

    valid_results = [success_results[k] for k in PROPERTIES.keys() if k in success_results]

    if len(valid_results) != len(PROPERTIES):
        missing = [PROPERTIES[k]["name"] for k in PROPERTIES.keys() if k not in success_results]
        raise RuntimeError(f"DATA INCOMPLETE: Missing properties: {missing}")

    # ================= EXCEL =================
    wb = Workbook()
    wb.remove(wb.active)

    all_dfs = []
    consolidated_daily_collect = {}

    for name, df, total_rooms, prop_details, daily_collect in valid_results:
        all_dfs.append(df)

        # consolidate daily collection (INCLUDING DISCOUNT)
        for dkey, vals in (daily_collect or {}).items():
            if dkey not in consolidated_daily_collect:
                consolidated_daily_collect[dkey] = {
                    "cash": 0.0,
                    "qr": 0.0,
                    "online": 0.0,
                    "discount": 0.0
                }

            consolidated_daily_collect[dkey]["cash"] += float(vals.get("cash", 0) or 0)
            consolidated_daily_collect[dkey]["qr"] += float(vals.get("qr", 0) or 0)
            consolidated_daily_collect[dkey]["online"] += float(vals.get("online", 0) or 0)
            consolidated_daily_collect[dkey]["discount"] += float(vals.get("discount", 0) or 0)

        ws = wb.create_sheet(name)

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        beautify(ws)

        ws.append([])
        ws.append([])
        add_payment_tables(ws, df, daily_collect, TF, TT)
        add_property_details_box(ws, prop_details)

    # ================= CONSOLIDATED SHEET =================
    big = pd.concat(all_dfs) if all_dfs else pd.DataFrame(columns=[
        "Date", "Booking Id", "Guest Name", "Status", "Booking Source",
        "Check In", "Check Out",
        "Cash", "QR", "Online", "Discount", "Total Paid", "Time"
    ])

    ws = wb.create_sheet("CONSOLIDATED STATISTICS")

    add_payment_tables(ws, big, consolidated_daily_collect, TF, TT, title_prefix="CONSOLIDATED — ")
    beautify(ws)

    # ================= SEND EXCEL =================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    await send_telegram_excel_buffer(
        buffer,
        filename=f"Collection_{MONTH_LABEL}.xlsx",
        caption="📊 Date Wise Collection Report (Paid Only)"
    )

    print("✅ EXCEL SENT TO TELEGRAM")
    return


# ================= RUN =================
if __name__ == "__main__":
    try:
        asyncio.run(main())
    except Exception as e:
        print("SCRIPT CRASHED")
        print(e)
        traceback.print_exc()
        print("SCRIPT CRASHED", e, flush=True)
