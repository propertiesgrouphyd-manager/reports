# ==============================
# ULTRA FAST ASYNC MULTI PROPERTY AUTOMATION
# PER-DAY STAY CASH + QR + ONLINE + DISCOUNT + BALANCE REPORT
# ==============================

import os
import json
import asyncio
import aiohttp
from datetime import datetime, timedelta
import traceback
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from io import BytesIO
import pytz


IST = pytz.timezone("Asia/Kolkata")
now = datetime.now(IST)

MAX_FULL_RUN_RETRIES = 5
FULL_RUN_RETRY_DELAY = 10

PROP_PARALLEL_LIMIT = 3
DETAIL_PARALLEL_LIMIT = 10

prop_semaphore = asyncio.Semaphore(PROP_PARALLEL_LIMIT)

DETAIL_TIMEOUT = 25
BATCH_TIMEOUT = 35


# ================= TELEGRAM =================

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
CHAT_MAP = json.loads(os.getenv("TELEGRAM_CHAT_MAP", "{}"))

def get_chat_id(name: str):
    return int(CHAT_MAP[name])

TELEGRAM_CHAT_ID = get_chat_id("6am")


async def send_telegram_excel_buffer(buffer, filename, caption=None):

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"

    data = aiohttp.FormData()
    data.add_field("chat_id", str(TELEGRAM_CHAT_ID))

    if caption:
        data.add_field("caption", caption)

    data.add_field(
        "document",
        buffer,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    async with aiohttp.ClientSession() as session:
        async with session.post(url, data=data, timeout=120) as resp:
            if resp.status != 200:
                raise RuntimeError(await resp.text())


# ================= PROPERTIES =================

PROPERTIES_RAW = json.loads(os.getenv("OYO_PROPERTIES", "{}"))
PROPERTIES = {int(k): v for k, v in PROPERTIES_RAW.items()}


# ================= COLOR =================

def get_hour_color(day_index: int, total_days: int = 31):

    if total_days <= 1:
        total_days = 31

    t = day_index / (total_days - 1)

    colors = [
        (238,243,251),
        (217,242,255),
        (255,244,204),
        (255,221,128),
        (255,204,153),
        (255,193,193),
        (232,234,246),
        (227,242,253)
    ]

    seg = t * (len(colors) - 1)
    i = int(seg)
    frac = seg - i

    if i >= len(colors) - 1:
        r,g,b = colors[-1]
    else:
        r1,g1,b1 = colors[i]
        r2,g2,b2 = colors[i+1]

        r = int(r1 + (r2-r1)*frac)
        g = int(g1 + (g2-g1)*frac)
        b = int(b1 + (b2-b1)*frac)

    return f"{r:02X}{g:02X}{b:02X}"


# ================= FETCH DETAILS =================

async def fetch_booking_details(session, P, booking_no):

    url = "https://www.oyoos.com/hms_ms/api/v1/visibility/booking_details_with_entities"

    params = {
        "qid": P["QID"],
        "booking_id": booking_no,
        "role": 0,
        "platform": "OYOOS",
        "country_code": 1
    }

    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}

    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    for attempt in range(1, 4):

        try:
            async with session.get(
                url,
                params=params,
                headers=headers,
                cookies=cookies,
                timeout=DETAIL_TIMEOUT
            ) as r:

                if r.status != 200:
                    raise RuntimeError("DETAIL API FAILED")

                data = await r.json()

                booking = next(
                    iter(data.get("entities", {}).get("bookings", {}).values()),
                    {}
                )

                payments = booking.get("payments", [])

                cash = qr = online = discount = 0.0

                for p in payments:

                    amt = float(p.get("amount", 0) or 0)
                    mode = p.get("mode", "")

                    if mode == "oyo_wizard_discount":
                        discount += amt
                    elif mode == "Cash at Hotel":
                        cash += amt
                    elif mode == "UPI QR":
                        qr += amt
                    else:
                        online += amt

                # ✅ CORRECT BALANCE LOGIC
                paid = float(booking.get("get_amount_paid") or 0)
                balance = float(booking.get("payable_amount") or 0)

                return cash, qr, online, discount, balance, paid

        except Exception:
            await asyncio.sleep(2 + attempt)

    raise RuntimeError("DETAIL FETCH FAILED")


# ================= FETCH BOOKINGS =================

async def fetch_bookings_batch(session, offset, f, t, P):

    url = "https://www.oyoos.com/hms_ms/api/v1/get_booking_with_ids"

    params = {
        "qid": P["QID"],
        "checkin_from": f,
        "checkin_till": t,
        "batch_count": 100,
        "batch_offset": offset,
        "visibility_required": "true",
        "additionalParams": "payment_hold_transaction,guest,stay_details",
        "decimal_price": "true",
        "ascending": "true",
        "sort_on": "checkin_date"
    }

    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}

    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    async with session.get(
        url,
        params=params,
        cookies=cookies,
        headers=headers,
        timeout=BATCH_TIMEOUT
    ) as r:

        if r.status != 200:
            raise RuntimeError("BATCH API FAILED")

        return await r.json()


# ================= PROCESS PROPERTY =================

async def process_property(P, TF, TT, HF, HT):

    print(f"PROCESSING → {P['name']}")

    tf_dt = datetime.strptime(TF, "%Y-%m-%d").date()
    tt_dt = datetime.strptime(TT, "%Y-%m-%d").date()

    async with aiohttp.ClientSession() as session:

        detail_semaphore = asyncio.Semaphore(DETAIL_PARALLEL_LIMIT)

        async def limited_detail_call(booking_no):
            async with detail_semaphore:
                return await fetch_booking_details(session, P, booking_no)

        # DATE MAP
        date_map = {}
        d = tf_dt
        while d <= tt_dt:
            date_map[d] = {
                "cash": 0.0,
                "qr": 0.0,
                "online": 0.0,
                "discount": 0.0,
                "balance": 0.0,
                "total": 0.0
            }
            d += timedelta(days=1)

        offset = 0

        while True:

            data = await fetch_bookings_batch(session, offset, HF, HT, P)

            if not data or not data.get("bookingIds"):
                break

            bookings = data.get("entities", {}).get("bookings", {})
            if not bookings:
                break

            tasks = []
            mapping = []

            for b in bookings.values():

                status = (b.get("status") or "").strip()
                if status not in ["Checked In", "Checked Out"]:
                    continue

                booking_no = b.get("booking_no")
                if not booking_no:
                    continue

                ci = datetime.strptime(b["checkin"], "%Y-%m-%d").date()
                co = datetime.strptime(b["checkout"], "%Y-%m-%d").date()

                if co <= ci:
                    continue

                tasks.append(limited_detail_call(booking_no))
                mapping.append((ci, co))

            results = await asyncio.gather(*tasks, return_exceptions=True)

            for res, (ci, co) in zip(results, mapping):

                if isinstance(res, Exception):
                    continue

                cash, qr, online, discount, balance, paid = res

                stay_days = max((co - ci).days, 1)

                cash_d = cash / stay_days
                qr_d = qr / stay_days
                online_d = online / stay_days
                disc_d = discount / stay_days
                bal_d = balance / stay_days

                stay_day = ci

                while stay_day < co:

                    if tf_dt <= stay_day <= tt_dt:

                        row = date_map[stay_day]

                        row["cash"] += cash_d
                        row["qr"] += qr_d
                        row["online"] += online_d
                        row["discount"] += disc_d
                        row["balance"] += bal_d
                        row["total"] += (
                            cash_d + qr_d + online_d + disc_d + bal_d
                        )

                    stay_day += timedelta(days=1)

            if len(data.get("bookingIds", [])) < 100:
                break

            offset += 100

        return (P["name"], date_map)


# ================= RETRY =================

async def run_property_with_retry(P, TF, TT, HF, HT, retries=5):

    for attempt in range(1, retries + 1):

        try:
            return await process_property(P, TF, TT, HF, HT)

        except Exception as e:

            print(f"RETRY {attempt}/{retries} → {P['name']} :: {e}")
            await asyncio.sleep(2 + attempt * 2)

    raise RuntimeError(f"PROPERTY FAILED → {P['name']}")


async def run_property_limited(P, TF, TT, HF, HT):
    async with prop_semaphore:
        return await run_property_with_retry(P, TF, TT, HF, HT)


# ================= MAIN =================

async def main():

    target_date = (now - timedelta(days=1)).date()

    TF = target_date.replace(day=1).strftime("%Y-%m-%d")
    TT = target_date.strftime("%Y-%m-%d")

    HF = (target_date - timedelta(days=120)).strftime("%Y-%m-%d")
    HT = TT

    display_month = target_date.strftime("%B %Y")

    pending = {k: v for k, v in PROPERTIES.items()}
    success_results = {}

    for run_attempt in range(1, MAX_FULL_RUN_RETRIES + 1):

        if not pending:
            break

        tasks = [run_property_limited(P, TF, TT, HF, HT) for P in pending.values()]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        new_pending = {}

        for key, (P, result) in zip(list(pending.keys()), zip(pending.values(), results)):

            if isinstance(result, Exception):
                new_pending[key] = P
                continue

            success_results[key] = result

        pending = new_pending

        if pending:

            if run_attempt == MAX_FULL_RUN_RETRIES:
                raise RuntimeError("FINAL FAILURE")

            await asyncio.sleep(FULL_RUN_RETRY_DELAY)

    valid_results = [success_results[k] for k in PROPERTIES.keys() if k in success_results]

    # ================= EXCEL =================

    wb = Workbook()
    wb.remove(wb.active)

    start_dt = datetime.strptime(TF, "%Y-%m-%d").date()
    end_dt = datetime.strptime(TT, "%Y-%m-%d").date()

    date_list = []
    d = start_dt
    while d <= end_dt:
        date_list.append(d)
        d += timedelta(days=1)

    consolidated = {
        d: {"cash":0,"qr":0,"online":0,"discount":0,"balance":0,"total":0}
        for d in date_list
    }

    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    def create_sheet(ws, date_map):

        ws.append(["Date","Cash","QR","Online","Discount","Balance","Total"])

        header_fill = PatternFill("solid", fgColor="1F4E78")

        for col in range(1,8):
            c = ws.cell(row=1,column=col)
            c.fill = header_fill
            c.font = Font(bold=True,color="FFFFFF")
            c.alignment = Alignment(horizontal="center")

        # Column widths
        widths = [16,14,14,14,14,14,18]
        for i,w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64+i)].width = w

        sum_cash=sum_qr=sum_online=sum_discount=sum_balance=sum_total=0

        for idx,d in enumerate(date_list):

            row = date_map[d]

            cash=round(row["cash"],2)
            qr=round(row["qr"],2)
            online=round(row["online"],2)
            discount=round(row["discount"],2)
            balance=round(row["balance"],2)
            total=round(row["total"],2)

            sum_cash+=cash
            sum_qr+=qr
            sum_online+=online
            sum_discount+=discount
            sum_balance+=balance
            sum_total+=total

            ws.append([
                d.strftime("%d-%m-%Y"),
                cash,qr,online,discount,balance,total
            ])

            r=ws.max_row
            fill=PatternFill("solid", fgColor=get_hour_color(idx,len(date_list)))

            for c in range(1,8):
                cell=ws.cell(row=r,column=c)
                cell.fill=fill
                cell.border=thin
                cell.alignment=Alignment(horizontal="center")

        ws.append([
            "TOTAL",
            round(sum_cash,2),
            round(sum_qr,2),
            round(sum_online,2),
            round(sum_discount,2),
            round(sum_balance,2),
            round(sum_total,2)
        ])

        total_row=ws.max_row

        for c in range(1,8):
            cell=ws.cell(row=total_row,column=c)
            cell.fill=PatternFill("solid", fgColor="000000")
            cell.font=Font(bold=True,color="FFFFFF")
            cell.alignment=Alignment(horizontal="center")


        # ===== CHARTS =====

        chart_titles = ["Cash","QR","Online","Discount","Balance","Total"]

        base_row = ws.max_row + 3
        gap = 22

        for i, col in enumerate(range(2, 8)):

            chart = BarChart()
            chart.title = f"{chart_titles[i]} Trend"
            chart.height = 12
            chart.width = 26
            chart.legend = None

            data = Reference(ws, min_col=col, min_row=1, max_row=len(date_list)+1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(date_list)+1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            series = chart.series[0]
            pts = []

            for idx in range(len(date_list)):
                dp = DataPoint(idx=idx)
                dp.graphicalProperties.solidFill = get_hour_color(idx, len(date_list))
                pts.append(dp)

            series.dPt = pts

            ws.add_chart(chart, f"A{base_row + i*gap}")




    ranking=[]

    for name,date_map in valid_results:

        ws=wb.create_sheet(name[:31])
        create_sheet(ws,date_map)

        total_prop=0

        for d in date_list:
            for k in consolidated[d]:
                consolidated[d][k]+=date_map[d][k]

            total_prop+=date_map[d]["total"]

        ranking.append({"name":name,"total":total_prop})

    ws=wb.create_sheet("CONSOLIDATED")
    create_sheet(ws,consolidated)

    # ================= RANKING =================

    ws=wb.create_sheet("PROPERTY RANKING")

    headers=["Rank","Property","Total","Badge"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")

    for col in range(1,5):
        c=ws.cell(row=1,column=col)
        c.fill=header_fill
        c.font=Font(bold=True,color="FFFFFF")
        c.alignment=Alignment(horizontal="center")

    widths=[10,28,16,18]
    for i,w in enumerate(widths,start=1):
        ws.column_dimensions[chr(64+i)].width=w

    ranking.sort(key=lambda x:x["total"], reverse=True)

    def medal(r):
        if r==1: return "🥇 Gold"
        if r==2: return "🥈 Silver"
        if r==3: return "🥉 Bronze"
        return ""

    rnk=1
    for idx,p in enumerate(ranking):

        ws.append([rnk,p["name"],round(p["total"],2),medal(rnk)])

        r=ws.max_row
        fill=PatternFill("solid", fgColor=get_hour_color(idx,len(ranking)))

        for c in range(1,5):
            cell=ws.cell(row=r,column=c)
            cell.fill=fill
            cell.border=thin
            cell.alignment=Alignment(horizontal="center")

        rnk+=1

    buffer=BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    await send_telegram_excel_buffer(
        buffer,
        filename=f"Revenue_Monthly.xlsx",
        caption="📊 Per-Day Stay Collection Report"
    )


if __name__ == "__main__":

    try:
        asyncio.run(main())
    except Exception:
        traceback.print_exc()
