# ==============================
# ULTRA FAST ASYNC MULTI PROPERTY AUTOMATION
# MILLION BOOKING READY
# BEAUTIFUL PREMIUM EXCEL
# ==============================

import os
import asyncio
import aiohttp
import pandas as pd
from datetime import datetime, timedelta
import traceback
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import pytz
IST = pytz.timezone("Asia/Kolkata")

now = datetime.now(IST)

MAX_FULL_RUN_RETRIES = 5
FULL_RUN_RETRY_DELAY = 10  # seconds

# ================= NEW: GLOBAL THROTTLES (NO FEATURE REMOVED) =================
PROP_PARALLEL_LIMIT = 3      # max properties running in parallel
DETAIL_PARALLEL_LIMIT = 10   # max detail calls in parallel per property

prop_semaphore = asyncio.Semaphore(PROP_PARALLEL_LIMIT)

# ================= NEW: TIMEOUTS (MORE STABLE MONTHLY) =================
DETAIL_TIMEOUT = 25
ROOMS_TIMEOUT = 25
BATCH_TIMEOUT = 35


# ================= TELEGRAM =================

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")

if not TELEGRAM_BOT_TOKEN:
    raise RuntimeError("❌ TELEGRAM_BOT_TOKEN missing")


CHAT_MAP = json.loads(os.getenv("TELEGRAM_CHAT_MAP", "{}"))

def get_chat_id(name: str):
    if name not in CHAT_MAP:
        raise RuntimeError(f"❌ Chat ID not configured: {name}")
    return int(CHAT_MAP[name])


# choose chat key from secret map
TELEGRAM_CHAT_ID = get_chat_id("hourly")   # change if needed


# ================= PROPERTIES =================

PROPERTIES_RAW = json.loads(os.getenv("OYO_PROPERTIES", "{}"))

PROPERTIES = {int(k): v for k, v in PROPERTIES_RAW.items()}

if not PROPERTIES:
    raise RuntimeError("❌ OYO_PROPERTIES secret missing or empty")

# ================= TELEGRAM =================
# ✅ ULTRA SAFE: Guaranteed ordered sending (one-by-one)
# ✅ Handles 429 retry_after properly
# ✅ Strictly verifies JSON "ok": true
# ✅ Auto-split long messages
# ✅ Exponential backoff retries
# ✅ Never loses a property message unless Telegram is fully down

TELEGRAM_SEND_LOCK = asyncio.Lock()

async def send_telegram_message(text, retries=15, session=None):
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"

    def split_message(msg, limit=3900):
        msg = str(msg or "")
        if len(msg) <= limit:
            return [msg]

        parts = []
        while len(msg) > limit:
            cut = msg.rfind("\n", 0, limit)
            if cut == -1 or cut < 1000:
                cut = limit
            parts.append(msg[:cut].strip())
            msg = msg[cut:].strip()
        if msg:
            parts.append(msg)
        return parts

    async def _post(sess, msg_part):
        payload = {"chat_id": TELEGRAM_CHAT_ID, "text": msg_part, "parse_mode": "HTML"}

        async with sess.post(url, json=payload, timeout=25) as resp:
            # ✅ 429 flood control
            if resp.status == 429:
                retry_after = 5
                try:
                    data = await resp.json()
                    retry_after = int(data.get("parameters", {}).get("retry_after", 5))
                except Exception:
                    retry_after = 5

                print(f"⚠️ TELEGRAM 429 RATE LIMIT → sleeping {retry_after}s")
                await asyncio.sleep(retry_after + 1)
                return False

            # ✅ Non-200
            if resp.status != 200:
                try:
                    err = await resp.text()
                except Exception:
                    err = ""
                raise RuntimeError(f"Telegram HTTP {resp.status} {err}")

            # ✅ Very important: Telegram sometimes returns 200 but ok:false
            try:
                data = await resp.json()
            except Exception:
                data = {}

            if data.get("ok") is True:
                return True

            # ok:false -> treat as failure and retry
            desc = data.get("description", "Unknown Telegram error")
            raise RuntimeError(f"Telegram API ok:false → {desc}")

    parts = split_message(text)

    # ✅ Guarantee ONE-BY-ONE send (no burst)
    async with TELEGRAM_SEND_LOCK:

        # If session not provided (still works as before)
        if session is None:
            async with aiohttp.ClientSession() as s:
                for part in parts:
                    last_err = None
                    for attempt in range(1, retries + 1):
                        try:
                            ok = await _post(s, part)
                            if ok:
                                await asyncio.sleep(0.3)  # small safe gap
                                break
                        except Exception as e:
                            last_err = e
                            wait = min(60, 2 * attempt)  # backoff
                            print(f"⚠️ Telegram retry {attempt}/{retries} → {wait}s :: {e}")
                            await asyncio.sleep(wait)
                    else:
                        raise RuntimeError(f"Telegram send failed after retries: {last_err}")
            return

        # ✅ Session provided (fast + stable)
        for part in parts:
            last_err = None
            for attempt in range(1, retries + 1):
                try:
                    ok = await _post(session, part)
                    if ok:
                        await asyncio.sleep(0.3)  # small safe gap
                        break
                except Exception as e:
                    last_err = e
                    wait = min(60, 2 * attempt)  # backoff
                    print(f"⚠️ Telegram retry {attempt}/{retries} → {wait}s :: {e}")
                    await asyncio.sleep(wait)
            else:
                raise RuntimeError(f"Telegram send failed after retries: {last_err}")

# ================= BEAUTIFY EXCEL =================
def beautify(ws):
    blue = PatternFill("solid", fgColor="1F4E78")
    light1 = PatternFill("solid", fgColor="DDEBF7")
    light2 = PatternFill("solid", fgColor="F2F2F2")
    yellow = PatternFill("solid", fgColor="FFF4CC")

    bold_white = Font(color="FFFFFF", bold=True, size=12)
    bold_black = Font(color="000000", bold=True, size=12)

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    max_row = ws.max_row
    max_col = ws.max_column
    ws.freeze_panes = "A2"

    for col in range(1, max_col + 1):
        c = ws.cell(row=1, column=col)
        c.fill = blue
        c.font = bold_white
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin

    for r in range(2, max_row + 1):
        fill = light1 if r % 2 == 0 else light2
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            cell.fill = fill
            cell.border = thin

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 5

    for r in range(2, max_row + 1):
        text = str(ws.cell(row=r, column=1).value or "")
        if text.strip() == "":
            continue
        if "Booking" in text or "Amount" in text or "Total" in text or "OYO" in text:
            ws.cell(row=r, column=1).fill = yellow
            ws.cell(row=r, column=1).font = bold_black

# ================= BOOKING SOURCE =================
def get_booking_source(b):
    source = str(b.get("source", "") or "").strip()
    ota = str(b.get("ota_source", "") or "").strip()
    sub = str(b.get("sub_source", "") or "").strip()
    corp = bool(b.get("is_corporate", False))

    booking_identifier = str(b.get("booking_identifier", "") or "").strip()

    # ✅ strongest indicator first
    if booking_identifier == "TA":
        return "TA"

    # ✅ walk-in
    if source == "Walk In":
        return "Walk-in"

    # ✅ corporate
    if corp or sub == "corporate":
        return "CB"

    # ✅ OTA
    if "Booking.com" in ota:
        return "BDC"
    if "GoMMT" in ota:
        return "MMT"
    if "Agoda" in ota:
        return "Agoda"

    # ✅ travel agent / TPO (THIS fixes your sub_source issue)
    if source == "Travel Agent" or sub == "TPO":
        return "TA"

    # ✅ OYO direct
    if source in [
        "Android App",
        "IOS App",
        "Web Booking",
        "Mobile Web Booking",
        "Website Booking",
        "Direct"
    ]:
        return "OYO"

    # ✅ fallback
    return "OBA"


# ================= FETCH DETAILS =================
async def fetch_booking_details(session, P, booking_no):
    url = "https://www.oyoos.com/hms_ms/api/v1/visibility/booking_details_with_entities"
    params = {
        "qid": P["QID"],
        "booking_id": booking_no,
        "role": 0,
        "platform": "OYOOS",
        "country_code": 1
    }
    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}
    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    for attempt in range(1, 4):
        try:
            async with session.get(
                url, params=params, headers=headers, cookies=cookies, timeout=DETAIL_TIMEOUT
            ) as r:
                if r.status != 200:
                    raise RuntimeError("DETAIL API FAILED")

                data = await r.json()

                rooms = []
                stay = data.get("entities", {}).get("stayDetails", {})
                for s in stay.values():
                    rn = s.get("room_number")
                    if rn:
                        rooms.append(rn)

                booking = next(iter(data.get("entities", {}).get("bookings", {}).values()), {})
                payments = booking.get("payments", [])

                cash = qr = online = discount = 0
                for p in payments:
                    mode = p.get("mode", "")
                    amt = float(p.get("amount", 0))
                    if mode == "oyo_wizard_discount":
                        discount += amt
                    elif mode == "Cash at Hotel":
                        cash += amt
                    elif mode == "UPI QR":
                        qr += amt
                    else:
                        online += amt

                balance = booking.get("payable_amount", 0)
                return rooms, cash, qr, online, discount, balance

        except Exception:
            # NEW FEATURE: Better backoff (reduces rate limits)
            await asyncio.sleep(2 + attempt)

    raise RuntimeError("DETAIL FETCH FAILED")

# ================= BATCH FETCH =================
async def fetch_bookings_batch(session, offset, f, t, P):
    url = "https://www.oyoos.com/hms_ms/api/v1/get_booking_with_ids"
    params = {
        "qid": P["QID"],
        "checkin_from": f,
        "checkin_till": t,
        "batch_count": 100,
        "batch_offset": offset,
        "visibility_required": "true",
        "additionalParams": "payment_hold_transaction,guest,stay_details",
        "decimal_price": "true",
        "ascending": "true",
        "sort_on": "checkin_date"
    }
    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}
    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    async with session.get(
        url, params=params, cookies=cookies, headers=headers, timeout=BATCH_TIMEOUT
    ) as r:
        if r.status != 200:
            raise RuntimeError("BATCH API FAILED")
        return await r.json()

# ================= FETCH TOTAL ROOMS =================
async def fetch_total_rooms(session, P):
    url = "https://www.oyoos.com/hms_ms/api/v1/hotels/roomsNew"
    params = {"qid": P["QID"]}
    cookies = {"uif": P["UIF"], "uuid": P["UUID"]}
    headers = {
        "accept": "application/json",
        "x-qid": str(P["QID"]),
        "x-source-client": "merchant"
    }

    for attempt in range(1, 4):
        try:
            async with session.get(
                url, params=params, cookies=cookies, headers=headers, timeout=ROOMS_TIMEOUT
            ) as r:
                if r.status != 200:
                    raise RuntimeError("ROOM API FAILED")

                data = await r.json()
                rooms = data.get("rooms", {})
                return len(rooms)

        except Exception:
            await asyncio.sleep(2 + attempt)

    return 0

# ================= PROCESS PROPERTY =================
async def process_property(P, TF, TT, HF, HT):
    print(f"PROCESSING FAST ASYNC → {P['name']}")

    async with aiohttp.ClientSession() as session:
        total_rooms = await fetch_total_rooms(session, P)

        if total_rooms == 0:
            raise RuntimeError("TOTAL ROOMS FETCH FAILED")

        detail_semaphore = asyncio.Semaphore(DETAIL_PARALLEL_LIMIT)

        async def limited_detail_call(booking_no):
            async with detail_semaphore:
                return await fetch_booking_details(session, P, booking_no)

        all_rows = []
        offset = 0
        upcoming_count = cancelled_count = inhouse_count = checkedout_count = 0

        early_checkins = set()
        late_checkouts = set()

        while True:
            data = await fetch_bookings_batch(session, offset, HF, HT, P)

            if not data or not data.get("bookingIds"):
                break

            bookings = data.get("entities", {}).get("bookings", {})

            if not bookings:
                raise RuntimeError("BOOKING ENTITY EMPTY")

            curr = datetime.strptime(TF, "%Y-%m-%d")
            end = datetime.strptime(TT, "%Y-%m-%d")

            while curr <= end:
                target = curr.strftime("%Y-%m-%d")
                target_dt = curr

                tasks, mapping = [], []

                for b in bookings.values():
                    status = (b.get("status") or "").strip()
                    ci = datetime.strptime(b["checkin"], "%Y-%m-%d")
                    co = datetime.strptime(b["checkout"], "%Y-%m-%d")
                    tf_date = datetime.strptime(TF, "%Y-%m-%d")

                    # ---- STATUS COUNTS ----
                    if status == "Checked In":
                        if ci <= tf_date or ci == tf_date + timedelta(days=1):
                            inhouse_count += 1

                    elif status == "Checked Out":
                        today = now.date()
                        if co.date() == today:
                            checkedout_count += 1

                    elif status == "Confirm Booking":
                        today = now.date()
                        if ci.date() == today:
                            upcoming_count += 1

                    elif status == "Cancelled Booking":
                        if ci == tf_date or ci == tf_date + timedelta(days=1):
                            cancelled_count += 1

                    # ---- EARLY / LATE DETECTION ----
                    if status == "Checked In":
                        if ci.date() > tf_date.date():
                            early_checkins.add(str(b.get("booking_no", "")).strip())

                        if co.date() == tf_date.date():
                            late_checkouts.add(str(b.get("booking_no", "")).strip())

                    # ---------- ROW FILTER ----------
                    if status != "Checked In":
                        continue

                    if not (ci <= target_dt < co or (ci == tf_date + timedelta(days=1) and target_dt < co)):
                        continue

                    tasks.append(limited_detail_call(b["booking_no"]))
                    mapping.append((b, target, ci, co))

                if tasks:
                    results = await asyncio.gather(*tasks, return_exceptions=True)
                else:
                    results = []

                for res, (b, target, ci, co) in zip(results, mapping):
                    if isinstance(res, Exception):
                        continue

                    rooms, cash, qr, online, discount, balance = res

                    stay = max((co - ci).days, 1)
                    paid = float(b.get("get_amount_paid") or 0)
                    total_amt = paid + float(balance or 0)

                    all_rows.append({
                        "Date": target,
                        "Booking Id": b["booking_no"],
                        "Guest Name": b["guest_name"],
                        "Status": b.get("status"),
                        "Booking Source": get_booking_source(b),
                        "Check In": b["checkin"],
                        "Check Out": b["checkout"],
                        "Rooms": b.get("no_of_rooms", 1),
                        "Room Numbers": ", ".join(rooms),
                        "Amount": round(total_amt / stay, 2),
                        "Cash": round(cash / stay, 2),
                        "QR": round(qr / stay, 2),
                        "Online": round(online / stay, 2),
                        "Discount": round(discount / stay, 2),
                        "Balance": round(balance / stay, 2),
                    })

                curr += timedelta(days=1)

            if len(data["bookingIds"]) < 100:
                break

            offset += 100

        df = pd.DataFrame(all_rows)

        if df.empty:
            print(f"⚠️ NO ROWS → {P['name']}")
            df = pd.DataFrame(columns=[
                "Date", "Booking Id", "Guest Name", "Status", "Booking Source",
                "Check In", "Check Out", "Rooms", "Room Numbers",
                "Amount", "Cash", "QR", "Online", "Discount", "Balance"
            ])

        df.columns = [str(c).strip() for c in df.columns]

        return (
            P["name"],
            df,
            total_rooms,
            inhouse_count,
            checkedout_count,
            upcoming_count,
            cancelled_count,
            sorted(list(early_checkins)),
            sorted(list(late_checkouts))
        )


# ================= RELIABILITY WRAPPER =================
async def run_property_with_retry(P, TF, TT, HF, HT, retries=3):
    last_error = None
    for attempt in range(1, retries + 1):
        try:
            return await process_property(P, TF, TT, HF, HT)
        except Exception as e:
            last_error = e
            print(f"RETRY {attempt}/{retries} → {P['name']} :: {e}")
            await asyncio.sleep(2 + attempt * 2)
    raise RuntimeError(f"PROPERTY FAILED → {P['name']}") from last_error

# NEW FEATURE: PROPERTY PARALLEL LIMITER (reduces failures)
async def run_property_limited(P, TF, TT, HF, HT):
    async with prop_semaphore:
        return await run_property_with_retry(P, TF, TT, HF, HT)





def build_early_late_alert_message(prop, report_date, early_list, late_list):
    """
    Telegram HTML does not support font colors.
    So we use 🟢 and 🔴 icons for clear green/red indication.
    """
    early_count = len(early_list)
    late_count = len(late_list)

    if early_count == 0 and late_count == 0:
        return None

    lines = []
    lines.append("<b>IN-HOUSE EXCEPTION ALERT</b>")
    lines.append(f"<b>🏢 Property :</b> {prop}")
    lines.append(f"<b>📅 Date     :</b> {report_date}")
    lines.append("")

    if early_count > 0:
        lines.append(f"🟢 <b>EARLY CHECK-IN FOUND :</b> <b>{early_count}</b>")
        # show max 10 booking ids for neatness
        show = early_list[:10]
        lines.append(f"<b>Booking IDs:</b> {', '.join(show)}" + (" ..." if early_count > 10 else ""))
        lines.append("")

    if late_count > 0:
        lines.append(f"🔴 <b>LATE CHECK-OUT FOUND :</b> <b>{late_count}</b>")
        show = late_list[:10]
        lines.append(f"<b>Booking IDs:</b> {', '.join(show)}" + (" ..." if late_count > 10 else ""))
        lines.append("")

    lines.append("<b>Action Required:</b> Please verify front-office status immediately.")

    return "\n".join(lines).strip()

# ================= COUNT / AMOUNT =================
def count(df, src):
    if df.empty:
        return 0
    return int(df[df["Booking Source"] == src]["Rooms"].sum())

def amt(df, src):
    if df.empty:
        return 0
    return round(df[df["Booking Source"] == src]["Amount"].sum(), 2)

def count_upcoming(df, tf):
    tf_date = datetime.strptime(tf, "%Y-%m-%d")
    next_date = tf_date + timedelta(days=1)

    c = 0
    for _, r in df.iterrows():
        if r.get("Status") != "Confirm Booking":
            continue

        d = datetime.strptime(r["Date"], "%Y-%m-%d")

        if d == tf_date and now.hour >= 12:
            c += 1
        elif d == next_date and now.hour < 12:
            c += 1

    return c

def count_cancelled(df, tf):
    tf_date = datetime.strptime(tf, "%Y-%m-%d")
    next_date = tf_date + timedelta(days=1)

    c = 0
    for _, r in df.iterrows():
        if r.get("Status") != "Cancelled Booking":
            continue

        d = datetime.strptime(r["Date"], "%Y-%m-%d")

        if d == tf_date or d == next_date:
            c += 1

    return c

def build_telegram_message(
    prop,
    total_rooms,
    booked_rooms,
    available_rooms,
    occupancy,
    inhouse,
    checkedout,
    upcoming,
    cancelled,
    counts,
    amounts,
    arr,
    app_arr
):
    now = datetime.now(IST)

    return f"""
<pre>
HOURLY REPORT : {prop}

🏢 Property Code     : {prop}
📅 Date              : {now.strftime("%d/%m/%Y")}
⏰ Time              : {now.strftime("%I:%M %p")}
📆 Day               : {now.strftime("%A")}

🔹 URN In-House      : {booked_rooms}
🔹 Checked Out       : {checkedout:02d}
🔹 Upcoming Bookings : {upcoming:02d}
🔹 Cancelled         : {cancelled:02d}

🔹 Total Rooms       : {total_rooms}
🔹 Booked Rooms      : {booked_rooms}
🔹 Available Rooms   : {available_rooms}
🔹 Occupancy         : {occupancy}%

🔹 Walk-in           : {counts['Walk-in']:02d}
🔹 OYO               : {counts['OYO']:02d}
🔹 MMT               : {counts['MMT']:02d}
🔹 Agoda             : {counts['Agoda']:02d}
🔹 CB                : {counts['CB']:02d}
🔹 BDC               : {counts['BDC']:02d}
🔹 TA                : {counts['TA']:02d}
🔹 OBA               : {counts['OBA']:02d}

🔹 Total Amount      : ₹{amounts['Total']:,}
🔹 Cash              : ₹{amounts['Cash']:,}
🔹 QR                : ₹{amounts['QR']:,}
🔹 Online            : ₹{amounts['Online']:,}
🔹 Discount          : ₹{amounts['Discount']:,}
🔹 Balance           : ₹{amounts['Balance']:,}

🔹 ARR               : ₹{arr}
🔹 App ARR           : ₹{app_arr}

</pre>
""".strip()


# ================= MAIN =================
# ================= MAIN =================
async def main():
    print("========================================")
    print(" OYO DAILY TELEGRAM AUTOMATION")
    print("========================================")

    now = datetime.now(IST)

    # ================= BUSINESS DATE CUTOVER (12 PM RULE) =================
    if now.hour < 12:
        target_date = (now - timedelta(days=1)).date()
    else:
        target_date = now.date()

    # ================= PREVIOUS MONTH (BASED ON TARGET_DATE) =================
    TF = target_date.strftime("%Y-%m-%d")
    TT = TF
    HF = (target_date - timedelta(days=30)).strftime("%Y-%m-%d")
    HT = now.strftime("%Y-%m-%d")

    # ================= SMART RETRY (ONLY FAILED PROPERTIES) =================
    pending = {k: v for k, v in PROPERTIES.items()}
    success_results = {}

    for run_attempt in range(1, MAX_FULL_RUN_RETRIES + 1):
        if not pending:
            break

        print(f"\n🔁 PARTIAL RUN ATTEMPT {run_attempt}/{MAX_FULL_RUN_RETRIES}")
        print(f"⏳ Pending Properties: {len(pending)}")

        tasks = [run_property_limited(P, TF, TT, HF, HT) for P in pending.values()]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        new_pending = {}
        for key, (P, result) in zip(list(pending.keys()), zip(pending.values(), results)):

            if isinstance(result, Exception):
                print(f"❌ FAILED → {P['name']} :: {result}")
                new_pending[key] = P
                continue

            name, df, *_ = result

            if df is None:
                print(f"❌ EMPTY DATA → {name}")
                new_pending[key] = P
                continue

            success_results[key] = result
            print(f"✅ OK → {name} rows={len(df)}")

        pending = new_pending

        if pending:
            if run_attempt == MAX_FULL_RUN_RETRIES:
                failed_names = [p["name"] for p in pending.values()]
                raise RuntimeError(f"FINAL FAILURE: Properties failed after retries: {failed_names}")

            print(f"🔁 RETRYING ONLY FAILED PROPERTIES after {FULL_RUN_RETRY_DELAY}s...")
            await asyncio.sleep(FULL_RUN_RETRY_DELAY)

    # ================= FINAL VERIFICATION =================
    valid_results = [success_results[k] for k in PROPERTIES.keys() if k in success_results]

    if len(valid_results) != len(PROPERTIES):
        missing = [PROPERTIES[k]["name"] for k in PROPERTIES.keys() if k not in success_results]
        raise RuntimeError(f"DATA INCOMPLETE: Missing properties: {missing}")

    print("✅ DATA VERIFIED — ALL PROPERTIES PRESENT")

    async with aiohttp.ClientSession() as tg_session:

        # ================= PER-PROPERTY REPORTS =================
        for name, df, total_rooms, inhouse, checkedout, upcoming, cancelled, early_checkins, late_checkouts in valid_results:

            booked_rooms = int(df["Rooms"].sum()) if not df.empty else 0

            booked_rooms = int(df["Rooms"].sum())
            available_rooms = total_rooms - booked_rooms
            occupancy = round((booked_rooms / total_rooms) * 100) if total_rooms else 0

            total_amount = float(df["Amount"].sum()) if not df.empty else 0.0
            arr = round(total_amount / booked_rooms, 2) if booked_rooms else 0

            oyo_df = df[df["Booking Source"] == "OYO"] if not df.empty else df
            oyo_rooms = int(oyo_df["Rooms"].sum()) if not oyo_df.empty else 0
            oyo_amount = float(oyo_df["Amount"].sum()) if not oyo_df.empty else 0.0
            app_arr = round(oyo_amount / oyo_rooms, 2) if oyo_rooms else 0

            counts = {
                "Walk-in": count(df, "Walk-in"),
                "OYO": count(df, "OYO"),
                "MMT": count(df, "MMT"),
                "Agoda": count(df, "Agoda"),
                "CB": count(df, "CB"),
                "BDC": count(df, "BDC"),
                "TA": count(df, "TA"),
                "OBA": count(df, "OBA")
            }

            amounts = {
                "Total": int(total_amount),
                "Cash": int(df["Cash"].sum()) if not df.empty else 0,
                "QR": int(df["QR"].sum()) if not df.empty else 0,
                "Online": int(df["Online"].sum()) if not df.empty else 0,
                "Discount": int(df["Discount"].sum()) if not df.empty else 0,
                "Balance": int(df["Balance"].sum()) if not df.empty else 0
            }

            hourly_message = build_telegram_message(
                prop=name,
                total_rooms=total_rooms,
                booked_rooms=booked_rooms,
                available_rooms=available_rooms,
                occupancy=occupancy,
                inhouse=inhouse,
                checkedout=checkedout,
                upcoming=upcoming,
                cancelled=cancelled,
                counts=counts,
                amounts=amounts,
                arr=arr,
                app_arr=app_arr
            )

            await send_telegram_message(hourly_message, session=tg_session)
            await asyncio.sleep(1.5)

            # ================= EARLY / LATE ALERT =================
            if (early_checkins and len(early_checkins) > 0) or (late_checkouts and len(late_checkouts) > 0):

                alert_msg = build_early_late_alert_message(
                    prop=name,
                    report_date=datetime.strptime(TF, "%Y-%m-%d").strftime("%d/%m/%Y"),
                    early_list=early_checkins or [],
                    late_list=late_checkouts or []
                )

                if alert_msg:
                    try:
                        await send_telegram_message(alert_msg, session=tg_session)
                        await asyncio.sleep(1.5)
                    except Exception as e:
                        print(f"⚠️ ALERT TELEGRAM FAILED → {name} :: {e}")

        # ================= CONSOLIDATED REPORT =================
        all_df = pd.concat([r[1] for r in valid_results], ignore_index=True)

        total_rooms_all = sum(r[2] for r in valid_results)
        inhouse_all = sum(r[3] for r in valid_results)
        checkedout_all = sum(r[4] for r in valid_results)
        upcoming_all = sum(r[5] for r in valid_results)
        cancelled_all = sum(r[6] for r in valid_results)

        booked_rooms_all = int(all_df["Rooms"].sum()) if not all_df.empty else 0

        booked_rooms_all = int(all_df["Rooms"].sum())
        available_rooms_all = total_rooms_all - booked_rooms_all
        occupancy_all = round((booked_rooms_all / total_rooms_all) * 100) if total_rooms_all else 0

        total_amount_all = float(all_df["Amount"].sum()) if not all_df.empty else 0.0
        arr_all = round(total_amount_all / booked_rooms_all, 2) if booked_rooms_all else 0

        oyo_df_all = all_df[all_df["Booking Source"] == "OYO"] if not all_df.empty else all_df
        oyo_rooms_all = int(oyo_df_all["Rooms"].sum()) if not oyo_df_all.empty else 0
        oyo_amount_all = float(oyo_df_all["Amount"].sum()) if not oyo_df_all.empty else 0.0
        app_arr_all = round(oyo_amount_all / oyo_rooms_all, 2) if oyo_rooms_all else 0

        counts_all = {k: count(all_df, k) for k in ["Walk-in","OYO","MMT","Agoda","CB","BDC","TA","OBA"]}

        amounts_all = {
            "Total": int(total_amount_all),
            "Cash": int(all_df["Cash"].sum()) if not all_df.empty else 0,
            "QR": int(all_df["QR"].sum()) if not all_df.empty else 0,
            "Online": int(all_df["Online"].sum()) if not all_df.empty else 0,
            "Discount": int(all_df["Discount"].sum()) if not all_df.empty else 0,
            "Balance": int(all_df["Balance"].sum()) if not all_df.empty else 0
        }

        consolidated_hourly = build_telegram_message(
            prop="ALL",
            total_rooms=total_rooms_all,
            booked_rooms=booked_rooms_all,
            available_rooms=available_rooms_all,
            occupancy=occupancy_all,
            inhouse=inhouse_all,
            checkedout=checkedout_all,
            upcoming=upcoming_all,
            cancelled=cancelled_all,
            counts=counts_all,
            amounts=amounts_all,
            arr=arr_all,
            app_arr=app_arr_all
        )

        await send_telegram_message(consolidated_hourly, session=tg_session)

    print("✅ ALL MONTHLY TELEGRAM REPORTS SENT — GUARANTEED")

# ================= RUN =================
if __name__ == "__main__":
    try:
        asyncio.run(main())
    except Exception as e:
        print("SCRIPT CRASHED")
        print(e)
        traceback.print_exc()
        print("SCRIPT CRASHED", e, flush=True)










